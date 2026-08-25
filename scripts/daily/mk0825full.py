# -*- coding: utf-8 -*-
"""2026-08-25 定例レポート本体（昨日=8/24 月）＋ 当日8/25 14:47時点の診断シート。"""
import pickle, csv, datetime, collections, statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
R = pickle.load(open('/tmp/rep0824w.pkl', 'rb')); FEE = 0.03452
D1, D3, D7, D30, CUM = R['D1'], R['D3'], R['D7'], R['D30'], R['CUM']
SNAP = '2026-08-25'
rows = list(csv.reader(open('data/budget-snapshots.csv'))); body = [r for r in rows[1:] if r]
assert any(r[0] == SNAP for r in body), f'{SNAP} スナップショット未記録'
BUD = collections.defaultdict(int)
for r in body:
    if r[0] == SNAP: BUD[r[1]] += int(r[3])
assert sum(BUD.values()) == 258000, sum(BUD.values())
BSNAP = collections.defaultdict(lambda: collections.defaultdict(int))
for r in body: BSNAP[r[1]][r[0].rstrip('b')] += int(r[3])

MAP = R['MAP']; INV = {v: k for k, v in MAP.items()}
# --- 7日(8/18-24) Shopify注文数（ShopifyQL GROUP BY product_title 実測）---
ORD = {'ナノガラス脱毛パッド':260,'W固定スマホ車載ホルダー':115,'快適マジックインソール':69,'形状記憶日傘':47,
'ムダ毛シェーバー':36,'2WAYシートボックス':34,'偏光・調光サングラス':34,'完全遮光・接触冷感UVハット':30,
'姿勢サポートチェア':24,'4WAY':16,'バランスケアスリッパ':15,'3D足臭リセットブラシ':13,'優先配送':12,
'接触冷感UVアームカバー':10,'完全遮光・形状記憶':10,'むくみ取りかっさ':8,'接触冷感UVパーカー':7,
'携帯電動シェーバー':4,'健康サンダル':1,'癒しの指圧マット':1,'5WAY腰掛けファン':1}
STORE_ORD7, STORE_SESS7 = 722, 23039
assert sum(ORD.values()) == 747, sum(ORD.values())
# --- 7日(8/18-24) Metaファネル（キャンペーン別実測: cost, impressions, link_clicks）---
MF_RAW = {'ナノガラス脱毛パッド':(559791,282813,6536),'W固定スマホ車載ホルダー':(245253,119951,3378),
'形状記憶日傘':(200878,46213,1874),'ムダ毛シェーバー':(137388,103294,1315),'快適マジックインソール':(123811,73356,1777),
'偏光・調光サングラス':(101086,41425,1143),'2WAYシートボックス':(96954,27797,844),'姿勢サポートチェア':(77916,46960,922),
'完全遮光・接触冷感UVハット':(70578,14560,696),'4WAY 取り付けOK 小型瞬間冷却ハンディファン':(47379,12126,325),
'3D足臭リセットブラシ':(43334,9413,173),'接触冷感UVアームカバー':(39723,8912,304),'バランスケアスリッパ':(38200,23779,454),
'接触冷感UVパーカー':(19070,5260,135),'むくみ取りかっさ':(12921,3030,100),'1秒折り畳みチェア':(5704,1773,50),
'完全遮光・形状記憶・晴雨兼用・UV日傘':(2641,427,22)}
MF = {INV.get(k, k): v for k, v in MF_RAW.items()}

SEA = {'4WAY','3WAYサーキュレーター','卓上冷感クーラー','5WAY腰掛けファン','瞬間冷却ハンディファン','接触冷感UVパーカー',
'接触冷感UVアームカバー','瞬間冷感ポンチョ','完全遮光・接触冷感UVハット','形状記憶日傘','完全遮光・形状記憶',
'偏光・調光サングラス','害虫ブロッカー','湯上がりガーゼワンピース'}

wb = Workbook(); wb.remove(wb.active)
TH = Font(name='Arial', bold=True, color='FFFFFF', size=10); TD = Font(name='Arial', size=10)
NEG = Font(name='Arial', size=10, color='CC0000'); HEAD = PatternFill('solid', fgColor='305496')
thin = Border(*[Side(style='thin', color='CCCCCC')] * 4)

def sh(title, note, cols, data, widths, fmt=None):
    ws = wb.create_sheet(title); r = 1
    if note:
        for ln in note: ws.cell(r, 1, ln).font = Font(name='Arial', size=10, bold=(r == 1)); r += 1
        r += 1
    for c, v in enumerate(cols, 1): ws.cell(r, c, v)
    for c in range(1, len(cols) + 1):
        x = ws.cell(r, c); x.font = TH; x.fill = HEAD; x.border = thin
        x.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 30; hr = r; r += 1
    for row in data:
        for c, v in enumerate(row, 1):
            x = ws.cell(r, c, v); x.border = thin; x.font = TD
            if isinstance(v, (int, float)) and v < 0: x.font = NEG
            if fmt and c in fmt: x.number_format = fmt[c]
        r += 1
    for i, w_ in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i < 27 else 'A' + chr(38 + i)].width = w_
    ws.freeze_panes = ws.cell(hr + 1, 1).coordinate
    return ws

# ===== 日次CSVを読む（3窓判定・週次・日次推移で使う）=====
Sd = collections.defaultdict(lambda: collections.defaultdict(lambda: [0, 0]))
for r_ in csv.DictReader(open('data/daily/daily_sales.csv')):
    Sd[r_['name']][r_['date']][0] += int(r_['gross']); Sd[r_['name']][r_['date']][1] += int(r_['disc'])
ADd = collections.defaultdict(lambda: collections.defaultdict(float))
for r_ in csv.DictReader(open('data/daily/daily_ad.csv')): ADd[r_['campaign']][r_['date']] += float(r_['cost'])
COSTa = dict(R['COST']); PRICEa = dict(R['PRICE'])
COSTa['ナノガラス脱毛パッド'] = (855 * 757 + 373 * 740) / 1228          # 30日実測ミックスの加重平均【近似】
COSTa['壁掛けディスペンサー'] = (41 * 2528 + 13 * 1981) / 54
PRICEa['壁掛けディスペンサー'] = 363920 / 54
ALLD = sorted({d for v in Sd.values() for d in v if d <= D1[0]})
D5 = ALLD[-5:]
def win(n, days):
    """商品nのその窓の（実売, 原価【近似】, 広告費）"""
    g = sum(Sd[n][d][0] for d in days if d in Sd[n]); dc = sum(Sd[n][d][1] for d in days if d in Sd[n])
    c = g / PRICEa[n] * COSTa[n] if g and n in PRICEa and n in COSTa else 0
    a = sum(ADd[MAP.get(n, n)].get(d, 0) for d in days)
    return g - dc, c, a

# ---------- 1 シンプル判定（3窓判定つき） ----------
def blk(n, N, K, A_):
    s = N.get(n, 0); k = K.get(n, 0); a = A_.get(n, 0); return s, k, a, s - k - a
SIMPLE = []
for n in sorted(R['N7'], key=lambda x: -R['N7'][x]):
    a7 = R['A7'].get(n, 0)
    if not a7: continue
    s7, k7, _, p7 = blk(n, R['N7'], R['K7'], R['A7'])
    s3, k3, a3, p3 = blk(n, R['N3'], R['K3'], R['A3']); s1, k1, a1, p1 = blk(n, R['N1'], R['K1'], R['A1'])
    gm = 1 - k7 / s7; be = 1 / gm; tg = 1 / (gm - 0.35) if gm - 0.35 > 0 else None; mer = s7 / a7
    cp = MAP.get(n, n); days = [d for d in D7 if d in BSNAP[cp]]
    wu = (a7 / sum(BSNAP[cp][d] for d in days)) if len(days) == 7 and sum(BSNAP[cp][d] for d in days) else None
    # ★3窓判定（3日/5日/7日・すべて前日終端）
    w3 = []
    for dd in (D3, D5, D7):
        sx, cx, ax = win(n, dd)
        w3.append((sx / ax if ax else None, 1 / (1 - cx / sx) if sx and sx > cx else None))
    below = all(m is not None and b is not None and m < b for m, b in w3)
    tight = all(m is not None and b is not None and m - b < 0.30 for m, b in w3)
    rich = all(m is not None and tg and m >= tg for m, _ in w3) and wu is not None and wu >= 0.95
    tri = '🚨3窓すべて分岐割れ→停止/−50%' if below else ('⚠️3窓すべて余裕<0.30→−25%' if tight else
          ('🚀3窓すべて目標超＋消化95%→+25%' if rich else '⏸窓が割れた→据え置き'))
    if p7 < 0 or mer < be: st = '🚨縮小・停止'
    elif (a3 and s3 / a3 < be) or mer < be + 0.30: st = '🔧改善'
    elif tg and mer >= tg and wu and wu >= 0.95: st = '🚀伸ばす候補'
    else: st = '✅維持'
    mm = mer * 0.7; per = mm * gm - 1
    # 頑健性: 限界MER = 平均MER×0.6/0.7/0.8 の3端で符号が変わらないか
    ends = [mer * x * gm - 1 for x in (0.6, 0.7, 0.8)]
    rob = '一致' if (all(e > 0 for e in ends) or all(e < 0 for e in ends)) else '⚠️符号が割れる'
    SIMPLE.append([n, round(mer, 2), round(s3 / a3, 2) if a3 else '', round(s1 / a1, 2) if a1 else '',
        round(be, 2), round(tg, 2) if tg else '', round(p7), round(p3), round(p1),
        round(wu, 3) if wu else '', round(per, 3), rob, st, tri])
sh('シンプル判定', ['LOOTY 2026-08-25 定例（昨日=8/24 月）',
 '状態: 🚨縮小・停止=7日MER<分岐 or 7日利益マイナス ／ 🔧改善=3日MER<分岐 or 余裕<0.30 ／ 🚀伸ばす候補=7日MER≥目標かつ週消化率≥95% ／ ✅維持',
 '3窓判定（3日/5日/7日・すべて8/24終端）が本番の意思決定ルール。窓が1つでも割れたら据え置き＝動かさない',
 '「広告費1円あたり利益」= 限界MER×粗利率 − 1（限界MER = 平均MER×0.7）。頑健性は×0.6/0.7/0.8 の3端で符号が変わらないかを見る'],
 ['商品','7日MER','3日MER','前日MER','分岐','目標','7日利益','3日利益','前日利益','週消化率','1円あたり利益','頑健性','状態','3窓判定'],
 SIMPLE, [26,9,9,9,8,8,12,12,12,10,12,13,14,30], {7:'#,##0',8:'#,##0',9:'#,##0',10:'0.0%'})

# ---------- 2 ネクストアクション ----------
p3ma = (sum(R['N3'].values()) - sum(R['K3'].values()) - R['ACCT3']) / 3
NA = [
 ['🔴 確認待ち','むくみ取りかっさ の予算',
  f'8/24b(256,000)→8/25(258,000)の差分は1件のみ: むくみ取りかっさ 8,000→10,000（+25%）。ユーザーが8/24に意図的に実施と確認済み。'
  '中間判定は明日8/26で、しきい値は宣言済みの「RPM/CPM<0.7で停止」。しきい値は率なので予算変更では無効化されないが、'
  '8/25は14:47時点で6,062円消化・Shopify注文ゼロ。判定は宣言どおり執行する','—','8/26'],
 ['🔴 判定日','3D足臭リセットブラシ CR「B」',
  'Bの3日累計消化 7,852円 ≧ 関門1,500円 → 消化ゲートは前倒しで通過。消化シェアは 8/23 8.3%→8/24 34.6%→8/25 63.0% でBが主役。'
  'BのCPMは3日平均3,059円でAの5,611円より45%安いが、BのCTRは1.40%でAの1.81%より低い。本判定は宣言どおり8/30（広告セット全体のRPM/CPM）',
  '—','8/30'],
 ['🚨 要対応','形状記憶日傘',
  f"7日 実売257,964 / 広告200,857 / 利益 −12,270円。MER 1.28 < 分岐1.37 で稼働中で唯一の実質赤字。3窓判定の結果は左表『3窓判定』列を参照。"
  '撤退プロトコル step2（7日利益マイナス→テコ入れ着手）に該当。8/24に22,000→17,000へ下げた直後で3クリーン日ロック中のため、判定日は8/27',
  '週+約15,000円','8/27'],
 ['🔴 判定日','快適マジックインソール',
  '8/24に18,000→22,000へ増額済み。宣言済み基準は「1日あたり利益 ≥ 23,589円」。7日利益+171,033円（1日24,433円）・利益率47.5%で店内最高。'
  'MER 2.91 > 目標2.13。判定日は8/27','—','8/27'],
 ['🔴 判定日','ナノガラス脱毛パッド のコピー修正',
  '宣言済み基準は「点数/注文 1.06 → 1.15」。7日実測は 279個/260注文 = 1.07 でほぼ動いていない。判定日は8/28',
  '—','8/28'],
 ['🔴 判定日','4WAY・接触冷感UVアームカバー の停止',
  '宣言済み基準は「全店週次利益(8/24-30) ≥ 696,901円」。両方とも7日窓に残消化のみ（4WAYは8/24消化ゼロ）','—','8/30'],
 ['🔴 判定日','優先配送 790円',
  '宣言済み基準は「14日で件数 ≥ 28件」。7日実測12件（8/18-24）。8/22に536→790へ値上げ済み','—','9/4'],
 ['👀 監視','非常ブレーキ',
  f'3日移動平均利益 {p3ma:,.0f}円 < 発動ライン130,000円 → **ライン割れ中**。ただし宣言済みの対処（カタログ減額・赤字商品一律−25%）は、'
  f'日予算を590,000→258,000（−56.3%）まで下げる形で既に方向として実行済み。前日利益は134,495円で床を上回った','—','毎日'],
 ['👀 監視','当日8/25のCVR低下',
  '00:00-13:59で昨日とクリック数が同じ（1,486 vs 1,487）なのに注文が58→42。CVRのみ −27.6%。ただし z=−2.15・両側p≒0.032 で'
  '32日に1回は起きる幅、かつ先週火曜比では注文 −2.3% ＝ ほぼ同水準。単日では動かさない（詳細は「当日診断」シート）','—','8/26'],
 ['🟡 推奨','購入通知アプリ',
  'Judge.me の Pop-up Reviews Widget は導入済みアプリの機能なので追加0円。Sales Popを足すならQikify Basic $6.99（訪問者無制限）。'
  'ProveSourceは訪問者137,022人だとMonster $109で15.6倍。効果は±15%未満だと測れないので「13日で−15%以上悪化していないか」で見る','—','任意'],
 ['🟡 推奨','偽タイマー・偽の人数の削除',
  'custom_liquid_T3RXEV（偽カウントダウン）／ custom_liquid_LgNzfb（偽の閲覧人数）。テーマエディタでブロックを削除するだけ。'
  'ステマ規制・不当表示のリスク除去が確定でプラス','—','任意'],
 ['🟡 推奨','カテゴリタグ40件の付与',
  'data/tags/category-mutation-2026-08-23.graphql を GraphiQL で1回実行 ＋ ナノバブルの季節タグ是正。'
  f'判定は実施+14日の異商品ミックス率。現在 (747−722)/722 = 3.46%（前回4.32%から低下）・合格6.0%','—','任意'],
]
sh('ネクストアクション', ['今日やること・判定カレンダー・持ち越しタスク（実施が確認できるまで毎日残す）'],
   ['区分','商品/対象','内容','効果','期限'], NA, [12, 26, 78, 16, 10])

# ---------- 3 全体サマリー ----------
SUMR = []
for lbl, N, K, ACC in [('前日 8/24(月)', R['N1'], R['K1'], R['ACCT1']), ('3日 8/22-24', R['N3'], R['K3'], R['ACCT3']),
        ('7日 8/18-24', R['N7'], R['K7'], R['ACCT7']), ('30日 7/26-8/24', R['N30'], R['K30'], R['ACCT30']),
        ('★当月累積 8/01-24', R['NC'], R['KC'], R['ACCTC'])]:
    s = sum(N.values()); c = sum(K.values()); p = s - c - ACC
    SUMR.append([lbl, s, c, round(ACC), round(c + ACC), round((c + ACC) / s, 4), round(p), round(p / s, 4),
                 round(s / ACC, 2), round(1 / (1 - c / s), 2), round(s * FEE), round(p - s * FEE), round((p - s * FEE) / s, 4)])
sh('全体サマリー', ['全体サマリー（売上=Shopify gross−値引・広告費=Meta実測。返品は売上からも利益からも控除しない=A案）',
 '総合原価＝原価＋広告費。恒等式 総合原価率＋利益率＝100% が全窓で成立していることを確認済み',
 '縦照合3本すべて通過: Σ商品gross=全店554,020 ／ Σ値引=24,637 ／ Σキャンペーン広告費=アカウント実消化',
 '決済ブレンド率 3.452%（2026-08-01再計測: SP76.2%×3.25% ＋ KOMOJUスマホ16.3%×4.1% ＋ Paidy7.5%×4.1%）',
 f'7日の全店CVR（Shopifyセッション基準）= {STORE_ORD7}注文 ÷ {STORE_SESS7:,}セッション = {STORE_ORD7/STORE_SESS7:.2%}'],
 ['窓','実売','原価','広告費','総合原価','総合原価率','利益','利益率','MER','分岐MER','決済手数料','手数料控除後利益','手数料後利益率'],
 SUMR, [20,14,13,13,13,11,13,10,8,9,12,15,12],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'#,##0',6:'0.0%',7:'#,##0',8:'0.0%',9:'0.00',10:'0.00',11:'#,##0',12:'#,##0',13:'0.0%'})
ws = wb['全体サマリー']; r = ws.max_row + 3
ws.cell(r, 1, '■ 昨日(8/24 月) × 期間比較').font = Font(name='Arial', bold=True, size=11); r += 1
S1 = sum(R['N1'].values()); C1 = sum(R['K1'].values()); A1 = R['ACCT1']
per = [('実売', S1, sum(R['N3'].values())/3, sum(R['N7'].values())/7, sum(R['N30'].values())/30),
       ('原価', C1, sum(R['K3'].values())/3, sum(R['K7'].values())/7, sum(R['K30'].values())/30),
       ('広告費', A1, R['ACCT3']/3, R['ACCT7']/7, R['ACCT30']/30),
       ('利益', S1-C1-A1, (sum(R['N3'].values())-sum(R['K3'].values())-R['ACCT3'])/3,
        (sum(R['N7'].values())-sum(R['K7'].values())-R['ACCT7'])/7,
        (sum(R['N30'].values())-sum(R['K30'].values())-R['ACCT30'])/30)]
for c, v in enumerate(['指標','8/24(月)','3日平均/日','7日平均/日','30日平均/日','vs7日平均'], 1): ws.cell(r, c, v)
for c in range(1, 7):
    x = ws.cell(r, c); x.font = TH; x.fill = HEAD; x.border = thin; x.alignment = Alignment(horizontal='center')
r += 1
for lbl, d1, m3, m7, m30 in per:
    for c, v in enumerate([lbl, round(d1), round(m3), round(m7), round(m30), round(d1/m7-1, 4)], 1):
        x = ws.cell(r, c, v); x.border = thin; x.font = NEG if (isinstance(v, (int, float)) and v < 0) else TD
        if c in (2,3,4,5): x.number_format = '#,##0'
        if c == 6: x.number_format = '+0.0%;-0.0%'
    r += 1
ws.cell(r+1, 1, '⚠️ 「同曜日平均比 −37.9%」は読まないこと。直近30日の月曜（7/27・8/3・8/10・8/17）が夏物崩壊前で汚染されている。'
  f'実態は7日平均比 +2.4%。3日移動平均利益 {p3ma:,.0f}円 は非常ブレーキ130,000円を下回っているが、'
  '前日利益134,495円は床を上回った').font = Font(name='Arial', size=10, color='CC0000')

# ---------- 4 商品別 ----------
P = []
for n in sorted(R['N7'], key=lambda x: -R['N7'][x]):
    s7,k7,a7,p7 = blk(n,R['N7'],R['K7'],R['A7']); s1,k1,a1,p1 = blk(n,R['N1'],R['K1'],R['A1'])
    s3,k3,a3,p3 = blk(n,R['N3'],R['K3'],R['A3']); s30,k30,a30,p30 = blk(n,R['N30'],R['K30'],R['A30'])
    sc,kc,ac,pc = blk(n,R['NC'],R['KC'],R['AC'])
    gm = 1-k7/s7 if s7 else 0; mer = s7/a7 if a7 else None; be = 1/gm if gm else None
    tg = 1/(gm-0.35) if gm-0.35 > 0 else None; yo = (mer-be) if mer and be else None
    q7 = R['Q7'].get(n, 0); o = ORD.get(n); ipo = q7/o if o else None
    bcpa = (s7-k7)/q7*ipo if q7 and ipo else None; cpa = a7/o if o and a7 else None
    jud = ('📦広告なし' if not a7 else ('🚨赤字' if p7 < 0 else ('🔧テコ入れ' if yo is not None and yo < 0.30 else
           ('🚀増額候補' if tg and mer >= tg else '✅維持'))))
    P.append([n, '季節' if n in SEA else '通年', s7, k7, round(a7), round(p7), round(p7/s7,4) if s7 else '',
        round(mer,2) if mer else '—', round(be,2) if be else '—', round(tg,2) if tg else '—',
        round(yo,2) if yo is not None else '—', s1, round(p1), round(p1/s1,4) if s1 else '',
        round(p3), round(p30), round(pc), q7, o or '', round(ipo,2) if ipo else '',
        round(cpa) if cpa else '', round(bcpa) if bcpa else '', BUD.get(MAP.get(n,n), '—'), jud])
P.append(['カタログ全部（テスト）※全商品横断','—','—','—',round(R['CAT7'])]+['—']*18+[BUD.get('カタログ全部（テスト）','—'),'—'])
sh('商品別', ['商品別（7日 8/18-24 = 判定単位）。カタログ行の広告費は「7日合計」（前日ではない）',
  '分岐CPA(注文) = 1個あたり粗利 × 点数/注文。まとめ買い商品（点数/注文>1.2）は必ず注文ベースで比較する',
  'Σ商品広告費 + カタログ = Metaアカウント7日消化（差0円・w0824.pyでassert済み）',
  '★3D足臭の点数/注文は 15/13 = 1.15（n=13）。以前 n=1 で 2.00 と報告したのは誤り'],
 ['商品','区分','7日売上','7日原価','7日広告','7日利益','7日利益率','MER','分岐','目標','余裕','前日売上','前日利益','前日利益率',
  '3日利益','30日利益','当月累積利益','7日販売数','7日注文数','点数/注文','CPA(注文)','分岐CPA(注文)','現日予算','判定'],
 P, [26,6]+[12]*22, {3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',12:'#,##0',13:'#,##0',14:'0.0%',
                     15:'#,##0',16:'#,##0',17:'#,##0',21:'#,##0',22:'#,##0',23:'#,##0'})

# ---------- 5 レバー一覧 ----------
LV = []
for n in sorted(R['N7'], key=lambda x: -R['N7'][x]):
    cp = MAP.get(n, n); b = BUD.get(cp)
    if not b: continue
    s7, k7, a7, p7 = blk(n, R['N7'], R['K7'], R['A7'])
    if not a7: continue
    gm = 1-k7/s7; mer = s7/a7; mm = mer*0.7
    days = [d for d in D7 if d in BSNAP[cp]]
    wu = (a7/sum(BSNAP[cp][d] for d in days)) if len(days) == 7 and sum(BSNAP[cp][d] for d in days) else None
    ends = [mer*x*gm-1 for x in (0.6, 0.7, 0.8)]
    LV.append([n, '季節' if n in SEA else '通年', b, round(wu,3) if wu else '', round(mer,2), round(1/gm,2),
        round(mer-1/gm,2), round(mm*gm-1,3), round(ends[0],3), round(ends[2],3),
        '一致' if (all(e>0 for e in ends) or all(e<0 for e in ends)) else '⚠️符号が割れる',
        round((1-mm*gm)*b*0.25*7), round((mm*gm-1)*b*0.25*7), round(p7)])
LV.sort(key=lambda x: -max(x[11], x[12]))
sh('レバー一覧', ['「どれを動かすと一番効くか」は7日平均の利益額ではなく、動かしたときの変化量で並べる',
  '削減1円あたりの利益変化 = 1 − 限界MER×粗利率 ／ 増額1円あたり = 限界MER×粗利率 − 1（限界MER = 平均MER×0.7）',
  '★頑健性: ×0.6 と ×0.8 の両端で符号が変わらないものだけ動かしてよい。割れたら据え置き',
  '増額は週消化率95%以上のものだけが対象（それ未満は増やしても使われない）'],
 ['商品','区分','日予算','週消化率','MER','分岐','余裕','1円あたり利益(×0.7)','×0.6端','×0.8端','頑健性','−25%で週','+25%で週','7日利益'],
 LV, [26,6,11,10,8,8,8,15,10,10,13,12,12,12], {3:'#,##0',4:'0.0%',12:'#,##0',13:'#,##0',14:'#,##0'})

# ---------- 6 週次診断 ----------
def wblk(days, names=None):
    s = c = 0
    for n, v in Sd.items():
        if names is not None and n not in names: continue
        g = sum(v[d][0] for d in days if d in v); dc = sum(v[d][1] for d in days if d in v)
        if not g: continue
        s += g-dc
        if n in PRICEa and n in COSTa: c += g/PRICEa[n]*COSTa[n]
    a = sum(ADd[MAP.get(n, n)].get(d, 0) for n in (names if names is not None else Sd) for d in days) if names is not None \
        else sum(ADd[cp].get(d, 0) for cp in ADd for d in days)
    return s, c, a, s-c-a
YR = {n for n in Sd if n and n not in SEA}
end = datetime.date(2026, 8, 24); WK = []
for i in range(6):
    a = end - datetime.timedelta(days=7*i+6)
    WK.append((f"{a.strftime('%m/%d')}-{(a+datetime.timedelta(days=6)).strftime('%m/%d')}",
               [(a+datetime.timedelta(days=j)).isoformat() for j in range(7)]))
WK.reverse()
WD_ = []
for lbl, days in WK:
    if days[0] < '2026-06-29': continue
    s, c, a, p = wblk(days); ss, _, sa, sp = wblk(days, SEA); ys, _, ya, yp = wblk(days, YR)
    WD_.append([lbl, round(s), round(c), round(a), round(p), round(p/s,4), round(s/a,2), round(c/s,4), round(a/s,4),
                round(sp), round(ss/sa,2) if sa else '', round(yp), round(ys/ya,2) if ya else ''])
sh('週次診断', ['全店の週次分解。原価は日次のバリアント構成が取れないナノガラス/ディスペンサーのみ30日実測ミックスの加重平均単価【近似】',
  '★夏物の崩壊が全体の正体。壊れたのは季節物だけで、通年物は伸びている',
  '★最新週(8/18-24)は日予算256,000〜258,000で運転。日予算590,000だった7月とは水準が違うので、利益額の絶対比較はしない'],
 ['週','実売','原価','広告費','利益','利益率','MER','原価率','広告費率','季節物 利益','季節物MER','通年物 利益','通年物MER'],
 WD_, [14,13,12,12,12,9,8,9,10,15,11,15,11],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'#,##0',6:'0.0%',7:'0.00',8:'0.0%',9:'0.0%',10:'#,##0',11:'0.00',12:'#,##0',13:'0.00'})

# ---------- 7 CVR順位・ファネル（RPM/CPM つき）----------
FN = []
for n, (c, imp, clk) in MF.items():
    s = R['N7'].get(n, 0); k = R['K7'].get(n, 0); q = R['Q7'].get(n, 0); o = ORD.get(n, 0)
    if not s or not o or not clk: continue
    ipo = q/o; bcpa = (s-k)/q*ipo; cpm = c/imp*1000
    ctr = clk/imp; cvr = o/clk; gpo = (s-k)/o          # 粗利/注文
    need = cpm/(gpo*1000)                              # 必要 CTR×CVR
    FN.append([n, '季節' if n in SEA else '通年', imp, round(ctr,4), clk, round(cvr,4), round(s/clk),
               round(c/clk), round(c/o), round(bcpa), round(bcpa-c/o), round(cpm), round(gpo),
               round(ctr*cvr*10000, 2), round(need*10000, 2), round((ctr*cvr)/need, 2)])
mc = statistics.median([x[5] for x in FN]) if FN else 0; mr = statistics.median([x[6] for x in FN]) if FN else 0
FN.sort(key=lambda x: -x[15])
sh('CVR順位', ['CVR = Shopify注文 ÷ Metaリンククリック（Metaの購入数は損益に使わない）',
  f'中央値 CVR {mc*100:.2f}% ／ 売上perクリック {mr:.0f}円',
  '★RPM/CPM = CTR×CVR×粗利/注文 ÷ CPM。1.0未満が赤字圏。★CTR単独では成否を判別できない（実測で反証済み）',
  '★必要CTR×CVR(‱) = CPM ÷ (粗利/注文 × 1000)。この列を下回っている商品が赤字'],
 ['商品','区分','インプ','CTR','クリック','CVR','売上/クリック','CPC','CPA(注文)','分岐CPA(注文)','余裕','CPM','粗利/注文',
  '実測CTR×CVR(‱)','必要CTR×CVR(‱)','RPM/CPM'],
 FN, [26,6,12,9,11,9,13,9,12,15,11,10,11,15,15,10],
 {3:'#,##0',4:'0.00%',5:'#,##0',6:'0.00%',7:'#,##0',8:'#,##0',9:'#,##0',10:'#,##0',11:'#,##0',12:'#,##0',13:'#,##0'})

# ---------- 8 日次推移 ----------
WDJ = ['月','火','水','木','金','土','日']
pr = {}
for d in ALLD:
    s = c = 0
    for n, v in Sd.items():
        if n == '' or d not in v: continue
        g, dc = v[d]; s += g-dc
        if g and n in PRICEa and n in COSTa: c += g/PRICEa[n]*COSTa[n]
    a = sum(ADd[cp].get(d, 0) for cp in ADd); pr[d] = (s, c, a, s-c-a)
DD = []
for j, d in enumerate(ALLD):
    if d < '2026-07-13': continue
    s, c, a, p = pr[d]; ma = sum(pr[x][3] for x in ALLD[j-2:j+1])/3
    DD.append([d, WDJ[datetime.date.fromisoformat(d).weekday()], round(s), round(c), round(a), round(p),
               round(p/s,4), round(s/a,2), round(ma), '⚠️非常ブレーキ' if ma < 130000 else ''])
sh('日次推移', ['日次の利益は上記【近似】原価ベース。4期間サマリーはバリアント別実数量の正確値',
  '非常ブレーキ: 3日移動平均利益 < 130,000円 で発動（カタログ→赤字商品一律−25%）'],
 ['日付','曜','実売','原価','広告費','利益','利益率','MER','3日移動平均利益','警告'],
 DD, [12,5,13,12,12,12,9,8,16,16], {3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',8:'0.00',9:'#,##0'})

# ---------- 9 当日診断（8/25 14:47時点）----------
IH = [
 ['広告費', 130987, 125201], ['インプ', 59042, 51986], ['クリック', 1486, 1487],
 ['Shopify gross', 316100, 227000], ['注文', 58, 42],
]
INTRA = [[k, a, b, round(b/a-1, 4)] for k, a, b in IH]
INTRA += [['CTR', round(1486/59042,4), round(1487/51986,4), round((1487/51986)/(1486/59042)-1,4)],
          ['CPM', round(130987/59042*1000), round(125201/51986*1000), round((125201/51986)/(130987/59042)-1,4)],
          ['CVR(注文÷クリック)', round(58/1486,4), round(42/1487,4), round((42/1487)/(58/1486)-1,4)],
          ['客単価', round(316100/58), round(227000/42), round((227000/42)/(316100/58)-1,4)]]
sh('当日診断', ['2026-08-25 14:47 JST 時点。両日とも 00:00〜13:59 JST の14時間（完全な時間のみ）で揃えた like-for-like 比較',
  '★クリック数はほぼ同一（1,486 vs 1,487）なのに注文が58→42。配信量でもCTRでもCPMでもなく、CVRだけが落ちている',
  '★ただし単日では動かさない。昨日のCVRが続く前提の期待注文58.0件（±7.5）に対し実測42件 → z=−2.15・両側p≒0.032 ＝ 32日に1回は起きる幅',
  '★先週火曜(8/18)の同時刻比では 売上 −5.8% / 注文 −2.3% ＝ ほぼ同水準。「昨日(月)が強すぎた」側面が大きい',
  '★注文はクリックより遅れて立つが、末尾1時間を両日から落としてもCVR −21.4% で結論は変わらない',
  '★日中の着地予測は出さない（既存ルール）'],
 ['指標','8/24(月) 00-13時','8/25(火) 00-13時','差'], INTRA, [22,18,18,12],
 {2:'#,##0',3:'#,##0',4:'+0.0%;-0.0%'})
ws = wb['当日診断']; r = ws.max_row+2
for c, v in enumerate(['当日ゼロ注文のまま消化中（14:47時点）','消化','クリック','Shopify注文'], 1):
    x = ws.cell(r, c, v); x.font = TH; x.fill = HEAD; x.border = thin
r += 1
for nm, cst, clk in [('むくみ取りかっさ',6062,44), ('3D足臭リセットブラシ',5335,22), ('偏光・調光サングラス',3942,41)]:
    for c, v in enumerate([nm, cst, clk, 0], 1):
        x = ws.cell(r, c, v); x.border = thin; x.font = TD
        if c == 2: x.number_format = '#,##0'
    r += 1
ws.cell(r+1, 1, '3商品とも事故停止の基準（分岐CPA×3の消化で購入ゼロ）には未達。本日は動かさない。').font = Font(name='Arial', size=10, color='CC0000')

# ---------- 10 CVR推移 ----------
SES = {'2026-08-04':6639,'2026-08-05':5425,'2026-08-06':5138,'2026-08-07':4827,'2026-08-08':5527,'2026-08-09':5886,
'2026-08-10':5245,'2026-08-11':6001,'2026-08-12':4900,'2026-08-13':5126,'2026-08-14':5035,'2026-08-15':4871,
'2026-08-16':5349,'2026-08-17':3438,'2026-08-18':3523,'2026-08-19':3190,'2026-08-20':3356,'2026-08-21':3120,
'2026-08-22':3220,'2026-08-23':3438,'2026-08-24':3192,'2026-08-25':1685}
ORDD = {'2026-08-04':188,'2026-08-05':152,'2026-08-06':157,'2026-08-07':173,'2026-08-08':184,'2026-08-09':187,
'2026-08-10':141,'2026-08-11':176,'2026-08-12':124,'2026-08-13':125,'2026-08-14':120,'2026-08-15':142,
'2026-08-16':143,'2026-08-17':98,'2026-08-18':89,'2026-08-19':106,'2026-08-20':113,'2026-08-21':110,
'2026-08-22':98,'2026-08-23':106,'2026-08-24':100,'2026-08-25':44}
CV = [[d, WDJ[datetime.date.fromisoformat(d).weekday()], SES[d], ORDD[d], round(ORDD[d]/SES[d], 4),
       '★当日・14:47時点' if d == '2026-08-25' else ''] for d in sorted(SES)]
o1 = sum(ORDD[d] for d in sorted(SES) if d <= '2026-08-16'); s1_ = sum(SES[d] for d in sorted(SES) if d <= '2026-08-16')
o2 = sum(ORDD[d] for d in sorted(SES) if '2026-08-17' <= d <= '2026-08-24'); s2 = sum(SES[d] for d in sorted(SES) if '2026-08-17' <= d <= '2026-08-24')
sh('CVR推移', ['全店CVR = Shopify注文 ÷ Shopifyセッション。商品別レポートのCVR（注文÷Metaクリック）とは分母が違うので混ぜないこと',
  f'★8/04-16（広告費が大きかった時期）: {o1}注文 / {s1_:,}セッション = {o1/s1_:.2%}',
  f'★8/17-24（広告費を絞った後）: {o2}注文 / {s2:,}セッション = {o2/s2:.2%} → 広告費を削ったらCVRは **{o2/s2/(o1/s1_)-1:+.1%}** 改善している',
  '★つまり全店CVRは低下トレンドにない。8/25の落ち込みは単日の揺れとして扱う',
  '★セッションは 5,000〜6,000/日 → 3,200〜3,500/日 へ減った。これは日予算 590,000→258,000 の直接の結果'],
 ['日付','曜','セッション','注文','全店CVR','備考'], CV, [12,5,12,10,10,20],
 {3:'#,##0',4:'#,##0',5:'0.00%'})

# ---------- 11 曜日指数 ----------
sh('曜日指数', ['直近30日(7/26-8/24)・祝日 8/11(山の日) を除外。売上=gross−値引',
  '★7月の指数（火87で最弱・日125）はもう使えない。夏物崩壊で構成が変わり、火が107.0まで上がった',
  '★同曜日平均との比較は当面使わない。直近30日の月曜（7/27・8/3・8/10・8/17）が夏物崩壊前で汚染されているため',
  '★代わりに「7日平均比」で読む。8/24は +2.4%'],
 ['曜日','指数(全体=100)'], [[k, round(v, 1)] for k, v in sorted(R['IDX'].items(), key=lambda x: -x[1])], [10, 16])

wb.save('data/reports/report-2026-08-25.xlsx')
print('シート:', wb.sheetnames)
print(f"3日移動平均利益 {p3ma:,.0f}円 / 非常ブレーキ130,000円")
