# -*- coding: utf-8 -*-
"""2026-08-11 定例レポート本体。8/06まであったシート構成を復元（8/07-8/10は4シートに欠落していた）。"""
import pickle, csv, datetime, collections, statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
R=pickle.load(open('/tmp/rep0811.pkl','rb')); FEE=0.03452
D1,D3,D7,D30,CUM=R['D1'],R['D3'],R['D7'],R['D30'],R['CUM']
# --- 予算スナップショット（8/11。前日から変更なし＝8/10の値を引き継ぐ）
rows=list(csv.reader(open('data/budget-snapshots.csv'))); hdr=rows[0]; body=[r for r in rows[1:] if r]
if not any(r[0]=='2026-08-11' for r in body):
    for r in [x for x in body if x[0]=='2026-08-10']: body.append(['2026-08-11',r[1],r[2],r[3]])
with open('data/budget-snapshots.csv','w',newline='') as f:
    w=csv.writer(f); w.writerow(hdr); w.writerows(body)
BUD=collections.defaultdict(int)
for r in body:
    if r[0]=='2026-08-11': BUD[r[1]]+=int(r[3])
BSNAP=collections.defaultdict(lambda: collections.defaultdict(int))
for r in body: BSNAP[r[1]][r[0]]+=int(r[3])
# --- Meta 7日ファネル実測（8/04-8/10, キャンペーン別）
MF={'ナノガラス脱毛パッド':(557010,307877,7989),'害虫ブロッカー':(335071,153900,3017),'4WAY':(312281,71565,2290),
'形状記憶日傘':(264419,86617,3262),'接触冷感UVパーカー':(206732,64057,2266),'完全遮光・形状記憶':(196320,50298,2298),
'偏光・調光サングラス':(156379,72011,1663),'W固定スマホ車載ホルダー':(145600,85005,2466),'ムダ毛シェーバー':(140227,120294,1510),
'5WAY腰掛けファン':(114926,39607,1605),'卓上冷感クーラー':(86707,30112,1253),'接触冷感UVアームカバー':(77385,20598,537),
'姿勢サポートチェア':(73719,43511,853),'携帯電動シェーバー':(64988,29895,485),'完全遮光・接触冷感UVハット':(63984,14061,678),
'瞬間冷感ポンチョ':(62841,13845,583),'2WAYシートボックス':(60114,25111,695),'バランスケアスリッパ':(35803,20339,381),
'3WAYサーキュレーター':(35716,8035,292),'健康サンダル':(29984,10666,292),'UV歯ブラシ除菌器':(3366,1075,21)}
ORD={'ナノガラス脱毛パッド':312,'W固定スマホ車載ホルダー':124,'接触冷感UVパーカー':100,'形状記憶日傘':91,'害虫ブロッカー':91,
'4WAY':87,'偏光・調光サングラス':54,'完全遮光・形状記憶':50,'ムダ毛シェーバー':45,'5WAY腰掛けファン':41,
'完全遮光・接触冷感UVハット':40,'2WAYシートボックス':37,'接触冷感UVアームカバー':26,'卓上冷感クーラー':24,
'姿勢サポートチェア':23,'携帯電動シェーバー':19,'バランスケアスリッパ':12,'瞬間冷感ポンチョ':8,
'3WAYサーキュレーター':6,'健康サンダル':3,'UV歯ブラシ除菌器':1}
MAP=R['MAP']; INV={v:k for k,v in MAP.items()}
SEA={'4WAY','3WAYサーキュレーター','卓上冷感クーラー','5WAY腰掛けファン','瞬間冷却ハンディファン','接触冷感UVパーカー',
'接触冷感UVアームカバー','瞬間冷感ポンチョ','完全遮光・接触冷感UVハット','形状記憶日傘','完全遮光・形状記憶','偏光・調光サングラス','害虫ブロッカー'}
wb=Workbook(); wb.remove(wb.active)
TH=Font(name='Arial',bold=True,color='FFFFFF',size=10); TD=Font(name='Arial',size=10)
NEG=Font(name='Arial',size=10,color='CC0000'); HEAD=PatternFill('solid',fgColor='305496')
GREEN=PatternFill('solid',fgColor='D6EFD8'); RED=PatternFill('solid',fgColor='FADBD8'); YEL=PatternFill('solid',fgColor='FFF2A8')
thin=Border(*[Side(style='thin',color='CCCCCC')]*4)
def sh(title,note,cols,data,widths,fmt=None,fills=None):
    ws=wb.create_sheet(title); r=1
    if note:
        for ln in note: ws.cell(r,1,ln).font=Font(name='Arial',size=10,bold=(r==1)); r+=1
        r+=1
    for c,v in enumerate(cols,1): ws.cell(r,c,v)
    for c in range(1,len(cols)+1):
        x=ws.cell(r,c); x.font=TH; x.fill=HEAD; x.border=thin
        x.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.row_dimensions[r].height=30; hr=r; r+=1
    for row in data:
        for c,v in enumerate(row,1):
            x=ws.cell(r,c,v); x.border=thin; x.font=TD
            if isinstance(v,(int,float)) and v<0: x.font=NEG
            if fmt and c in fmt: x.number_format=fmt[c]
        if fills: fills(ws,r,row)
        r+=1
    for i,w_ in enumerate(widths,1):
        ws.column_dimensions[chr(64+i) if i<27 else 'A'+chr(38+i)].width=w_
    ws.freeze_panes=ws.cell(hr+1,1).coordinate
    return ws
# ---------- 1 シンプル判定 ----------
def blk(n,N,K,A_): 
    s=N.get(n,0); k=K.get(n,0); a=A_.get(n,0); return s,k,a,s-k-a
SIMPLE=[]
for n in sorted(R['N7'],key=lambda x:-R['N7'][x]):
    a7=R['A7'].get(n,0)
    if not a7: continue
    s7,k7,_,p7=blk(n,R['N7'],R['K7'],R['A7']); s3,k3,a3,p3=blk(n,R['N3'],R['K3'],R['A3']); s1,k1,a1,p1=blk(n,R['N1'],R['K1'],R['A1'])
    gm=1-k7/s7; be=1/gm; tg=1/(gm-0.35) if gm-0.35>0 else None; mer=s7/a7
    cp=MAP.get(n,n); days=[d for d in D7 if d in BSNAP[cp]]
    wu=(a7/sum(BSNAP[cp][d] for d in days)) if len(days)==7 and sum(BSNAP[cp][d] for d in days) else None
    if p7<0 or mer<be: st='🚨縮小・停止'
    elif (a3 and s3/a3<be) or mer<be+0.30: st='🔧改善'
    elif tg and mer>=tg and wu and wu>=0.95: st='🚀伸ばす候補'
    else: st='✅維持'
    mm=mer*0.7; per=mm*gm-1
    if per<0: act='−25%（週+'+f"{(1-mm*gm)*BUD.get(cp,0)*0.25*7:,.0f}"+'円）'
    elif wu and wu>=0.95: act='+25%（週+'+f"{per*BUD.get(cp,0)*0.25*7:,.0f}"+'円）'
    else: act='据え置き（週消化率が95%未満＝増やしても使われない）'
    SIMPLE.append([n,round(mer,2),round(s3/a3,2) if a3 else '',round(s1/a1,2) if a1 else '',round(be,2),
        round(tg,2) if tg else '',round(p7),round(p3),round(p1),round(wu,3) if wu else '',round(per,3),st,act])
sh('シンプル判定',['LOOTY 2026-08-11 定例（昨日=8/10 月）',
 '状態: 🚨縮小・停止=7日MER<分岐 or 7日利益マイナス ／ 🔧改善=3日MER<分岐 or 余裕<0.30 ／ 🚀伸ばす候補=7日MER≥目標かつ週消化率≥95% ／ ✅維持',
 '「広告費1円あたり利益」= 限界MER×粗利率 − 1（限界MER = 平均MER×0.7、実測レンジ0.6〜0.8の中央）。正なら出すほど利益額が増える'],
 ['商品','7日MER','3日MER','前日MER','分岐','目標','7日利益','3日利益','前日利益','週消化率','1円あたり利益','状態','今やること'],
 SIMPLE,[26,9,9,9,8,8,12,12,12,10,12,14,26],{7:'#,##0',8:'#,##0',9:'#,##0',10:'0.0%'})
# ---------- 2 ネクストアクション ----------
NA=[['🔴 必須','姿勢サポートチェア','本判定。CPA(注文)3,205円=薄利帯・CVR2.70%（合格3.15%未達）→ 売価 5,980→7,980円へ戻す','週+20,200円【推計】','8/11'],
    ['🟡 推奨','W固定スマホ車載ホルダー','32,500→40,000（週消化102.9%で天井・1円あたり+0.803）','週+42,158円','8/11'],
    ['🟡 推奨','ナノガラス脱毛パッド２','63,000→79,000（セット２のみ。セット３は週消化89.8%で入れても使われない）','週+36,624円','8/11'],
    ['🟡 推奨','完全遮光・接触冷感UVハット','12,000→15,000（週消化106.6%・1円あたり+0.672）','週+14,112円','8/11'],
    ['🟡 推奨','4WAY','28,000→21,000（1円あたり−0.216・残4週）','週+10,584円','8/11'],
    ['🟡 推奨','携帯電動シェーバー','8,250→6,000（7日−3,474円・余裕−0.08）','週+5,308円','8/11'],
    ['👀 様子見','害虫ブロッカー','8/13の秋CR本判定まで予算を触らない（動かすとCR効果が測れない）','—','8/13'],
    ['👀 様子見','日傘2種','短は長の減額効果の測定窓／長は3クリーン日ロック中','—','8/13'],
    ['👀 様子見','偏光・調光サングラス','LP判定（7日CVR≥3.0%）','—','8/13'],
    ['👀 様子見','2WAYシートボックス','本判定（7日CPA<3,037円／RPM_CPM≥1.3なら+60%）','—','8/13'],
    ['👀 様子見','瞬間冷感ポンチョ','判定（7日利益>0 かつ購入ゼロ日なし）→ 両方すでにアウト','—','8/12'],
    ['👀 様子見','数量割引','7日 点数/注文≥1.25（現1.182）→ 不合格見通し','—','8/14'],
    ['👀 様子見','バランスケアスリッパ','新素材D 第2関門（外部CTR≥2.00%かつ売上/クリック≥163円）','—','8/14'],
    ['👀 様子見','W固定スマホ車載ホルダー','週次ランプ判定 W-1(8/2-8) vs W0(8/9-15)・増分MER≥2.88','—','8/16'],
    ['🔴 必須','CJ原価の確認2件','①車載シートヒーター(目標1,400円以下) ②電動スカルプブラシ(目標900円以下)。新商品パイプラインが空','9月末の崖 週505,088円','未着手'],
    ['🔴 必須','CJ原価の再交渉','主力3品（ナノガラス749・4WAY1,730・害虫867）。広告を触らないのでリスクゼロ','週+64,647円','未着手']]
sh('ネクストアクション',['今日やること・判定カレンダー・持ち越しタスク（実施が確認できるまで毎日残す）'],
   ['区分','商品/対象','内容','効果','期限'],NA,[12,26,58,20,10])
# ---------- 3 全体サマリー ----------
SUMR=[]
for lbl,N,K,ACC in [('前日 8/10(月)',R['N1'],R['K1'],R['ACCT1']),('3日 8/08-10',R['N3'],R['K3'],R['ACCT3']),
        ('7日 8/04-10',R['N7'],R['K7'],R['ACCT7']),('30日 7/12-8/10',R['N30'],R['K30'],R['ACCT30']),
        ('★当月累積 8/01-10',R['NC'],R['KC'],R['ACCTC'])]:
    s=sum(N.values()); c=sum(K.values()); p=s-c-ACC
    SUMR.append([lbl,s,c,round(ACC),round(c+ACC),round((c+ACC)/s,4),round(p),round(p/s,4),round(s/ACC,2),
                 round(1/(1-c/s),2),round(s*FEE),round(p-s*FEE),round((p-s*FEE)/s,4)])
sh('全体サマリー',['全体サマリー（売上=Shopify gross−値引・広告費=Meta実測。返品は売上からも利益からも控除しない=A案）',
 '総合原価＝原価＋広告費。恒等式 総合原価率＋利益率＝100% が全窓で成立していることを確認済み',
 '当月累積 8/01-10: 返品 56,821円（返品率0.55%）／注文 1,868件／点数 2,307（点数/注文 1.235）',
 '決済ブレンド率 3.452%（2026-08-01再計測: SP76.2%×3.25% ＋ KOMOJUスマホ16.3%×4.1% ＋ Paidy7.5%×4.1%）'],
 ['窓','実売','原価','広告費','総合原価','総合原価率','利益','利益率','MER','分岐MER','決済手数料','手数料控除後利益','手数料後利益率'],
 SUMR,[18,14,13,13,13,11,13,10,8,9,12,15,12],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'#,##0',6:'0.0%',7:'#,##0',8:'0.0%',9:'0.00',10:'0.00',11:'#,##0',12:'#,##0',13:'0.0%'})
# 昨日×期間比較
ws=wb['全体サマリー']; r=ws.max_row+3
ws.cell(r,1,'■ 昨日(8/10 月) × 期間比較').font=Font(name='Arial',bold=True,size=11); r+=1
S1=sum(R['N1'].values()); C1=sum(R['K1'].values()); A1=R['ACCT1']
per=[('実売',S1,sum(R['N3'].values())/3,sum(R['N7'].values())/7,sum(R['N30'].values())/30),
     ('原価',C1,sum(R['K3'].values())/3,sum(R['K7'].values())/7,sum(R['K30'].values())/30),
     ('広告費',A1,R['ACCT3']/3,R['ACCT7']/7,R['ACCT30']/30),
     ('利益',S1-C1-A1,(sum(R['N3'].values())-sum(R['K3'].values())-R['ACCT3'])/3,
      (sum(R['N7'].values())-sum(R['K7'].values())-R['ACCT7'])/7,
      (sum(R['N30'].values())-sum(R['K30'].values())-R['ACCT30'])/30)]
for c,v in enumerate(['指標','8/10(月)','平常月曜平均','3日平均/日','7日平均/日','30日平均/日','vs7日平均'],1): ws.cell(r,c,v)
for c in range(1,8):
    x=ws.cell(r,c); x.font=TH; x.fill=HEAD; x.border=thin; x.alignment=Alignment(horizontal='center')
r+=1
for lbl,d1,m3,m7,m30 in per:
    mo=1117783 if lbl=='実売' else ''
    for c,v in enumerate([lbl,round(d1),mo,round(m3),round(m7),round(m30),round(d1/m7-1,4)],1):
        x=ws.cell(r,c,v); x.border=thin; x.font=NEG if (isinstance(v,(int,float)) and v<0) else TD
        if c in (2,3,4,5,6): x.number_format='#,##0'
        if c==7: x.number_format='+0.0%;-0.0%'
    r+=1
ws.cell(r+1,1,'対同曜日（平常月曜 1,117,783円）= 64.8%。3日移動平均利益 183,236円 > 非常ブレーキ130,000円 → 未発動').font=Font(name='Arial',size=10,color='CC0000')
# ---------- 4 商品別 ----------
P=[]
for n in sorted(R['N7'],key=lambda x:-R['N7'][x]):
    s7,k7,a7,p7=blk(n,R['N7'],R['K7'],R['A7']); s1,k1,a1,p1=blk(n,R['N1'],R['K1'],R['A1'])
    s3,k3,a3,p3=blk(n,R['N3'],R['K3'],R['A3']); s30,k30,a30,p30=blk(n,R['N30'],R['K30'],R['A30'])
    sc,kc,ac,pc=blk(n,R['NC'],R['KC'],R['AC'])
    gm=1-k7/s7 if s7 else 0; mer=s7/a7 if a7 else None; be=1/gm if gm else None
    tg=1/(gm-0.35) if gm-0.35>0 else None; yo=(mer-be) if mer and be else None
    q7=R['Q7'].get(n,0); o=ORD.get(n); ipo=q7/o if o else None
    bcpa=(s7-k7)/q7*ipo if q7 and ipo else None; cpa=a7/o if o and a7 else None
    jud=('📦広告なし' if not a7 else ('🚨赤字' if p7<0 else ('🔧テコ入れ' if yo is not None and yo<0.30 else
         ('🚀増額候補' if tg and mer>=tg else '✅維持'))))
    P.append([n,'季節' if n in SEA else '通年',s7,k7,round(a7),round(p7),round(p7/s7,4) if s7 else '',
        round(mer,2) if mer else '—',round(be,2) if be else '—',round(tg,2) if tg else '—',round(yo,2) if yo is not None else '—',
        s1,round(p1),round(p1/s1,4) if s1 else '',round(p3),round(p30),round(pc),q7,o or '',
        round(ipo,2) if ipo else '',round(cpa) if cpa else '',round(bcpa) if bcpa else '',BUD.get(MAP.get(n,n),'—'),jud])
P.append(['カタログ全部（テスト）※全商品横断','—','—','—',round(R['CAT7']),'—','—','—','—','—','—','—','—','—','—','—','—','—','—','—','—','—',BUD.get('カタログ全部（テスト）','—'),'—'])
sh('商品別',['商品別（7日=判定単位）。カタログ行の広告費は「7日合計」（前日ではない）',
  '分岐CPA(注文) = 1個あたり粗利 × 点数/注文。まとめ買い商品（点数/注文>1.2）は必ず注文ベースで比較する',
  'Σ商品利益 − カタログ7日広告費 = 全店7日利益（差0円）／ Σ商品広告費 + カタログ = Metaアカウント7日消化（差0円）'],
 ['商品','区分','7日売上','7日原価','7日広告','7日利益','7日利益率','MER','分岐','目標','余裕','前日売上','前日利益','前日利益率',
  '3日利益','30日利益','当月累積利益','7日販売数','7日注文数','点数/注文','CPA(注文)','分岐CPA(注文)','現日予算','判定'],
 P,[26,6]+[12]*22,{3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',12:'#,##0',13:'#,##0',14:'0.0%',15:'#,##0',16:'#,##0',17:'#,##0',21:'#,##0',22:'#,##0',23:'#,##0'})
# ---------- 5 レバー一覧 ----------
LV=[]
for n in sorted(R['N7'],key=lambda x:-R['N7'][x]):
    cp=MAP.get(n,n); b=BUD.get(cp)
    if not b: continue
    s7,k7,a7,p7=blk(n,R['N7'],R['K7'],R['A7'])
    if not a7: continue
    gm=1-k7/s7; mer=s7/a7; mm=mer*0.7
    days=[d for d in D7 if d in BSNAP[cp]]
    wu=(a7/sum(BSNAP[cp][d] for d in days)) if len(days)==7 and sum(BSNAP[cp][d] for d in days) else None
    LV.append([n,'季節' if n in SEA else '通年',b,round(wu,3) if wu else '',round(mer,2),round(1/gm,2),round(mer-1/gm,2),
        round(mm*gm-1,3),round((1-mm*gm)*b*0.25*7),round((mm*gm-1)*b*0.25*7),round(p7)])
LV.sort(key=lambda x:-max(x[8],x[9]))
sh('レバー一覧',['「どれを動かすと一番効くか」は7日平均の利益額ではなく、動かしたときの変化量で並べる',
  '削減1円あたりの利益変化 = 1 − 限界MER×粗利率 ／ 増額1円あたり = 限界MER×粗利率 − 1（限界MER = 平均MER×0.7）',
  '増額は週消化率95%以上のものだけが対象（それ未満は増やしても使われない）。P-1ガードは8/11時点で解除（3日比較 +4.3%）'],
 ['商品','区分','日予算','週消化率','MER','分岐','余裕','1円あたり利益','−25%で週','+25%で週','7日利益'],
 LV,[26,6,11,10,8,8,8,13,13,13,12],{3:'#,##0',4:'0.0%',9:'#,##0',10:'#,##0',11:'#,##0'})
# ---------- 6 週次診断 ----------
Sd=collections.defaultdict(lambda: collections.defaultdict(lambda:[0,0]))
for r_ in csv.DictReader(open('data/daily/daily_sales.csv')):
    Sd[r_['name']][r_['date']][0]+=int(r_['gross']); Sd[r_['name']][r_['date']][1]+=int(r_['disc'])
ADd=collections.defaultdict(lambda: collections.defaultdict(float))
for r_ in csv.DictReader(open('data/daily/daily_ad.csv')): ADd[r_['campaign']][r_['date']]+=float(r_['cost'])
COST=dict(R['COST']); PRICE=dict(R['PRICE'])
COST['ナノガラス脱毛パッド']=(835*757+364*740)/1199; COST['壁掛けディスペンサー']=(45*2528+13*1981)/58
PRICE['壁掛けディスペンサー']=391840/58
end=datetime.date(2026,8,10); WK=[]
for i in range(6):
    a=end-datetime.timedelta(days=7*i+6); WK.append((f"{a.strftime('%m/%d')}-{(a+datetime.timedelta(days=6)).strftime('%m/%d')}",
        [(a+datetime.timedelta(days=j)).isoformat() for j in range(7)]))
WK.reverse()
def wblk(days,names=None):
    s=c=0
    for n,v in Sd.items():
        if names is not None and n not in names: continue
        g=sum(v[d][0] for d in days if d in v); dc=sum(v[d][1] for d in days if d in v)
        if not g: continue
        s+=g-dc
        if n in PRICE and n in COST: c+=g/PRICE[n]*COST[n]
    a=sum(ADd[MAP.get(n,n)].get(d,0) for n in (names if names is not None else Sd) for d in days) if names is not None \
      else sum(ADd[cp].get(d,0) for cp in ADd for d in days)
    return s,c,a,s-c-a
YR={n for n in Sd if n and n not in SEA}
WD_=[]
for lbl,days in WK:
    if days[0]<'2026-06-29': continue
    s,c,a,p=wblk(days); ss,_,sa,sp=wblk(days,SEA); ys,_,ya,yp=wblk(days,YR)
    WD_.append([lbl,round(s),round(c),round(a),round(p),round(p/s,4),round(s/a,2),round(c/s,4),round(a/s,4),
                round(sp),round(ss/sa,2),round(yp),round(ys/ya,2)])
sh('週次診断',['全店の週次分解。原価は日次のバリアント構成が取れないナノガラス/ディスペンサーのみ30日実測ミックスの加重平均単価【近似】',
  '★結論: 原価率は6週間ほぼ不変（27.9%→28.8%）。利益率の−18.7ptは全部が広告費率の上昇（33.8%→51.6%）',
  '★利益減 −2,157,846円/週 のうち 季節物13品が −2,100,537円＝97.3%。通年物14品のMERは3週連続で回復中'],
 ['週','実売','原価','広告費','利益','利益率','MER','原価率','広告費率','季節物13品 利益','季節物MER','通年物14品 利益','通年物MER'],
 WD_,[14,13,12,12,12,9,8,9,10,15,11,15,11],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'#,##0',6:'0.0%',7:'0.00',8:'0.0%',9:'0.0%',10:'#,##0',11:'0.00',12:'#,##0',13:'0.00'})
# ---------- 7 CVR順位・ファネル ----------
FN=[]
for n,(c,imp,clk) in MF.items():
    s=R['N7'].get(n,0); k=R['K7'].get(n,0); q=R['Q7'].get(n,0); o=ORD.get(n,0)
    if not s or not o: continue
    ipo=q/o; bcpa=(s-k)/q*ipo
    FN.append([n,'季節' if n in SEA else '通年',imp,round(clk/imp,4),clk,round(o/clk,4),round(s/clk),
               round(c/clk),round(c/o),round(bcpa),round(bcpa-c/o),round(c/imp*1000)])
mc=statistics.median([x[5] for x in FN]); mr=statistics.median([x[6] for x in FN])
FN.sort(key=lambda x:-x[6])
sh('CVR順位',['CVR = Shopify注文 ÷ Metaアウトバウンドクリック（Metaの購入数は損益に使わない）',
  f'中央値 CVR {mc*100:.2f}% ／ 売上perクリック {mr:.0f}円。LP改修の対象は「CVRも売上/クリックも中央値未満」の商品だけ',
  'CVRの単純順位は価格帯を混同する。売上perクリック（=AOV×CVR）と同価格帯内で比較すること'],
 ['商品','区分','インプ','CTR','クリック','CVR','売上/クリック','CPC','CPA(注文)','分岐CPA(注文)','余裕','CPM'],
 FN,[26,6,12,9,11,9,13,9,12,15,11,10],
 {3:'#,##0',4:'0.00%',5:'#,##0',6:'0.00%',7:'#,##0',8:'#,##0',9:'#,##0',10:'#,##0',11:'#,##0',12:'#,##0'})
# ---------- 8 日次推移 ----------
days=sorted({d for v in Sd.values() for d in v}); WDJ=['月','火','水','木','金','土','日']
pr={}
for d in days:
    s=c=0
    for n,v in Sd.items():
        if n=='' or d not in v: continue
        g,dc=v[d]; s+=g-dc
        if g and n in PRICE and n in COST: c+=g/PRICE[n]*COST[n]
    a=sum(ADd[cp].get(d,0) for cp in ADd); pr[d]=(s,c,a,s-c-a)
DD=[]
for j,d in enumerate(days):
    if d<'2026-07-13': continue
    s,c,a,p=pr[d]; ma=sum(pr[x][3] for x in days[j-2:j+1])/3
    DD.append([d,WDJ[datetime.date.fromisoformat(d).weekday()],round(s),round(c),round(a),round(p),round(p/s,4),
               round(s/a,2),round(ma),'⚠️非常ブレーキ' if ma<130000 else ''])
sh('日次推移',['日次の利益は上記【近似】原価ベース。4期間サマリーはバリアント別実数量の正確値',
  '非常ブレーキ: 3日移動平均利益 < 130,000円 で発動（カタログ→4WAY→日傘の順に緊急減額）'],
 ['日付','曜','実売','原価','広告費','利益','利益率','MER','3日移動平均利益','警告'],
 DD,[12,5,13,12,12,12,9,8,16,16],{3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',8:'0.00',9:'#,##0'})
# ---------- 9 曜日指数 ----------
sh('曜日指数',['直近30日(7/12-8/10)・祝日7/20除外。売上=gross−値引',
  '★月曜の実績は7日平均ではなく「平常月曜平均」と比べる。祝日には平日の曜日指数を当てない（祝日は日曜級113〜127で評価）',
  '2026-08-11は山の日（火・祝）→ 8/12の評価は祝日基準で行う'],
 ['曜日','指数(全体=100)'],[[k,round(v,1)] for k,v in sorted(R['IDX'].items(),key=lambda x:-x[1])],[10,16])
wb.save('data/reports/report-2026-08-11.xlsx')
print('本体シート:',wb.sheetnames)
