# -*- coding: utf-8 -*-
"""2026-09-05 定例レポート本体（昨日=9/4 金）。"""
import pickle, csv, datetime, collections, statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
R = pickle.load(open('/tmp/rep0904w.pkl', 'rb'))
FEE = 0.03436   # 2026-09-01 月次再計測。次回 10/1
D1, D3, D7, D30, CUM = R['D1'], R['D3'], R['D7'], R['D30'], R['CUM']
SNAP = '2026-09-05'   # サングラス停止（9/5 00:15 直読でPAUSED確認）を反映した13セット・243,000円/日
rows = list(csv.reader(open('data/budget-snapshots.csv'))); body = [r for r in rows[1:] if r]
assert any(r[0] == SNAP for r in body), f'{SNAP} スナップショット未記録'
BUD = collections.defaultdict(int)
for r in body:
    if r[0] == SNAP: BUD[r[1]] += int(r[3])
assert sum(BUD.values()) == 243000, sum(BUD.values())
_BS = collections.defaultdict(dict)
for r in body:
    day = _BS[r[1]].setdefault(r[0][:10], {}); day[r[0]] = day.get(r[0], 0) + int(r[3])
BSNAP = collections.defaultdict(lambda: collections.defaultdict(int))
for cp, byday in _BS.items():
    for d, snaps in byday.items(): BSNAP[cp][d] = snaps[max(snaps)]

MAP = R['MAP']; INV = {v: k for k, v in MAP.items()}
# --- 7日(8/26-9/1) Shopify注文数（ShopifyQL GROUP BY product_title 実測）---
ORD = {'ナノガラス脱毛パッド':271,'W固定スマホ車載ホルダー':104,'快適マジックインソール':83,'ムダ毛シェーバー':43,
'バランスケアスリッパ':39,'むくみ取りかっさ':33,'姿勢サポートチェア':23,'2WAYシートボックス':20,
'ナノバブルシャワーヘッド':20,'完全遮光・接触冷感UVハット':11,'優先配送':7,'偏光・調光サングラス':7,
'温感EMSフェイシャルワンド':6,'携帯電動シェーバー':5,'ビジュアル耳かき':4,'壁掛けディスペンサー':2,
'ネックマッサージャー':1,'瞬間冷感ポンチョ':1,'接触冷感UVパーカー':1,'3D足臭リセットブラシ':1,
'ヘアドライタオル':1,'癒しの指圧マット':1,'1秒折り畳みチェア':1,'完全遮光・形状記憶':1}
STORE_ORD7, STORE_SESS7 = 670, 22319
assert sum(ORD.values()) == 686, sum(ORD.values())   # 全店670より多いのは複数商品を含む注文が両方で数えられるため

# --- 30日(8/03-9/01) Shopify注文数（ShopifyQL GROUP BY product_title 実測）---
ORD30 = {'ナノガラス脱毛パッド':1181,'W固定スマホ車載ホルダー':536,'快適マジックインソール':238,
'ムダ毛シェーバー':187,'形状記憶日傘':180,'2WAYシートボックス':157,'接触冷感UVパーカー':136,
'偏光・調光サングラス':135,'完全遮光・接触冷感UVハット':111,'害虫ブロッカー':108,'姿勢サポートチェア':101,
'4WAY':99,'バランスケアスリッパ':95,'完全遮光・形状記憶':70,'優先配送':62,'むくみ取りかっさ':60,
'接触冷感UVアームカバー':52,'5WAY腰掛けファン':44,'ナノバブルシャワーヘッド':31,'携帯電動シェーバー':29,
'3D足臭リセットブラシ':21,'卓上冷感クーラー':15,'瞬間冷感ポンチョ':8,'温感EMSフェイシャルワンド':6,
'ビジュアル耳かき':5,'ヘアドライタオル':4,'壁掛けディスペンサー':4,'健康サンダル':3,'癒しの指圧マット':2,
'ジェットウォッシャー':2,'湯上がりガーゼワンピース':2,'3WAYサーキュレーター':2,'瞬間冷却ハンディファン':1,
'UV歯ブラシ除菌器':1,'姿勢サポートベルト':1,'1秒折り畳みチェア':1,'ネックマッサージャー':1}

# --- 30日(8/03-9/01) Metaファネル（data_query 実測・Asia/Tokyo・アウトバウンドクリック）---
MF_RAW = {'ナノガラス脱毛パッド':(2408822,1297810,32989),'W固定スマホ車載ホルダー':(978272,500915,14214),
'形状記憶日傘':(626763,174012,6766),'カタログ全部（テスト）':(603732,149962,5461),
'ムダ毛シェーバー':(563572,457195,5379),'快適マジックインソール':(502400,305486,7767),
'害虫ブロッカー':(406335,186624,3534),'偏光・調光サングラス':(381788,164319,4247),
'2WAYシートボックス':(375989,129226,3622),'接触冷感UVパーカー':(345470,111265,3453),
'姿勢サポートチェア':(325270,203715,3986),'4WAY 取り付けOK 小型瞬間冷却ハンディファン':(321156,77651,2346),
'完全遮光・接触冷感UVハット':(275766,62860,2977),'完全遮光・形状記憶・晴雨兼用・UV日傘':(243947,59486,2662),
'バランスケアスリッパ':(219842,153717,2652),'5WAY腰掛けファン':(148444,48892,1969),
'接触冷感UVアームカバー':(140682,33016,1070),'むくみ取りかっさ':(133991,36792,1123),
'ナノバブルシャワーヘッド':(92522,43406,1200),'3D足臭リセットブラシ':(85838,24927,338),
'卓上冷感クーラー':(60364,23346,857),'携帯電動シェーバー':(53925,22174,423),
'瞬間冷感ポンチョ':(47030,10272,386),'温感EMSフェイシャルワンド':(26539,6085,145),
'ビジュアル耳かき':(26464,9402,148),'高吸水・速乾ヘアドライタオル':(10669,2706,75),
'カタログ全部（テスト） 夏以外':(6080,1903,50),'1秒折り畳みチェア':(5704,1773,49),
'3WAYサーキュレーター扇風機':(5602,1180,35),'健康サンダル':(4819,1768,57)}
MF = {INV.get(k, k): v for k, v in MF_RAW.items()}

SEA = {'4WAY','3WAYサーキュレーター','卓上冷感クーラー','5WAY腰掛けファン','瞬間冷却ハンディファン','接触冷感UVパーカー',
'接触冷感UVアームカバー','瞬間冷感ポンチョ','完全遮光・接触冷感UVハット','形状記憶日傘','完全遮光・形状記憶',
'偏光・調光サングラス','害虫ブロッカー','湯上がりガーゼワンピース'}

wb = Workbook(); wb.remove(wb.active)
TH = Font(name='Arial', bold=True, color='FFFFFF', size=10); TD = Font(name='Arial', size=10)
NEG = Font(name='Arial', size=10, color='CC0000'); HEAD = PatternFill('solid', fgColor='305496')
thin = Border(*[Side(style='thin', color='CCCCCC')] * 4)

def sh(title, note, cols, data, widths, fmt=None):
    ws = wb.create_sheet(title); r = 1
    if note:
        for ln in note: ws.cell(r, 1, ln).font = Font(name='Arial', size=10, bold=(r == 1)); r += 1
        r += 1
    for c, v in enumerate(cols, 1): ws.cell(r, c, v)
    for c in range(1, len(cols) + 1):
        x = ws.cell(r, c); x.font = TH; x.fill = HEAD; x.border = thin
        x.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 30; hr = r; r += 1
    for row in data:
        for c, v in enumerate(row, 1):
            x = ws.cell(r, c, v); x.border = thin; x.font = TD
            if isinstance(v, (int, float)) and v < 0: x.font = NEG
            if fmt and c in fmt: x.number_format = fmt[c]
        r += 1
    for i, w_ in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i < 27 else 'A' + chr(38 + i)].width = w_
    ws.freeze_panes = ws.cell(hr + 1, 1).coordinate
    return ws

# ===== 日次CSVを読む =====
Sd = collections.defaultdict(lambda: collections.defaultdict(lambda: [0, 0]))
for r_ in csv.DictReader(open('data/daily/daily_sales.csv')):
    Sd[r_['name']][r_['date']][0] += int(r_['gross']); Sd[r_['name']][r_['date']][1] += int(r_['disc'])
ADd = collections.defaultdict(lambda: collections.defaultdict(float))
for r_ in csv.DictReader(open('data/daily/daily_ad.csv')): ADd[r_['campaign']][r_['date']] += float(r_['cost'])
COSTa = dict(R['COST']); PRICEa = dict(R['PRICE'])
COSTa['ナノガラス脱毛パッド'] = (855 * 757 + 373 * 740) / 1228          # 30日実測ミックスの加重平均【近似】
COSTa['壁掛けディスペンサー'] = (41 * 2528 + 13 * 1981) / 54
PRICEa['壁掛けディスペンサー'] = 363920 / 54
ALLD = sorted({d for v in Sd.values() for d in v if d <= D1[0]})
D5 = ALLD[-5:]
def win(n, days):
    g = sum(Sd[n][d][0] for d in days if d in Sd[n]); dc = sum(Sd[n][d][1] for d in days if d in Sd[n])
    c = g / PRICEa[n] * COSTa[n] if g and n in PRICEa and n in COSTa else 0
    a = sum(ADd[MAP.get(n, n)].get(d, 0) for d in days)
    return g - dc, c, a

# ---------- 0 収益 ----------
WDJ0 = ['月', '火', '水', '木', '金', '土', '日']
def dayblk(d):
    s = c = 0
    for n, v in Sd.items():
        if n == '' or d not in v: continue
        g, dc = v[d]
        if not g: continue
        s += g - dc
        if n in PRICEa and n in COSTa: c += g / PRICEa[n] * COSTa[n]
    a = sum(ADd[cp].get(d, 0) for cp in ADd)
    return s, c, a, s - c - a

PROF = []
_pl = [dayblk(d) for d in ALLD]
for i, d in enumerate(ALLD):
    if d < ALLD[-21]: continue
    s, c, a, p = _pl[i]
    ma = sum(_pl[j][3] for j in range(i-2, i+1)) / 3 if i >= 2 else None
    PROF.append([d, WDJ0[datetime.date.fromisoformat(d).weekday()], round(s), round(c), round(a),
                 round(p), round(p/s, 4), round(p - s*FEE), round(s/a, 2) if a else '',
                 round(ma) if ma else '', ('🚨' if ma and ma < 130000 else ('✅' if ma else '')),
                 '★昨日' if d == D1[0] else ''])
# 8月確定（広告費の上書き後の最終値）と9月の滑り出し
AUG = [d for d in ALLD if '2026-08-01' <= d <= '2026-08-31']
_au = [dayblk(d) for d in AUG]
aus = sum(x[0] for x in _au); auc = sum(x[1] for x in _au); aua = sum(x[2] for x in _au); aup = aus-auc-aua
_cum = [dayblk(d) for d in CUM]
cs = sum(x[0] for x in _cum); cc = sum(x[1] for x in _cum); ca = sum(x[2] for x in _cum); cp = cs-cc-ca
p3ma = (sum(R['N3'].values()) - sum(R['K3'].values()) - R['ACCT3']) / 3
sh('収益', [f'LOOTY 収益レポート 2026-09-05（昨日 = 9/4 金）',
  f'★昨日の利益 {round(_pl[-1][3]):,}円（利益率 {_pl[-1][3]/_pl[-1][0]:.1%}・MER {_pl[-1][0]/_pl[-1][2]:.2f}）／手数料控除後 {round(_pl[-1][3]-_pl[-1][0]*FEE):,}円',
  f'★9月累積（9/01-04）: 実売 {cs:,.0f} / 利益 {cp:,.0f}円（{cp/cs:.1%}）／1日あたり {cp/4:,.0f}円（8月確定は {aup:,.0f}円・{aup/31:,.0f}円/日）',
  '★9/4は9月最悪の日（利益+20,130円・5.4%）だが分岐割れは回避。夜(20-24時)は週内最多23件＝解毒の初期シグナル（詳細はdecisions.md）',
  f'⚠️ 3日移動平均利益 {p3ma:,.0f}円。全店CVRトリガー（単日<2.27%）が9/4に初発動（2.25%）→ 処置のD1-D5診断は9/4昼に前倒し実施済み',
  '⚠️ 楽天スーパーセール 9/4 20:00〜9/11 01:59 開催中。セール期の悪化を理由に停止・減額を打たない（宣言済みガード）',
  '原価は日次のバリアント構成が取れないナノガラス/壁掛けディスペンサーのみ30日実測ミックスの加重平均単価【近似】。'
  '期間別サマリー（次シート）はバリアント別実数量の正確値'],
 ['日付','曜','実売','原価','広告費','利益','利益率','手数料後利益','MER','3日移動平均利益','ブレーキ','備考'],
 PROF, [12,5,13,12,12,13,9,14,8,16,9,10],
 {3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',8:'#,##0',9:'0.00',10:'#,##0'})

# ---------- 1 シンプル判定 ----------
def blk(n, N, K, A_):
    s = N.get(n, 0); k = K.get(n, 0); a = A_.get(n, 0); return s, k, a, s - k - a
SIMPLE = []
for n in sorted(R['N7'], key=lambda x: -R['N7'][x]):
    a7 = R['A7'].get(n, 0)
    if not a7: continue
    s7, k7, _, p7 = blk(n, R['N7'], R['K7'], R['A7'])
    s3, k3, a3, p3 = blk(n, R['N3'], R['K3'], R['A3']); s1, k1, a1, p1 = blk(n, R['N1'], R['K1'], R['A1'])
    gm = 1 - k7 / s7; be = 1 / gm; tg = 1 / (gm - 0.35) if gm - 0.35 > 0 else None; mer = s7 / a7
    cp = MAP.get(n, n); days = [d for d in D7 if d in BSNAP[cp]]
    wu = (a7 / sum(BSNAP[cp][d] for d in days)) if len(days) == 7 and sum(BSNAP[cp][d] for d in days) else None
    w3 = []
    for dd in (D3, D5, D7):
        sx, cx, ax = win(n, dd)
        w3.append((sx / ax if ax else None, 1 / (1 - cx / sx) if sx and sx > cx else None))
    below = all(m is not None and b is not None and m < b for m, b in w3)
    tight = all(m is not None and b is not None and m - b < 0.30 for m, b in w3)
    rich = all(m is not None and tg and m >= tg for m, _ in w3) and wu is not None and wu >= 0.95
    tri = '🚨3窓すべて分岐割れ→停止/−50%' if below else ('⚠️3窓すべて余裕<0.30→−25%' if tight else
          ('🚀3窓すべて目標超＋消化95%→+25%' if rich else '⏸窓が割れた→据え置き'))
    if p7 < 0 or mer < be: st = '🚨縮小・停止'
    elif (a3 and s3 / a3 < be) or mer < be + 0.30: st = '🔧改善'
    elif tg and mer >= tg and wu and wu >= 0.95: st = '🚀伸ばす候補'
    else: st = '✅維持'
    mm = mer * 0.7; per = mm * gm - 1
    ends = [mer * x * gm - 1 for x in (0.6, 0.7, 0.8)]
    rob = '一致' if (all(e > 0 for e in ends) or all(e < 0 for e in ends)) else '⚠️符号が割れる'
    SIMPLE.append([n, round(mer, 2), round(s3 / a3, 2) if a3 else '', round(s1 / a1, 2) if a1 else '',
        round(be, 2), round(tg, 2) if tg else '', round(p7), round(p3), round(p1),
        round(wu, 3) if wu else '', round(per, 3), rob, st, tri])
sh('シンプル判定', ['LOOTY 2026-09-05 定例（昨日=9/4 金）',
 '状態: 🚨縮小・停止=7日MER<分岐 or 7日利益マイナス ／ 🔧改善=3日MER<分岐 or 余裕<0.30 ／ 🚀伸ばす候補=7日MER≥目標かつ週消化率≥95% ／ ✅維持',
 '3窓判定（3日/5日/7日・すべて9/1終端）が本番の意思決定ルール。窓が1つでも割れたら据え置き＝動かさない',
 '「広告費1円あたり利益」= 限界MER×粗利率 − 1（限界MER = 平均MER×0.7）。頑健性は×0.6/0.7/0.8 の3端で符号が変わらないかを見る'],
 ['商品','7日MER','3日MER','前日MER','分岐','目標','7日利益','3日利益','前日利益','週消化率','1円あたり利益','頑健性','状態','3窓判定'],
 SIMPLE, [26,9,9,9,8,8,12,12,12,10,12,13,14,30], {7:'#,##0',8:'#,##0',9:'#,##0',10:'0.0%'})

# ---------- 2 ネクストアクション ----------
NA = [
 ['🎉 決着','9/4（金）は9月最悪の日。ただし黒字死守＋夜に回復シグナル',
  '実売372,740（前週金曜比−12.0%）/ 利益+20,130円（5.4%・手数料後+7,323円）/ MER1.45（分岐1.34）。'
  '全店CVR 2.25%で宣言済み−2σトリガーが初発動（処置のD1-D5診断は9/4昼に前倒し済み）。'
  '**20-24時の注文23件は週内最多** — 昼の巻き戻し4件のあと最初の時間帯で正常化し、増額希釈の解毒が効き始めた形。'
  '楽天セール本番と衝突した夜に沈まなかったため「楽天に吸われる」説は大幅に弱まった',
  '—','完了'],
 ['✅ 実行済み','偏光・調光サングラス キャンペーン停止（ユーザー実行・Meta直読でPAUSED確認）',
  '3日判定(9/2-4) MER1.18 < 分岐1.30 で宣言どおり停止。9/2の「復活」の正体は同日に作られた複製セットの学習初期配信だった（訂正済み）。'
  '空く5,000円/日は新商品テスト枠へ保留。再開するなら秋CR前提の「秋の再テスト」（decisions.md宣言）',
  '+約150円/日','完了'],
 ['🔴 判定執行','温感EMSワンド 中間判定 → ✅通過（10,000円で継続）',
  '3日CTR 2.38% ≥ 閾値2.2% ／ 倍率1.03 ≥ 0.70。2日目に黒字化した良い出だし。'
  '**本判定 9/9朝**: 7日倍率 ≥1.3 継続／1.0〜1.3 −50%／<1.0 停止',
  '—','9/9朝'],
 ['⏸ 延期','2WAYシートボックス 減額判定 → 9/7朝へ（理由明記）',
  'クリーン3日(9/2-4) 3,135円/日 < 基準3,913円で機械的には不合格。ただし不合格は全店−2σ異常日9/4（−644円）だけで決まり、'
  '9/2-3の2日平均は5,025円で基準超え。汚染窓での執行を避け9/5-6を加えた5日窓で9/7朝に判定（基準不変）',
  '—','9/7朝'],
 ['👀 判定日','バランスケアスリッパ 増額判定（15,000の実験）— 9/7朝',
  'クリーン3日(9/4-6)の1日利益 ≥ 8,055円。初日9/4は+7,178円で基準を878円下回るスタート（9/4は全店異常日である点を判定時に併記）',
  '—','9/7朝'],
 ['👀 監視','ナノガラスのCVR軟化（週次−17%）— 来週の週次診断で継続なら新CR検討',
  '打ち方は決定済み: 本体セットに足さず小さい別セット（8,000円/日）で受ける。今週は触らない',
  '—','9/8週'],
 ['📅 セール週の運用','楽天スーパーセール 9/4 20:00〜9/11 01:59',
  'セール期の悪化を理由に停止・減額を打たない（宣言済みガード）。きれいな答え合わせは**9/12-14の反発**。'
  '今日9/5(土)は解毒後最初のフル日: 先週土曜（8/29 実売583,139）にどこまで戻るかが本試験',
  '—','〜9/11'],
 ['🟡 推奨','秋テスト第1波の立ち上げ（candidates30-2026-09-04.txt）',
  '在庫あり2本（ニットトート=LP納品済み・1秒折り畳みチェア再テスト）は今日出せる。'
  'CJ問い合わせ（cj-inquiry-2026-09-04.txt 貼るだけ）でスカルプブラシ・靴乾燥機ほか第1波が来週並ぶ',
  '当たり1本=週6万円〜','今週'],
 ['🟡 推奨','購入者2万人へのメール導線（アイデア済み・GO待ち）',
  '9/12のセール明けに第1号を当てる設計。GOがあれば文面を納品する',
  '広告費ゼロの新チャネル','ユーザー判断'],
 ['🟡 推奨','優先配送536円へ戻す（9/4判定で不合格確定済み）','アップセルアプリ側の価格設定変更のみ','週+8,000円','実行待ち'],
 ['🟡 推奨','アップセル手動ペア5組の設定','data/lp/upsell-pairs-2026-08-26.txt。持ち越し10日目','週+約4万円','任意'],
]
sh('ネクストアクション', ['今日やること・判定カレンダー・持ち越しタスク（実施が確認できるまで毎日残す）'],
   ['区分','商品/対象','内容','効果','期限'], NA, [12, 26, 78, 16, 10])

# ---------- 3 全体サマリー ----------
SUMR = []
for lbl, N, K, ACC in [('前日 9/4(金)', R['N1'], R['K1'], R['ACCT1']), ('3日 9/2-4', R['N3'], R['K3'], R['ACCT3']),
        ('7日 8/29-9/4', R['N7'], R['K7'], R['ACCT7']), ('30日 8/06-9/04', R['N30'], R['K30'], R['ACCT30']),
        ('9月累積 9/01-04', R['NC'], R['KC'], R['ACCTC'])]:
    s = sum(N.values()); c = sum(K.values()); p = s - c - ACC
    SUMR.append([lbl, s, c, round(ACC), round(c + ACC), round((c + ACC) / s, 4), round(p), round(p / s, 4),
                 round(s / ACC, 2), round(1 / (1 - c / s), 2), round(s * FEE), round(p - s * FEE), round((p - s * FEE) / s, 4)])
sh('全体サマリー', ['全体サマリー（売上=Shopify gross−値引・広告費=Meta実測。返品は売上からも利益からも控除しない=A案）',
 '総合原価＝原価＋広告費。恒等式 総合原価率＋利益率＝100% が全窓で成立していることを確認済み',
 '縦照合3本すべて通過: Σ商品gross=全店381,300 ／ Σ値引=8,560 ／ Σキャンペーン広告費=257,125円（9/4・アカウントレベルと差0円）',
 '✅ 9/3の広告費は再取得で確定済み（264,897円・+138円）',
 '⚠️ 9/4の広告費は 9/5 00:08 の取得。Metaの確定まで +0.1%程度動きうるので、明朝に再取得して上書きする',
 '決済ブレンド率 **3.436%**（2026-09-01再計測。次回10/1）',
 f'7日の全店CVR（Shopifyセッション基準）= {STORE_ORD7}注文 ÷ {STORE_SESS7:,}セッション = {STORE_ORD7/STORE_SESS7:.2%}'],
 ['窓','実売','原価','広告費','総合原価','総合原価率','利益','利益率','MER','分岐MER','決済手数料','手数料控除後利益','手数料後利益率'],
 SUMR, [20,14,13,13,13,11,13,10,8,9,12,15,12],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'#,##0',6:'0.0%',7:'#,##0',8:'0.0%',9:'0.00',10:'0.00',11:'#,##0',12:'#,##0',13:'0.0%'})
ws = wb['全体サマリー']; r = ws.max_row + 3
ws.cell(r, 1, '■ 昨日(9/4 金) × 期間比較').font = Font(name='Arial', bold=True, size=11); r += 1
S1 = sum(R['N1'].values()); C1 = sum(R['K1'].values()); A1 = R['ACCT1']
per = [('実売', S1, sum(R['N3'].values())/3, sum(R['N7'].values())/7, sum(R['N30'].values())/30),
       ('原価', C1, sum(R['K3'].values())/3, sum(R['K7'].values())/7, sum(R['K30'].values())/30),
       ('広告費', A1, R['ACCT3']/3, R['ACCT7']/7, R['ACCT30']/30),
       ('利益', S1-C1-A1, (sum(R['N3'].values())-sum(R['K3'].values())-R['ACCT3'])/3,
        (sum(R['N7'].values())-sum(R['K7'].values())-R['ACCT7'])/7,
        (sum(R['N30'].values())-sum(R['K30'].values())-R['ACCT30'])/30)]
for c, v in enumerate(['指標','9/4(金)','3日平均/日','7日平均/日','30日平均/日','vs7日平均'], 1): ws.cell(r, c, v)
for c in range(1, 7):
    x = ws.cell(r, c); x.font = TH; x.fill = HEAD; x.border = thin; x.alignment = Alignment(horizontal='center')
r += 1
for lbl, d1, m3, m7, m30 in per:
    for c, v in enumerate([lbl, round(d1), round(m3), round(m7), round(m30), round(d1/m7-1, 4)], 1):
        x = ws.cell(r, c, v); x.border = thin; x.font = NEG if (isinstance(v, (int, float)) and v < 0) else TD
        if c in (2,3,4,5): x.number_format = '#,##0'
        if c == 6: x.number_format = '+0.0%;-0.0%'
    r += 1
ws.cell(r+1, 1, '9/4(金)は前週金曜(8/28 実売423,531)比 −12.0%・データセット最悪の日。分解は消化103.7%（金曜のペーシング押し込み）× CVR 2.25%（−2σ）。夜は週内最多で回復シグナルあり。単日で動かさない').font = Font(name='Arial', size=10, color='CC0000')

# ---------- 4 商品別 ----------
P = []
for n in sorted(R['N7'], key=lambda x: -R['N7'][x]):
    s7,k7,a7,p7 = blk(n,R['N7'],R['K7'],R['A7']); s1,k1,a1,p1 = blk(n,R['N1'],R['K1'],R['A1'])
    s3,k3,a3,p3 = blk(n,R['N3'],R['K3'],R['A3']); s30,k30,a30,p30 = blk(n,R['N30'],R['K30'],R['A30'])
    sc,kc,ac,pc = blk(n,R['NC'],R['KC'],R['AC'])
    gm = 1-k7/s7 if s7 else 0; mer = s7/a7 if a7 else None; be = 1/gm if gm else None
    tg = 1/(gm-0.35) if gm-0.35 > 0 else None; yo = (mer-be) if mer and be else None
    q7 = R['Q7'].get(n, 0); o = ORD.get(n); ipo = q7/o if o else None
    bcpa = (s7-k7)/q7*ipo if q7 and ipo else None; cpa = a7/o if o and a7 else None
    jud = ('📦広告なし' if not a7 else ('🚨赤字' if p7 < 0 else ('🔧テコ入れ' if yo is not None and yo < 0.30 else
           ('🚀増額候補' if tg and mer >= tg else '✅維持'))))
    P.append([n, '季節' if n in SEA else '通年', s7, k7, round(a7), round(p7), round(p7/s7,4) if s7 else '',
        round(mer,2) if mer else '—', round(be,2) if be else '—', round(tg,2) if tg else '—',
        round(yo,2) if yo is not None else '—', s1, round(p1), round(p1/s1,4) if s1 else '',
        round(p3), round(p30), round(pc), q7, o or '', round(ipo,2) if ipo else '',
        round(cpa) if cpa else '', round(bcpa) if bcpa else '', BUD.get(MAP.get(n,n), '—'), jud])
P.append(['カタログ全部（テスト）※全商品横断','—','—','—',round(R['CAT7'])]+['—']*18+[BUD.get('カタログ全部（テスト）','—'),'—'])
sh('商品別', ['商品別（7日 8/29-9/4 = 判定単位）。カタログ行の広告費は「7日合計」（前日ではない）',
  '分岐CPA(注文) = 1個あたり粗利 × 点数/注文。まとめ買い商品（点数/注文>1.2）は必ず注文ベースで比較する',
  'Σ商品広告費 + カタログ = Metaアカウント7日消化（差0円・w0901.pyでassert済み）',
  '当月累積利益は9月分（9/01-04）',
  '★耳かき(9/1停止)・サングラス(9/5停止・直読確認済み)は7日窓に残消化が入っている'],
 ['商品','区分','7日売上','7日原価','7日広告','7日利益','7日利益率','MER','分岐','目標','余裕','前日売上','前日利益','前日利益率',
  '3日利益','30日利益','当月累積利益','7日販売数','7日注文数','点数/注文','CPA(注文)','分岐CPA(注文)','現日予算','判定'],
 P, [26,6]+[12]*22, {3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',12:'#,##0',13:'#,##0',14:'0.0%',
                     15:'#,##0',16:'#,##0',17:'#,##0',21:'#,##0',22:'#,##0',23:'#,##0'})

# ---------- 5 レバー一覧 ----------
LV = []
for n in sorted(R['N7'], key=lambda x: -R['N7'][x]):
    cp = MAP.get(n, n); b = BUD.get(cp)
    if not b: continue
    s7, k7, a7, p7 = blk(n, R['N7'], R['K7'], R['A7'])
    if not a7: continue
    gm = 1-k7/s7; mer = s7/a7; mm = mer*0.7
    days = [d for d in D7 if d in BSNAP[cp]]
    wu = (a7/sum(BSNAP[cp][d] for d in days)) if len(days) == 7 and sum(BSNAP[cp][d] for d in days) else None
    ends = [mer*x*gm-1 for x in (0.6, 0.7, 0.8)]
    LV.append([n, '季節' if n in SEA else '通年', b, round(wu,3) if wu else '', round(mer,2), round(1/gm,2),
        round(mer-1/gm,2), round(mm*gm-1,3), round(ends[0],3), round(ends[2],3),
        '一致' if (all(e>0 for e in ends) or all(e<0 for e in ends)) else '⚠️符号が割れる',
        round((1-mm*gm)*b*0.25*7), round((mm*gm-1)*b*0.25*7), round(p7)])
LV.sort(key=lambda x: -max(x[11], x[12]))
sh('レバー一覧', ['「どれを動かすと一番効くか」は7日平均の利益額ではなく、動かしたときの変化量で並べる',
  '削減1円あたりの利益変化 = 1 − 限界MER×粗利率 ／ 増額1円あたり = 限界MER×粗利率 − 1（限界MER = 平均MER×0.7）',
  '★頑健性: ×0.6 と ×0.8 の両端で符号が変わらないものだけ動かしてよい。割れたら据え置き',
  '増額は週消化率95%以上のものだけが対象（それ未満は増やしても使われない）'],
 ['商品','区分','日予算','週消化率','MER','分岐','余裕','1円あたり利益(×0.7)','×0.6端','×0.8端','頑健性','−25%で週','+25%で週','7日利益'],
 LV, [26,6,11,10,8,8,8,15,10,10,13,12,12,12], {3:'#,##0',4:'0.0%',12:'#,##0',13:'#,##0',14:'#,##0'})

# ---------- 6 週次診断 ----------
def wblk(days, names=None):
    s = c = 0
    for n, v in Sd.items():
        if names is not None and n not in names: continue
        g = sum(v[d][0] for d in days if d in v); dc = sum(v[d][1] for d in days if d in v)
        if not g: continue
        s += g-dc
        if n in PRICEa and n in COSTa: c += g/PRICEa[n]*COSTa[n]
    a = sum(ADd[MAP.get(n, n)].get(d, 0) for n in (names if names is not None else Sd) for d in days) if names is not None \
        else sum(ADd[cp].get(d, 0) for cp in ADd for d in days)
    return s, c, a, s-c-a
YR = {n for n in Sd if n and n not in SEA}
end = datetime.date(2026, 8, 30); WK = []
for i in range(6):
    a = end - datetime.timedelta(days=7*i+6)
    WK.append((f"{a.strftime('%m/%d')}-{(a+datetime.timedelta(days=6)).strftime('%m/%d')}",
               [(a+datetime.timedelta(days=j)).isoformat() for j in range(7)]))
WK.reverse()
WD_ = []
for lbl, days in WK:
    if days[0] < '2026-06-29': continue
    s, c, a, p = wblk(days); ss, _, sa, sp = wblk(days, SEA); ys, _, ya, yp = wblk(days, YR)
    WD_.append([lbl, round(s), round(c), round(a), round(p), round(p/s,4), round(s/a,2), round(c/s,4), round(a/s,4),
                round(sp), round(ss/sa,2) if sa else '', round(yp), round(ys/ya,2) if ya else ''])
sh('週次診断', ['全店の週次分解（完全週のみ・最新は8/24-30週）。原価はナノガラス/ディスペンサーのみ加重平均単価【近似】',
  '★夏物の崩壊が8月前半の正体。壊れたのは季節物だけで、通年物へ入れ替えて回復した',
  '★8月後半は日予算244,000〜255,000で運転。日予算590,000だった7月とは水準が違うので、利益額の絶対比較はしない'],
 ['週','実売','原価','広告費','利益','利益率','MER','原価率','広告費率','季節物 利益','季節物MER','通年物 利益','通年物MER'],
 WD_, [14,13,12,12,12,9,8,9,10,15,11,15,11],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'#,##0',6:'0.0%',7:'0.00',8:'0.0%',9:'0.0%',10:'#,##0',11:'0.00',12:'#,##0',13:'0.00'})

# ---------- 7 ファネル（30日・RPM/CPM つき）----------
for _k, _v in MF_RAW.items():
    _c = sum(ADd[_k].get(d, 0) for d in D30)
    assert abs(_c - _v[0]) <= max(50, _v[0] * 0.0005), (_k, _c, _v[0])
FN = []
for n, (c, imp, clk) in MF.items():
    s = R['N30'].get(n, 0); k = R['K30'].get(n, 0); q = R['Q30'].get(n, 0); o = ORD30.get(n, 0)
    if not s or not o or not clk: continue
    ipo = q/o; bcpa = (s-k)/q*ipo; cpm = c/imp*1000
    ctr = clk/imp; cvr = o/clk; gpo = (s-k)/o
    need = cpm/(gpo*1000)
    FN.append([n, '季節' if n in SEA else '通年', imp, round(ctr,4), clk, round(cvr,4), round(s/clk),
               round(c/clk), round(c/o), round(bcpa), round(bcpa-c/o), round(cpm), round(gpo),
               round(ctr*cvr*10000, 2), round(need*10000, 2), round((ctr*cvr)/need, 2)])
mc = statistics.median([x[5] for x in FN]) if FN else 0; mr = statistics.median([x[6] for x in FN]) if FN else 0
FN.sort(key=lambda x: -x[15])
sh('ファネル30日', ['⚠️ このシートだけ窓が「30日(8/06-9/04)」。他シートの7日窓と混ぜて読まないこと',
  'クリックは **アウトバウンドクリック**（data_query 実測・Asia/Tokyo）',
  '✅ 検算: 全30キャンペーンで data_query の消化 = CSV(8/06-9/04) が許容差以内で一致（コード内assert）',
  'CVR = Shopify注文 ÷ Meta全クリック（Metaの購入数は損益に使わない）',
  f'中央値 CVR {mc*100:.2f}% ／ 売上perクリック {mr:.0f}円',
  '★RPM/CPM = CTR×CVR×粗利/注文 ÷ CPM。1.0未満が赤字圏。★CTR単独では成否を判別できない（実測で反証済み）',
  '★必要CTR×CVR(‱) = CPM ÷ (粗利/注文 × 1000)。この列を下回っている商品が赤字'],
 ['商品','区分','インプ','CTR','クリック','CVR','売上/クリック','CPC','CPA(注文)','分岐CPA(注文)','余裕','CPM','粗利/注文',
  '実測CTR×CVR(‱)','必要CTR×CVR(‱)','RPM/CPM'],
 FN, [26,6,12,9,11,9,13,9,12,15,11,10,11,15,15,10],
 {3:'#,##0',4:'0.00%',5:'#,##0',6:'0.00%',7:'#,##0',8:'#,##0',9:'#,##0',10:'#,##0',11:'#,##0',12:'#,##0',13:'#,##0'})

# ---------- 8 日次推移 ----------
WDJ = ['月','火','水','木','金','土','日']
pr = {}
for d in ALLD:
    s = c = 0
    for n, v in Sd.items():
        if n == '' or d not in v: continue
        g, dc = v[d]; s += g-dc
        if g and n in PRICEa and n in COSTa: c += g/PRICEa[n]*COSTa[n]
    a = sum(ADd[cp].get(d, 0) for cp in ADd); pr[d] = (s, c, a, s-c-a)
DD = []
for j, d in enumerate(ALLD):
    if d < '2026-07-13': continue
    s, c, a, p = pr[d]; ma = sum(pr[x][3] for x in ALLD[j-2:j+1])/3
    DD.append([d, WDJ[datetime.date.fromisoformat(d).weekday()], round(s), round(c), round(a), round(p),
               round(p/s,4), round(s/a,2), round(ma), '⚠️非常ブレーキ' if ma < 130000 else ''])
sh('日次推移', ['日次の利益は上記【近似】原価ベース。4期間サマリーはバリアント別実数量の正確値',
  '非常ブレーキ: 3日移動平均利益 < 130,000円 で発動（⚠️ 現体制244,000円/日に対して基準が古い。再校正を提案中 — ネクストアクション参照）'],
 ['日付','曜','実売','原価','広告費','利益','利益率','MER','3日移動平均利益','警告'],
 DD, [12,5,13,12,12,12,9,8,16,16], {3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',8:'0.00',9:'#,##0'})

# ---------- 9 本日の判定 ----------
JUDGE = [
 ['9/5','温感EMSワンド 中間（3日）','リンクCTR<2.2%即停止 ／ 倍率<0.70停止 ／ クリック100未満なら延期',
  '3日CTR (11+58+76)÷(886+2319+2877) = **2.38%** ／ 倍率 = 粗利3,901 ÷ CPA(個)3,791 = **1.03** ／ クリック145',
  '✅ 通過 — 10,000円で継続','本判定 9/9朝（7日倍率 ≥1.3継続／1.0-1.3で−50%／<1.0停止）'],
 ['9/5','偏光・調光サングラス 3日（9/2-4）','MER ≥ 分岐1.30 → 継続 ／ 割れ → 停止',
  '実売15,920 ÷ 広告13,523 = **MER 1.18**（9/4は消化4,995で売上ゼロ）',
  '❌ 停止 — ユーザー実行済み（Meta直読でPAUSED確認）',
  '配信死は8/30〜でセール前から。セール週ガードの対象外。空く5,000円は新商品枠へ'],
 ['9/5→9/7','2WAYシートボックス 減額判定','クリーン3日(9/2-4)の1日利益 ≥ 3,913円',
  '+3,211 / +6,840 / **−644**（9/4=全店−2σ異常日）= 3,135円/日。9/2-3の2日平均は5,025円で基準超え',
  '⏸ 9/7朝へ延期（汚染窓での執行を回避・基準不変）',''],
 ['9/7','バランスケアスリッパ 増額判定（15,000）','クリーン3日(9/4-6)の1日利益 ≥ 8,055円（12,000時代のP2実測）',
  '初日9/4: **+7,178円**（基準−878円。全店異常日）','進行中','不合格なら12,000へ戻す'],
 ['9/9','温感EMSワンド 本判定（7日）','倍率 ≥1.3 継続 ／ 1.0〜1.3 −50% ／ <1.0 停止','—','—',''],
 ['9/8週','ナノガラス CVR軟化の継続確認','週次CVRが2週連続−15%なら新CR（小さい別セット8,000円で受ける）',
  '直近週 −17%（消化フラットでインプ+14%・CPM−15%＝Meta側のリーチ拡大）','監視中',''],
 ['9/12-14','セール明けの反発確認','同曜日比が平常圏へ戻るか',
  '戻る→セール吸引が確定・平常運転へ ／ 戻らない→構造問題として予算総量の議論に進む','待機',''],
]
sh('本日の判定', ['宣言済みの基準に当てはめるだけ。基準は宣言時のまま動かさない',
  '★本日の執行: EMSワンド中間=✅通過 ／ サングラス3日=❌停止（実行済み）／ 2WAY=9/7へ延期',
  '★楽天スーパーセール開催中（〜9/11）: セール起因の悪化で停止・減額を打たない'],
 ['判定日','対象','宣言済み基準','実測','結果','備考'], JUDGE, [10,26,34,34,16,46])

# ---------- 9a 予算見直しチェック（9/1 → 9/2 の差分）----------
PREVB = collections.defaultdict(int)
for r in body:
    if r[0] == '2026-09-04b': PREVB[r[1]] += int(r[3])
BR, inc, dec = [], 0, 0
for cp in sorted(BUD, key=lambda c: -BUD[c]):
    p, nn = PREVB.get(cp, 0), BUD[cp]
    n = INV.get(cp, cp)
    s7 = R['N7'].get(n, 0); k7 = R['K7'].get(n, 0); a7 = R['A7'].get(n, 0)
    s3 = R['N3'].get(n, 0); k3 = R['K3'].get(n, 0); a3 = R['A3'].get(n, 0)
    q7 = R['Q7'].get(n, 0); q3 = R['Q3'].get(n, 0)
    gm = 1 - k7/s7 if s7 else None
    mer = s7/a7 if s7 and a7 else None; be = 1/gm if gm else None
    tg = 1/(gm-0.35) if gm and gm-0.35 > 0 else None
    m3 = s3/a3 if s3 and a3 else None; b3_ = 1/(1-k3/s3) if s3 else None
    r7 = ((s7-k7)/q7)/(a7/q7) if q7 and a7 else None
    r3 = ((s3-k3)/q3)/(a3/q3) if q3 and a3 else None
    dd = [d for d in D7 if BSNAP[cp].get(d)]
    w = (sum(ADd[cp].get(d, 0) for d in dd) / sum(BSNAP[cp][d] for d in dd)) if dd else None
    per = (mer*0.7*gm - 1) if mer and gm else None
    e6 = (mer*0.6*gm - 1) if mer and gm else None
    e8 = (mer*0.8*gm - 1) if mer and gm else None
    if nn > p: inc += nn - p
    elif nn < p: dec += p - nn
    BR.append([n if n != cp else cp, p, nn, nn-p, (nn/p-1) if p else '',
        round(mer, 2) if mer else '—', round(be, 2) if be else '—', round(tg, 2) if tg else '—',
        round(mer-be, 2) if mer and be else '—', round(m3, 2) if m3 else '—',
        round(m3-b3_, 2) if m3 and b3_ else '—', round(r7, 2) if r7 else '—', round(r3, 2) if r3 else '—',
        round(w, 3) if w else '—', round(per, 3) if per is not None else '—',
        ('一致' if (e6 and e8 and ((e6 > 0 and e8 > 0) or (e6 < 0 and e8 < 0))) else '⚠️符号が割れる') if per is not None else '—',
        '★変更' if p != nn else ''])
BR.sort(key=lambda x: (x[16] == '', -abs(x[3]), -x[2]))
sh('予算見直しチェック', ['前日(9/4昼・9/4b)スナップショット → 本日(9/5)',
  '合計 248,000 → **243,000円/日**。変更は1件: **偏光・調光サングラス 停止**（3日判定❌の執行・Meta直読でPAUSED確認）',
  '現行の実験は スリッパ15,000（9/7判定）と EMSワンド10,000（9/9本判定）の2本のみ',
  '⚠️ 楽天セール中（〜9/11）は新たな減額・停止を打たない（宣言済みガード）'],
 ['商品/キャンペーン','変更前','変更後','差','変化率','7日MER','分岐','目標','7日余裕','3日MER','3日余裕',
  '倍率7日','倍率3日','消化率(7日)','1円あたり利益','頑健性','印'],
 BR, [26,10,10,9,9,8,7,7,9,8,9,9,9,12,13,14,7],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'+0.0%;-0.0%',14:'0.0%'})

# ---------- 10 CVR推移 ----------
SES = {'2026-08-04':6639,'2026-08-05':5425,'2026-08-06':5138,'2026-08-07':4827,'2026-08-08':5527,'2026-08-09':5886,
'2026-08-10':5245,'2026-08-11':6001,'2026-08-12':4900,'2026-08-13':5126,'2026-08-14':5035,'2026-08-15':4871,
'2026-08-16':5349,'2026-08-17':3438,'2026-08-18':3523,'2026-08-19':3190,'2026-08-20':3356,'2026-08-21':3120,
'2026-08-22':3220,'2026-08-23':3438,'2026-08-24':3192,'2026-08-25':3280,'2026-08-26':2733,'2026-08-27':3145,
'2026-08-28':3149,'2026-08-29':3537,'2026-08-30':3310,'2026-08-31':2761,'2026-09-01':2936,'2026-09-02':3142,'2026-09-03':3340,'2026-09-04':3293}
ORDD = {'2026-08-04':188,'2026-08-05':152,'2026-08-06':157,'2026-08-07':173,'2026-08-08':184,'2026-08-09':187,
'2026-08-10':141,'2026-08-11':176,'2026-08-12':124,'2026-08-13':125,'2026-08-14':120,'2026-08-15':142,
'2026-08-16':143,'2026-08-17':98,'2026-08-18':89,'2026-08-19':106,'2026-08-20':113,'2026-08-21':110,
'2026-08-22':98,'2026-08-23':106,'2026-08-24':100,'2026-08-25':99,'2026-08-26':105,'2026-08-27':130,
'2026-08-28':85,'2026-08-29':115,'2026-08-30':114,'2026-08-31':84,'2026-09-01':88,'2026-09-02':92,'2026-09-03':103,'2026-09-04':74}
CV = [[d, WDJ[datetime.date.fromisoformat(d).weekday()], SES[d], ORDD[d], round(ORDD[d]/SES[d], 4),
       ('★昨日（確定・CVR2.25%=−2σトリガー初発動）' if d == '2026-09-04' else '')] for d in sorted(SES)]
sh('CVR推移', ['全店CVR = Shopify注文 ÷ Shopifyセッション。商品別レポートのCVR（注文÷Metaクリック）とは分母が違うので混ぜないこと',
  '★9/4は セッション3,293・注文74件 = **CVR 2.25%**（−2σ）。宣言済みトリガー初発動 → D診断は9/4昼に前倒し実施済み',
  '★セッション2,700〜3,500/日 は日予算253,000→244,000円体制の水準。日予算590,000だった8月上旬(5,000〜6,600)と比べない'],
 ['日付','曜','セッション','注文','全店CVR','備考'], CV, [12,5,12,10,10,20],
 {3:'#,##0',4:'#,##0',5:'0.00%'})

# ---------- 11 曜日指数 ----------
_ix = sorted(R['IDX'].items(), key=lambda x: -x[1])
_hi, _lo = _ix[0], _ix[-1]
sh('曜日指数', ['直近30日(8/06-9/04)・祝日 8/11(山の日) を除外。売上=gross−値引。次回の定期更新は 9/7(月)',
  '★同曜日平均との比較は当面使わない（8月上旬の高予算期が窓に混ざるため）。「前週同曜日比」で読む',
  '★9/4(金)は前週金曜(8/28 423,531円)比 −12.0%。楽天スーパーセール前日〜初日の特異日として注記',
  f'★{_hi[0]}曜が最強({_hi[1]:.1f})・{_lo[0]}曜が最弱({_lo[1]:.1f})。中位の曜日は窓次第で入れ替わるので差を根拠に判断しない'],
 ['曜日','指数(全体=100)'], [[k, round(v, 1)] for k, v in _ix], [10, 16])

wb.save('data/reports/report-2026-09-05.xlsx')
print('シート:', wb.sheetnames)
print(f"前日利益 {round(_pl[-1][3]):,} / 3日移動平均 {p3ma:,.0f}円 / 8月確定 {aup:,.0f}円")
