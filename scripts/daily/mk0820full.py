# -*- coding: utf-8 -*-
"""2026-08-18 定例レポート本体（昨日=8/17 月）。"""
import pickle, csv, datetime, collections, statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
R=pickle.load(open('/tmp/rep0819w.pkl','rb')); FEE=0.03452
D1,D3,D7,D30,CUM=R['D1'],R['D3'],R['D7'],R['D30'],R['CUM']
# --- 予算スナップショット（8/18・Meta直読。ヘアドライタオル停止、ナノガラス 80,000→74,000へリバランス）
rows=list(csv.reader(open('data/budget-snapshots.csv'))); hdr=rows[0]; body=[r for r in rows[1:] if r]
assert any(r[0]=='2026-08-18' for r in body), '8/18スナップショット未記録'
BUD=collections.defaultdict(int)
for r in body:
    if r[0]=='2026-08-18': BUD[r[1]]+=int(r[3])
BSNAP=collections.defaultdict(lambda: collections.defaultdict(int))
for r in body: BSNAP[r[1]][r[0]]+=int(r[3])
# --- Meta 7日ファネル実測（8/11-17, キャンペーン別）
MF={}   # 7日ファネルはチャット本文で提示（Excelは4期間の断面＋商品タブ）
# 7日(8/11-17)の注文数 = ShopifyQL実測（FROM sales SHOW orders GROUP BY product_title）
ORD={'ナノガラス脱毛パッド':281,'W固定スマホ車載ホルダー':138,'形状記憶日傘':65,'接触冷感UVパーカー':57,
'害虫ブロッカー':53,'ムダ毛シェーバー':48,'偏光・調光サングラス':43,'2WAYシートボックス':40,
'快適マジックインソール':36,'4WAY':32,'完全遮光・接触冷感UVハット':27,'完全遮光・形状記憶':26,
'姿勢サポートチェア':24,'接触冷感UVアームカバー':21,'優先配送':19,'バランスケアスリッパ':15,
'5WAY腰掛けファン':13,'携帯電動シェーバー':8,'健康サンダル':2,'卓上冷感クーラー':2,'ヘアドライタオル':2,
'瞬間冷感ポンチョ':2,'ジェットウォッシャー':1,'3WAYサーキュレーター':1,'ナノバブルシャワーヘッド':1,
'壁掛けディスペンサー':1,'姿勢サポートベルト':1}
MAP=R['MAP']; INV={v:k for k,v in MAP.items()}
SEA={'4WAY','3WAYサーキュレーター','卓上冷感クーラー','5WAY腰掛けファン','瞬間冷却ハンディファン','接触冷感UVパーカー',
'接触冷感UVアームカバー','瞬間冷感ポンチョ','完全遮光・接触冷感UVハット','形状記憶日傘','完全遮光・形状記憶','偏光・調光サングラス','害虫ブロッカー'}
wb=Workbook(); wb.remove(wb.active)
TH=Font(name='Arial',bold=True,color='FFFFFF',size=10); TD=Font(name='Arial',size=10)
NEG=Font(name='Arial',size=10,color='CC0000'); HEAD=PatternFill('solid',fgColor='305496')
GREEN=PatternFill('solid',fgColor='D6EFD8'); RED=PatternFill('solid',fgColor='FADBD8'); YEL=PatternFill('solid',fgColor='FFF2A8')
thin=Border(*[Side(style='thin',color='CCCCCC')]*4)
def sh(title,note,cols,data,widths,fmt=None,fills=None):
    ws=wb.create_sheet(title); r=1
    if note:
        for ln in note: ws.cell(r,1,ln).font=Font(name='Arial',size=10,bold=(r==1)); r+=1
        r+=1
    for c,v in enumerate(cols,1): ws.cell(r,c,v)
    for c in range(1,len(cols)+1):
        x=ws.cell(r,c); x.font=TH; x.fill=HEAD; x.border=thin
        x.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.row_dimensions[r].height=30; hr=r; r+=1
    for row in data:
        for c,v in enumerate(row,1):
            x=ws.cell(r,c,v); x.border=thin; x.font=TD
            if isinstance(v,(int,float)) and v<0: x.font=NEG
            if fmt and c in fmt: x.number_format=fmt[c]
        if fills: fills(ws,r,row)
        r+=1
    for i,w_ in enumerate(widths,1):
        ws.column_dimensions[chr(64+i) if i<27 else 'A'+chr(38+i)].width=w_
    ws.freeze_panes=ws.cell(hr+1,1).coordinate
    return ws
# ---------- 1 シンプル判定 ----------
def blk(n,N,K,A_): 
    s=N.get(n,0); k=K.get(n,0); a=A_.get(n,0); return s,k,a,s-k-a
SIMPLE=[]
for n in sorted(R['N7'],key=lambda x:-R['N7'][x]):
    a7=R['A7'].get(n,0)
    if not a7: continue
    s7,k7,_,p7=blk(n,R['N7'],R['K7'],R['A7']); s3,k3,a3,p3=blk(n,R['N3'],R['K3'],R['A3']); s1,k1,a1,p1=blk(n,R['N1'],R['K1'],R['A1'])
    gm=1-k7/s7; be=1/gm; tg=1/(gm-0.35) if gm-0.35>0 else None; mer=s7/a7
    cp=MAP.get(n,n); days=[d for d in D7 if d in BSNAP[cp]]
    wu=(a7/sum(BSNAP[cp][d] for d in days)) if len(days)==7 and sum(BSNAP[cp][d] for d in days) else None
    if p7<0 or mer<be: st='🚨縮小・停止'
    elif (a3 and s3/a3<be) or mer<be+0.30: st='🔧改善'
    elif tg and mer>=tg and wu and wu>=0.95: st='🚀伸ばす候補'
    else: st='✅維持'
    mm=mer*0.7; per=mm*gm-1
    if per<0: act='−25%（週+'+f"{(1-mm*gm)*BUD.get(cp,0)*0.25*7:,.0f}"+'円）'
    elif wu and wu>=0.95: act='+25%（週+'+f"{per*BUD.get(cp,0)*0.25*7:,.0f}"+'円）'
    else: act='据え置き（週消化率が95%未満＝増やしても使われない）'
    SIMPLE.append([n,round(mer,2),round(s3/a3,2) if a3 else '',round(s1/a1,2) if a1 else '',round(be,2),
        round(tg,2) if tg else '',round(p7),round(p3),round(p1),round(wu,3) if wu else '',round(per,3),st,act])
sh('シンプル判定',['LOOTY 2026-08-18 定例（昨日=8/17 月）',
 '状態: 🚨縮小・停止=7日MER<分岐 or 7日利益マイナス ／ 🔧改善=3日MER<分岐 or 余裕<0.30 ／ 🚀伸ばす候補=7日MER≥目標かつ週消化率≥95% ／ ✅維持',
 '「広告費1円あたり利益」= 限界MER×粗利率 − 1（限界MER = 平均MER×0.7、実測レンジ0.6〜0.8の中央）。正なら出すほど利益額が増える'],
 ['商品','7日MER','3日MER','前日MER','分岐','目標','7日利益','3日利益','前日利益','週消化率','1円あたり利益','状態','今やること'],
 SIMPLE,[26,9,9,9,8,8,12,12,12,10,12,14,26],{7:'#,##0',8:'#,##0',9:'#,##0',10:'0.0%'})
# ---------- 2 ネクストアクション ----------
NA=[['🔴 本日実行','形状記憶日傘 ＋ 完全遮光日傘','宣言済みの「日傘2本を1つの予算とみなした3窓判定」に該当。両方とも3窓すべて余裕<0.30（形状記憶 +0.20/+0.09/+0.14 ／ 完全遮光 −0.38/−0.09/+0.16）、頑健性も3端一致。形状記憶 35,000→26,250 ／ 完全遮光 10,000→7,500','週+約24,000円','本日'],
    ['✅ 判定済','日傘の統合判定（どちらかを停止）','宣言済み基準は「広告費1円あたり利益を3窓で比較し、3窓すべてで低いほうを停止」。実測は 3日 長−0.289/短+0.135、5日 長−0.081/短+0.054、7日 長+0.097/短+0.089。7日だけ短が低いので**3窓は揃わない → 据え置き（どちらも止めない）**','—','完了'],
    ['✅ 判定済','高吸水・速乾ヘアドライタオル','ユーザーが8/17に停止。宣言済みの中間判定（②CPM≥3,100円で即停止）に照らすと2日累計CPM 3,948円で**基準どおりの停止**。2日で広告10,664円・クリック74・注文2件。8/17は6,571円で注文ゼロ、CPM 4,522円','—','完了'],
    ['✅ 判定済','ナノガラス脱毛パッドのリバランス','ユーザーが8/18に実施。セット２ 63,000→54,000 ／ セット３（BC検証用）17,000→20,000。合計 80,000→74,000（−6,000）。Meta直読で確認。方向は正しい（セット３は7/29-8/02に週消化率113%で天井に張り付いていた側）','—','完了'],
    ['👀 様子見','接触冷感UVパーカー','3窓すべて分岐1.56割れ（1.10/1.06/1.36）・7日 −20,186円で**稼働中で唯一の赤字**。頑健性3端一致（+0.302〜+0.477）。ただし8/17に23,000→17,250（−25%）を実施したばかりで3クリーン日ロック中。判定日は8/20','週+約13,000円','8/20'],
    ['👀 様子見','快適マジックインソール','3窓すべて目標2.08超（2.65/2.45/2.70）で+25%の条件を満たすが、**週消化率90.5%**で増額の前提（95%以上）に4.5pt足りない。7日利益+85,729円・利益率44.9%は店内最高','週+約16,000円','8/19'],
    ['📌 記録','W固定スマホ車載ホルダー','週次ランプ再判定。W-1(8/04-10) MER 3.70 → W0(8/11-17) 2.50、Δ広告費+98,177円に対しΔ売上+71,242円で**増分MER 0.73＝分岐1.42割れ**。ただしW-1自体がランプ中の週で比較が汚染されている。経験則（平均2.50×0.6/0.7/0.8）だと削減1円あたりは−0.055/−0.230/−0.406で「削ると損」。**2手法が別ゾーンなので判定保留・据え置き**','—','据え置き'],
    ['🟡 推奨','新商品テスト3枠の立ち上げ','24,000円/日。原資はヘアドライタオル停止(8,000)＋ナノガラスのリバランス(6,000)＋夏物の縮小分。1枠目は在庫にあるスタイルアップインナー（秋冬本番・判定が終わっていない）。候補リストは data/lp/newproduct-2026-08-17.txt','8週で期待+116,000円/枠','本日'],
    ['🟡 推奨','優先配送のアタッチ率','7日 19件/全注文943件 = 2.01%。10%まで上げると週+約40,000円（原価0なので全額が利益）。商品ページ590円 vs 実課金536円の不一致が未解消','週+約40,000円','本日'],
    ['👀 様子見','非常ブレーキ','3日移動平均利益 136,840円。発動ライン130,000円までの余裕は6,840円しかない。割ったら カタログ23,000→10,000 ／ 赤字商品を一律−25% を即実行','—','毎日'],
    ['👀 様子見','安全弁','全店MERが実効分岐1.42を2日連続で下回ったら赤字商品を一律−25%／単日1.20割れで即停止。8/16 1.87 → 8/17 1.79（いずれも上回っている）','—','毎日']]
sh('ネクストアクション',['今日やること・判定カレンダー・持ち越しタスク（実施が確認できるまで毎日残す）'],
   ['区分','商品/対象','内容','効果','期限'],NA,[12,26,58,20,10])
# ---------- 3 全体サマリー ----------
SUMR=[]
for lbl,N,K,ACC in [('前日 8/17(月)',R['N1'],R['K1'],R['ACCT1']),('3日 8/15-17',R['N3'],R['K3'],R['ACCT3']),
        ('7日 8/11-17',R['N7'],R['K7'],R['ACCT7']),('30日 7/19-8/17',R['N30'],R['K30'],R['ACCT30']),
        ('★当月累積 8/01-17',R['NC'],R['KC'],R['ACCTC'])]:
    s=sum(N.values()); c=sum(K.values()); p=s-c-ACC
    SUMR.append([lbl,s,c,round(ACC),round(c+ACC),round((c+ACC)/s,4),round(p),round(p/s,4),round(s/ACC,2),
                 round(1/(1-c/s),2),round(s*FEE),round(p-s*FEE),round((p-s*FEE)/s,4)])
sh('全体サマリー',['全体サマリー（売上=Shopify gross−値引・広告費=Meta実測。返品は売上からも利益からも控除しない=A案）',
 '総合原価＝原価＋広告費。恒等式 総合原価率＋利益率＝100% が全窓で成立していることを確認済み',
 '当月累積 8/01-17: Shopify実測。返品・注文数は下の商品別シートを参照',
 '決済ブレンド率 3.452%（2026-08-01再計測: SP76.2%×3.25% ＋ KOMOJUスマホ16.3%×4.1% ＋ Paidy7.5%×4.1%）'],
 ['窓','実売','原価','広告費','総合原価','総合原価率','利益','利益率','MER','分岐MER','決済手数料','手数料控除後利益','手数料後利益率'],
 SUMR,[18,14,13,13,13,11,13,10,8,9,12,15,12],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'#,##0',6:'0.0%',7:'#,##0',8:'0.0%',9:'0.00',10:'0.00',11:'#,##0',12:'#,##0',13:'0.0%'})
# 昨日×期間比較
ws=wb['全体サマリー']; r=ws.max_row+3
ws.cell(r,1,'■ 昨日(8/17 月) × 期間比較').font=Font(name='Arial',bold=True,size=11); r+=1
S1=sum(R['N1'].values()); C1=sum(R['K1'].values()); A1=R['ACCT1']
per=[('実売',S1,sum(R['N3'].values())/3,sum(R['N7'].values())/7,sum(R['N30'].values())/30),
     ('原価',C1,sum(R['K3'].values())/3,sum(R['K7'].values())/7,sum(R['K30'].values())/30),
     ('広告費',A1,R['ACCT3']/3,R['ACCT7']/7,R['ACCT30']/30),
     ('利益',S1-C1-A1,(sum(R['N3'].values())-sum(R['K3'].values())-R['ACCT3'])/3,
      (sum(R['N7'].values())-sum(R['K7'].values())-R['ACCT7'])/7,
      (sum(R['N30'].values())-sum(R['K30'].values())-R['ACCT30'])/30)]
for c,v in enumerate(['指標','8/17(月)','非祝日の月曜平均','3日平均/日','7日平均/日','30日平均/日','vs7日平均'],1): ws.cell(r,c,v)
for c in range(1,8):
    x=ws.cell(r,c); x.font=TH; x.fill=HEAD; x.border=thin; x.alignment=Alignment(horizontal='center')
r+=1
for lbl,d1,m3,m7,m30 in per:
    mo=971627 if lbl=='実売' else ''
    for c,v in enumerate([lbl,round(d1),mo,round(m3),round(m7),round(m30),round(d1/m7-1,4)],1):
        x=ws.cell(r,c,v); x.border=thin; x.font=NEG if (isinstance(v,(int,float)) and v<0) else TD
        if c in (2,3,4,5,6): x.number_format='#,##0'
        if c==7: x.number_format='+0.0%;-0.0%'
    r+=1
ws.cell(r+1,1,'対同曜日（非祝日の月曜平均 971,627円）= 50.7%。ただし同曜日平均は日予算50万円台の時期を含むため水準比較に使えない。'
  '月曜どうしのMERは 7/27 2.09 → 8/03 2.06 → 8/10 1.83 → 8/17 1.79 で、効率はほぼ横ばい。'
  '3日移動平均利益 136,840円 > 非常ブレーキ130,000円 → 未発動（余裕は6,840円）').font=Font(name='Arial',size=10,color='CC0000')
# ---------- 4 商品別 ----------
P=[]
for n in sorted(R['N7'],key=lambda x:-R['N7'][x]):
    s7,k7,a7,p7=blk(n,R['N7'],R['K7'],R['A7']); s1,k1,a1,p1=blk(n,R['N1'],R['K1'],R['A1'])
    s3,k3,a3,p3=blk(n,R['N3'],R['K3'],R['A3']); s30,k30,a30,p30=blk(n,R['N30'],R['K30'],R['A30'])
    sc,kc,ac,pc=blk(n,R['NC'],R['KC'],R['AC'])
    gm=1-k7/s7 if s7 else 0; mer=s7/a7 if a7 else None; be=1/gm if gm else None
    tg=1/(gm-0.35) if gm-0.35>0 else None; yo=(mer-be) if mer and be else None
    q7=R['Q7'].get(n,0); o=ORD.get(n); ipo=q7/o if o else None
    bcpa=(s7-k7)/q7*ipo if q7 and ipo else None; cpa=a7/o if o and a7 else None
    jud=('📦広告なし' if not a7 else ('🚨赤字' if p7<0 else ('🔧テコ入れ' if yo is not None and yo<0.30 else
         ('🚀増額候補' if tg and mer>=tg else '✅維持'))))
    P.append([n,'季節' if n in SEA else '通年',s7,k7,round(a7),round(p7),round(p7/s7,4) if s7 else '',
        round(mer,2) if mer else '—',round(be,2) if be else '—',round(tg,2) if tg else '—',round(yo,2) if yo is not None else '—',
        s1,round(p1),round(p1/s1,4) if s1 else '',round(p3),round(p30),round(pc),q7,o or '',
        round(ipo,2) if ipo else '',round(cpa) if cpa else '',round(bcpa) if bcpa else '',BUD.get(MAP.get(n,n),'—'),jud])
P.append(['カタログ全部（テスト）※全商品横断','—','—','—',round(R['CAT7']),'—','—','—','—','—','—','—','—','—','—','—','—','—','—','—','—','—',BUD.get('カタログ全部（テスト）','—'),'—'])
sh('商品別',['商品別（7日=判定単位）。カタログ行の広告費は「7日合計」（前日ではない）',
  '分岐CPA(注文) = 1個あたり粗利 × 点数/注文。まとめ買い商品（点数/注文>1.2）は必ず注文ベースで比較する',
  'Σ商品利益 − カタログ7日広告費 = 全店7日利益（差0円）／ Σ商品広告費 + カタログ = Metaアカウント7日消化（差0円）'],
 ['商品','区分','7日売上','7日原価','7日広告','7日利益','7日利益率','MER','分岐','目標','余裕','前日売上','前日利益','前日利益率',
  '3日利益','30日利益','当月累積利益','7日販売数','7日注文数','点数/注文','CPA(注文)','分岐CPA(注文)','現日予算','判定'],
 P,[26,6]+[12]*22,{3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',12:'#,##0',13:'#,##0',14:'0.0%',15:'#,##0',16:'#,##0',17:'#,##0',21:'#,##0',22:'#,##0',23:'#,##0'})
# ---------- 5 レバー一覧 ----------
LV=[]
for n in sorted(R['N7'],key=lambda x:-R['N7'][x]):
    cp=MAP.get(n,n); b=BUD.get(cp)
    if not b: continue
    s7,k7,a7,p7=blk(n,R['N7'],R['K7'],R['A7'])
    if not a7: continue
    gm=1-k7/s7; mer=s7/a7; mm=mer*0.7
    days=[d for d in D7 if d in BSNAP[cp]]
    wu=(a7/sum(BSNAP[cp][d] for d in days)) if len(days)==7 and sum(BSNAP[cp][d] for d in days) else None
    LV.append([n,'季節' if n in SEA else '通年',b,round(wu,3) if wu else '',round(mer,2),round(1/gm,2),round(mer-1/gm,2),
        round(mm*gm-1,3),round((1-mm*gm)*b*0.25*7),round((mm*gm-1)*b*0.25*7),round(p7)])
LV.sort(key=lambda x:-max(x[8],x[9]))
sh('レバー一覧',['「どれを動かすと一番効くか」は7日平均の利益額ではなく、動かしたときの変化量で並べる',
  '削減1円あたりの利益変化 = 1 − 限界MER×粗利率 ／ 増額1円あたり = 限界MER×粗利率 − 1（限界MER = 平均MER×0.7）',
  '増額は週消化率95%以上のものだけが対象（それ未満は増やしても使われない）'],
 ['商品','区分','日予算','週消化率','MER','分岐','余裕','1円あたり利益','−25%で週','+25%で週','7日利益'],
 LV,[26,6,11,10,8,8,8,13,13,13,12],{3:'#,##0',4:'0.0%',9:'#,##0',10:'#,##0',11:'#,##0'})
# ---------- 6 週次診断 ----------
Sd=collections.defaultdict(lambda: collections.defaultdict(lambda:[0,0]))
for r_ in csv.DictReader(open('data/daily/daily_sales.csv')):
    Sd[r_['name']][r_['date']][0]+=int(r_['gross']); Sd[r_['name']][r_['date']][1]+=int(r_['disc'])
ADd=collections.defaultdict(lambda: collections.defaultdict(float))
for r_ in csv.DictReader(open('data/daily/daily_ad.csv')): ADd[r_['campaign']][r_['date']]+=float(r_['cost'])
COST=dict(R['COST']); PRICE=dict(R['PRICE'])
COST['ナノガラス脱毛パッド']=(855*757+373*740)/1228; COST['壁掛けディスペンサー']=(41*2528+13*1981)/54
PRICE['壁掛けディスペンサー']=363920/54
end=datetime.date(2026,8,10); WK=[]
for i in range(6):
    a=end-datetime.timedelta(days=7*i+6); WK.append((f"{a.strftime('%m/%d')}-{(a+datetime.timedelta(days=6)).strftime('%m/%d')}",
        [(a+datetime.timedelta(days=j)).isoformat() for j in range(7)]))
WK.reverse()
def wblk(days,names=None):
    s=c=0
    for n,v in Sd.items():
        if names is not None and n not in names: continue
        g=sum(v[d][0] for d in days if d in v); dc=sum(v[d][1] for d in days if d in v)
        if not g: continue
        s+=g-dc
        if n in PRICE and n in COST: c+=g/PRICE[n]*COST[n]
    a=sum(ADd[MAP.get(n,n)].get(d,0) for n in (names if names is not None else Sd) for d in days) if names is not None \
      else sum(ADd[cp].get(d,0) for cp in ADd for d in days)
    return s,c,a,s-c-a
YR={n for n in Sd if n and n not in SEA}
WD_=[]
for lbl,days in WK:
    if days[0]<'2026-06-29': continue
    s,c,a,p=wblk(days); ss,_,sa,sp=wblk(days,SEA); ys,_,ya,yp=wblk(days,YR)
    WD_.append([lbl,round(s),round(c),round(a),round(p),round(p/s,4),round(s/a,2),round(c/s,4),round(a/s,4),
                round(sp),round(ss/sa,2),round(yp),round(ys/ya,2)])
sh('週次診断',['全店の週次分解。原価は日次のバリアント構成が取れないナノガラス/ディスペンサーのみ30日実測ミックスの加重平均単価【近似】',
  '★結論: 原価率は6週間ほぼ不変（27.9%→28.8%）。利益率の−18.7ptは全部が広告費率の上昇（33.8%→51.6%）',
  '★利益減 −2,157,846円/週 のうち 季節物13品が −2,100,537円＝97.3%。通年物14品のMERは3週連続で回復中'],
 ['週','実売','原価','広告費','利益','利益率','MER','原価率','広告費率','季節物13品 利益','季節物MER','通年物14品 利益','通年物MER'],
 WD_,[14,13,12,12,12,9,8,9,10,15,11,15,11],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'#,##0',6:'0.0%',7:'0.00',8:'0.0%',9:'0.0%',10:'#,##0',11:'0.00',12:'#,##0',13:'0.00'})
# ---------- 7 CVR順位・ファネル ----------
FN=[]
for n,(c,imp,clk) in MF.items():
    s=R['N7'].get(n,0); k=R['K7'].get(n,0); q=R['Q7'].get(n,0); o=ORD.get(n,0)
    if not s or not o: continue
    ipo=q/o; bcpa=(s-k)/q*ipo
    FN.append([n,'季節' if n in SEA else '通年',imp,round(clk/imp,4),clk,round(o/clk,4),round(s/clk),
               round(c/clk),round(c/o),round(bcpa),round(bcpa-c/o),round(c/imp*1000)])
mc=statistics.median([x[5] for x in FN]) if FN else 0; mr=statistics.median([x[6] for x in FN]) if FN else 0
FN.sort(key=lambda x:-x[6])
sh('CVR順位',['CVR = Shopify注文 ÷ Metaアウトバウンドクリック（Metaの購入数は損益に使わない）',
  f'中央値 CVR {mc*100:.2f}% ／ 売上perクリック {mr:.0f}円。LP改修の対象は「CVRも売上/クリックも中央値未満」の商品だけ',
  'CVRの単純順位は価格帯を混同する。売上perクリック（=AOV×CVR）と同価格帯内で比較すること'],
 ['商品','区分','インプ','CTR','クリック','CVR','売上/クリック','CPC','CPA(注文)','分岐CPA(注文)','余裕','CPM'],
 FN,[26,6,12,9,11,9,13,9,12,15,11,10],
 {3:'#,##0',4:'0.00%',5:'#,##0',6:'0.00%',7:'#,##0',8:'#,##0',9:'#,##0',10:'#,##0',11:'#,##0',12:'#,##0'})
# ---------- 8 日次推移 ----------
days=sorted({d for v in Sd.values() for d in v}); WDJ=['月','火','水','木','金','土','日']
pr={}
for d in days:
    s=c=0
    for n,v in Sd.items():
        if n=='' or d not in v: continue
        g,dc=v[d]; s+=g-dc
        if g and n in PRICE and n in COST: c+=g/PRICE[n]*COST[n]
    a=sum(ADd[cp].get(d,0) for cp in ADd); pr[d]=(s,c,a,s-c-a)
DD=[]
for j,d in enumerate(days):
    if d<'2026-07-13': continue
    s,c,a,p=pr[d]; ma=sum(pr[x][3] for x in days[j-2:j+1])/3
    DD.append([d,WDJ[datetime.date.fromisoformat(d).weekday()],round(s),round(c),round(a),round(p),round(p/s,4),
               round(s/a,2),round(ma),'⚠️非常ブレーキ' if ma<130000 else ''])
sh('日次推移',['日次の利益は上記【近似】原価ベース。4期間サマリーはバリアント別実数量の正確値',
  '非常ブレーキ: 3日移動平均利益 < 130,000円 で発動（カタログ→4WAY→日傘の順に緊急減額）'],
 ['日付','曜','実売','原価','広告費','利益','利益率','MER','3日移動平均利益','警告'],
 DD,[12,5,13,12,12,12,9,8,16,16],{3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',8:'0.00',9:'#,##0'})
# ---------- 9 曜日指数 ----------
sh('曜日指数',['直近30日(7/19-8/17)・祝日 7/20・8/11 を除外。売上=gross−値引',
  '★月曜の実績は7日平均ではなく「平常月曜平均」と比べる。祝日には平日の曜日指数を当てない（祝日は日曜級113〜127で評価）',
  '★2026-08-15(土) 売上769,897円は非祝日土曜平均1,238,022円の62.2%。ただし7日平均比は+2.2%で、下げ止まっている',
  '★害虫ブロッカーは8/15 14:59にDRAFT化（PSE）。8/15は24,671円消化・8/16は消化ゼロで停止を確認',
  '★接触冷感UVパーカーは8/15に4,980→3,980へ戻した（値上げ撤回）。8/13-15は価格混在のためMIXEDDAYで分解',
  '★曜日指数は 7/20(海の日)・8/11(山の日) を除外して算出'],
 ['曜日','指数(全体=100)'],[[k,round(v,1)] for k,v in sorted(R['IDX'].items(),key=lambda x:-x[1])],[10,16])
wb.save('data/reports/report-2026-08-20.xlsx')
print('本体シート:',wb.sheetnames)
