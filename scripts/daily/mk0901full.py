# -*- coding: utf-8 -*-
"""2026-09-01 定例レポート本体（昨日=8/31 月）。"""
import pickle, csv, datetime, collections, statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
R = pickle.load(open('/tmp/rep0831w.pkl', 'rb')); FEE = 0.03423   # 2026-09-01 月次再計測（SP79.7% + スマホ17.1% + Paidy3.2%）
D1, D3, D7, D30, CUM = R['D1'], R['D3'], R['D7'], R['D30'], R['CUM']
SNAP = '2026-08-30b'   # ★8/30 昼にユーザーが6セット見直した後の実測（Meta直読で確認）
rows = list(csv.reader(open('data/budget-snapshots.csv'))); body = [r for r in rows[1:] if r]
assert any(r[0] == SNAP for r in body), f'{SNAP} スナップショット未記録'
BUD = collections.defaultdict(int)
for r in body:
    if r[0] == SNAP: BUD[r[1]] += int(r[3])
assert sum(BUD.values()) == 255000, sum(BUD.values())
# ★2026-08-30修正: 同じ日に複数回スナップショットを取った日（末尾 b/c/d）を
#   従来は `rstrip('b')` して **合算** していたため、その日の予算が2〜3倍に膨らみ、
#   週消化率が実態より低く出ていた（例: 偏光・調光サングラス 8/25 は 10,000+10,000+7,000+7,000 = 34,000 と誤集計）。
#   正しくは「その日の最後のスナップショット＝その日の大半で効いていた値」を採る。
_BS = collections.defaultdict(dict)
for r in body:
    day = _BS[r[1]].setdefault(r[0][:10], {}); day[r[0]] = day.get(r[0], 0) + int(r[3])
BSNAP = collections.defaultdict(lambda: collections.defaultdict(int))
for cp, byday in _BS.items():
    for d, snaps in byday.items(): BSNAP[cp][d] = snaps[max(snaps)]

MAP = R['MAP']; INV = {v: k for k, v in MAP.items()}
# --- 7日(8/25-31) Shopify注文数（ShopifyQL GROUP BY product_title 実測）---
ORD = {'ナノガラス脱毛パッド':257,'W固定スマホ車載ホルダー':128,'快適マジックインソール':83,'ムダ毛シェーバー':54,
'むくみ取りかっさ':38,'2WAYシートボックス':34,'姿勢サポートチェア':33,'バランスケアスリッパ':33,
'完全遮光・接触冷感UVハット':16,'ナノバブルシャワーヘッド':16,'偏光・調光サングラス':13,'3D足臭リセットブラシ':8,
'4WAY':5,'形状記憶日傘':5,'ビジュアル耳かき':5,'優先配送':4,'携帯電動シェーバー':3,
'湯上がりガーゼワンピース':1,'瞬間冷感ポンチョ':1,'接触冷感UVアームカバー':1,'接触冷感UVパーカー':1,
'ジェットウォッシャー':1,'壁掛けディスペンサー':1,'ヘアドライタオル':1,'1秒折り畳みチェア':1,'完全遮光・形状記憶':1}
STORE_ORD7, STORE_SESS7 = 732, 21915
assert sum(ORD.values()) == 744, sum(ORD.values())   # 全店732注文より多いのは複数商品を含む注文が両方で数えられるため

# --- 30日(8/02-8/31) Shopify注文数（ShopifyQL GROUP BY product_title 実測）---
ORD30 = {'ナノガラス脱毛パッド':1214,'W固定スマホ車載ホルダー':520,'形状記憶日傘':248,'ムダ毛シェーバー':209,
'接触冷感UVパーカー':204,'4WAY':191,'害虫ブロッカー':190,'快適マジックインソール':188,'偏光・調光サングラス':168,
'2WAYシートボックス':145,'完全遮光・接触冷感UVハット':116,'完全遮光・形状記憶':115,'姿勢サポートチェア':105,
'優先配送':93,'バランスケアスリッパ':78,'5WAY腰掛けファン':73,'接触冷感UVアームカバー':72,'むくみ取りかっさ':46,
'携帯電動シェーバー':43,'卓上冷感クーラー':41,'瞬間冷感ポンチョ':23,'3D足臭リセットブラシ':21,
'ナノバブルシャワーヘッド':19,'3WAYサーキュレーター':15,'健康サンダル':11,'ビジュアル耳かき':5,
'壁掛けディスペンサー':4,'ジェットウォッシャー':4,'ヘアドライタオル':3,'湯上がりガーゼワンピース':3,
'UV歯ブラシ除菌器':2,'瞬間冷却ハンディファン':1,'姿勢サポートベルト':1,'指圧マット':1,
'1秒折り畳みチェア':1,'リカバリーサンダル':1}

# --- 30日(8/02-8/31) Metaファネル ---
# ★★ data_query がハーネス側でブロックされているため、通る campaign_and_resource_get の
#    health_check（last_30_days = 8/02-8/31 の実測・UTC基準の窓）から取った。したがって:
#      ・窓は「7日」ではなく「30日」。過去レポートの CVR順位シートと数値を並べて比較しないこと
#      ・クリックは link_clicks ではなく **全クリック**（CTR/CVR/CPC は全クリック基準で低め/高めに出る）
#    全キャンペーンで health_check の spend = CSV(8/02-8/31) が **差0円** で一致することを検算済み。
MF_RAW = {'ナノガラス脱毛パッド':(2385747,1296479,38491),'W固定スマホ車載ホルダー':(895137,473600,15500),
'カタログ全部（テスト）':(638461,165872,6816),'ムダ毛シェーバー':(567015,463733,6810),
'偏光・調光サングラス':(451913,204466,6519),'快適マジックインソール':(368943,218790,6349),
'2WAYシートボックス':(343603,117623,4063),'姿勢サポートチェア':(304931,191239,4351),
'完全遮光・接触冷感UVハット':(282786,65346,3590),'バランスケアスリッパ':(188616,123986,2579),
'むくみ取りかっさ':(86010,23567,933),'ナノバブルシャワーヘッド':(59827,26940,1050),
'ビジュアル耳かき':(26408,9374,186)}
MF = {INV.get(k, k): v for k, v in MF_RAW.items()}

SEA = {'4WAY','3WAYサーキュレーター','卓上冷感クーラー','5WAY腰掛けファン','瞬間冷却ハンディファン','接触冷感UVパーカー',
'接触冷感UVアームカバー','瞬間冷感ポンチョ','完全遮光・接触冷感UVハット','形状記憶日傘','完全遮光・形状記憶',
'偏光・調光サングラス','害虫ブロッカー','湯上がりガーゼワンピース'}

wb = Workbook(); wb.remove(wb.active)
TH = Font(name='Arial', bold=True, color='FFFFFF', size=10); TD = Font(name='Arial', size=10)
NEG = Font(name='Arial', size=10, color='CC0000'); HEAD = PatternFill('solid', fgColor='305496')
thin = Border(*[Side(style='thin', color='CCCCCC')] * 4)

def sh(title, note, cols, data, widths, fmt=None):
    ws = wb.create_sheet(title); r = 1
    if note:
        for ln in note: ws.cell(r, 1, ln).font = Font(name='Arial', size=10, bold=(r == 1)); r += 1
        r += 1
    for c, v in enumerate(cols, 1): ws.cell(r, c, v)
    for c in range(1, len(cols) + 1):
        x = ws.cell(r, c); x.font = TH; x.fill = HEAD; x.border = thin
        x.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 30; hr = r; r += 1
    for row in data:
        for c, v in enumerate(row, 1):
            x = ws.cell(r, c, v); x.border = thin; x.font = TD
            if isinstance(v, (int, float)) and v < 0: x.font = NEG
            if fmt and c in fmt: x.number_format = fmt[c]
        r += 1
    for i, w_ in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i < 27 else 'A' + chr(38 + i)].width = w_
    ws.freeze_panes = ws.cell(hr + 1, 1).coordinate
    return ws

# ===== 日次CSVを読む（3窓判定・週次・日次推移で使う）=====
Sd = collections.defaultdict(lambda: collections.defaultdict(lambda: [0, 0]))
for r_ in csv.DictReader(open('data/daily/daily_sales.csv')):
    Sd[r_['name']][r_['date']][0] += int(r_['gross']); Sd[r_['name']][r_['date']][1] += int(r_['disc'])
ADd = collections.defaultdict(lambda: collections.defaultdict(float))
for r_ in csv.DictReader(open('data/daily/daily_ad.csv')): ADd[r_['campaign']][r_['date']] += float(r_['cost'])
COSTa = dict(R['COST']); PRICEa = dict(R['PRICE'])
COSTa['ナノガラス脱毛パッド'] = (855 * 757 + 373 * 740) / 1228          # 30日実測ミックスの加重平均【近似】
COSTa['壁掛けディスペンサー'] = (41 * 2528 + 13 * 1981) / 54
PRICEa['壁掛けディスペンサー'] = 363920 / 54
ALLD = sorted({d for v in Sd.values() for d in v if d <= D1[0]})
D5 = ALLD[-5:]
def win(n, days):
    """商品nのその窓の（実売, 原価【近似】, 広告費）"""
    g = sum(Sd[n][d][0] for d in days if d in Sd[n]); dc = sum(Sd[n][d][1] for d in days if d in Sd[n])
    c = g / PRICEa[n] * COSTa[n] if g and n in PRICEa and n in COSTa else 0
    a = sum(ADd[MAP.get(n, n)].get(d, 0) for d in days)
    return g - dc, c, a

# ---------- 0 収益（先頭シート・2026-08-31ユーザー要望）----------
# 「毎日のExcelで収益を見たい」に対する常設シート。
# 昨日の損益 → 日次の推移 → 期間別 → 当月の着地ペース の順で、上から読めば収益だけ分かる。
WDJ0 = ['月', '火', '水', '木', '金', '土', '日']
def dayblk(d):
    """1日ぶんの（実売, 原価【近似】, 広告費, 利益）。原価はバリアント構成が取れないため近似"""
    s = c = 0
    for n, v in Sd.items():
        if n == '' or d not in v: continue
        g, dc = v[d]
        if not g: continue
        s += g - dc
        if n in PRICEa and n in COSTa: c += g / PRICEa[n] * COSTa[n]
    a = sum(ADd[cp].get(d, 0) for cp in ADd)
    return s, c, a, s - c - a

PROF = []
_pl = [dayblk(d) for d in ALLD]
for i, d in enumerate(ALLD):
    if d < ALLD[-21]: continue
    s, c, a, p = _pl[i]
    ma = sum(_pl[j][3] for j in range(i-2, i+1)) / 3 if i >= 2 else None
    PROF.append([d, WDJ0[datetime.date.fromisoformat(d).weekday()], round(s), round(c), round(a),
                 round(p), round(p/s, 4), round(p - s*FEE), round(s/a, 2) if a else '',
                 round(ma) if ma else '', ('🚨' if ma and ma < 130000 else ('✅' if ma else '')),
                 '★昨日' if d == D1[0] else ''])
# 当月の着地ペース
_cum = [dayblk(d) for d in CUM]
cs = sum(x[0] for x in _cum); cc = sum(x[1] for x in _cum); ca = sum(x[2] for x in _cum); cp = cs-cc-ca
import calendar
_dim = calendar.monthrange(2026, 8)[1]
sh('収益', [f'LOOTY 収益レポート 2026-09-01（昨日 = 8/31 月）★8月が確定しました',
  f'★昨日の利益 {round(_pl[-1][3]):,}円（利益率 {_pl[-1][3]/_pl[-1][0]:.1%}）／手数料控除後 {round(_pl[-1][3]-_pl[-1][0]*FEE):,}円',
  f'★当月累積（8/01-{CUM[-1][8:]}・{len(CUM)}日）実売 {cs:,.0f} / 原価 {cc:,.0f} / 広告 {ca:,.0f} → '
  f'**利益 {cp:,.0f}円（{cp/cs:.1%}）**／1日あたり {cp/len(CUM):,.0f}円',
  f'★**8月の確定値: 実売 {cs:,.0f}円 / 利益 {cp:,.0f}円（{cp/cs:.1%}）**（31日・着地ではなく実績）',
  '★非常ブレーキ = 3日移動平均利益 < 130,000円。🚨が付いた日は発動水準',
  '原価は日次のバリアント構成が取れないナノガラス/壁掛けディスペンサーのみ30日実測ミックスの加重平均単価【近似】。'
  '期間別サマリー（次シート）はバリアント別実数量の正確値'],
 ['日付','曜','実売','原価','広告費','利益','利益率','手数料後利益','MER','3日移動平均利益','ブレーキ','備考'],
 PROF, [12,5,13,12,12,13,9,14,8,16,9,10],
 {3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',8:'#,##0',9:'0.00',10:'#,##0'})

# ---------- 1 シンプル判定（3窓判定つき） ----------
def blk(n, N, K, A_):
    s = N.get(n, 0); k = K.get(n, 0); a = A_.get(n, 0); return s, k, a, s - k - a
SIMPLE = []
for n in sorted(R['N7'], key=lambda x: -R['N7'][x]):
    a7 = R['A7'].get(n, 0)
    if not a7: continue
    s7, k7, _, p7 = blk(n, R['N7'], R['K7'], R['A7'])
    s3, k3, a3, p3 = blk(n, R['N3'], R['K3'], R['A3']); s1, k1, a1, p1 = blk(n, R['N1'], R['K1'], R['A1'])
    gm = 1 - k7 / s7; be = 1 / gm; tg = 1 / (gm - 0.35) if gm - 0.35 > 0 else None; mer = s7 / a7
    cp = MAP.get(n, n); days = [d for d in D7 if d in BSNAP[cp]]
    wu = (a7 / sum(BSNAP[cp][d] for d in days)) if len(days) == 7 and sum(BSNAP[cp][d] for d in days) else None
    # ★3窓判定（3日/5日/7日・すべて前日終端）
    w3 = []
    for dd in (D3, D5, D7):
        sx, cx, ax = win(n, dd)
        w3.append((sx / ax if ax else None, 1 / (1 - cx / sx) if sx and sx > cx else None))
    below = all(m is not None and b is not None and m < b for m, b in w3)
    tight = all(m is not None and b is not None and m - b < 0.30 for m, b in w3)
    rich = all(m is not None and tg and m >= tg for m, _ in w3) and wu is not None and wu >= 0.95
    tri = '🚨3窓すべて分岐割れ→停止/−50%' if below else ('⚠️3窓すべて余裕<0.30→−25%' if tight else
          ('🚀3窓すべて目標超＋消化95%→+25%' if rich else '⏸窓が割れた→据え置き'))
    if p7 < 0 or mer < be: st = '🚨縮小・停止'
    elif (a3 and s3 / a3 < be) or mer < be + 0.30: st = '🔧改善'
    elif tg and mer >= tg and wu and wu >= 0.95: st = '🚀伸ばす候補'
    else: st = '✅維持'
    mm = mer * 0.7; per = mm * gm - 1
    # 頑健性: 限界MER = 平均MER×0.6/0.7/0.8 の3端で符号が変わらないか
    ends = [mer * x * gm - 1 for x in (0.6, 0.7, 0.8)]
    rob = '一致' if (all(e > 0 for e in ends) or all(e < 0 for e in ends)) else '⚠️符号が割れる'
    SIMPLE.append([n, round(mer, 2), round(s3 / a3, 2) if a3 else '', round(s1 / a1, 2) if a1 else '',
        round(be, 2), round(tg, 2) if tg else '', round(p7), round(p3), round(p1),
        round(wu, 3) if wu else '', round(per, 3), rob, st, tri])
sh('シンプル判定', ['LOOTY 2026-09-01 定例（昨日=8/31 月）',
 '状態: 🚨縮小・停止=7日MER<分岐 or 7日利益マイナス ／ 🔧改善=3日MER<分岐 or 余裕<0.30 ／ 🚀伸ばす候補=7日MER≥目標かつ週消化率≥95% ／ ✅維持',
 '3窓判定（3日/5日/7日・すべて8/31終端）が本番の意思決定ルール。窓が1つでも割れたら据え置き＝動かさない',
 '「広告費1円あたり利益」= 限界MER×粗利率 − 1（限界MER = 平均MER×0.7）。頑健性は×0.6/0.7/0.8 の3端で符号が変わらないかを見る'],
 ['商品','7日MER','3日MER','前日MER','分岐','目標','7日利益','3日利益','前日利益','週消化率','1円あたり利益','頑健性','状態','3窓判定'],
 SIMPLE, [26,9,9,9,8,8,12,12,12,10,12,13,14,30], {7:'#,##0',8:'#,##0',9:'#,##0',10:'0.0%'})

# ---------- 2 ネクストアクション ----------
p3ma = (sum(R['N3'].values()) - sum(R['K3'].values()) - R['ACCT3']) / 3
NA = [
 ['📅 月次確定','**8月が確定しました**',
  f'実売 22,082,602円 / 原価 6,093,524 / 広告 11,071,845 → **利益 4,917,233円（22.3%）**／手数料控除後 4,161,346円。'
  '1日あたり利益 158,620円。7月（利益率32.9%）からは −10.6pt で、差はすべて広告費率（37.9%→50.1%）',
  '—','記録'],
 ['🔴 判定日','ナノバブルシャワーヘッド 本判定 → ✅合格・継続',
  '宣言済み基準「倍率 ≥1.3 継続 ／ 1.0〜1.3 −50% ／ <1.0 停止」。'
  '7日(8/25-31) 実売117,264円 / 17点 / 広告49,128円 → 実CPA 2,890円 vs 分岐CPA 4,725円 = **倍率1.64**。利益 +29,801円。'
  '3窓とも余裕+0.90〜1.21で安定 → **日8,000円で継続**。★増額は日10,000円が上限（3〜5月に壊した実測がある）',
  '—','本日 完了'],
 ['🔴 判定日','決済ミックス 月次再計測 → ✅完了',
  'Shopify Payments 79.7% / KOMOJUスマホ決済 17.1% / Paidy 3.2% → **実効ブレンド率 3.452% → 3.423%**。'
  '★動いたのは **Paidy が 7.5%→3.2% と半減**したこと。後払い需要の減少なので、来月も下がるようなら客層の変化として見る。'
  'FEE定数とSKILL.mdは更新済み',
  '年+7.8万円','本日 完了'],
 ['✅ 実行確認','ビジュアル耳かき は「減額」ではなく **停止** された',
  '昨朝の私の提案は「8,000→5,000円」だったが、Metaで **PAUSED** を確認。ユーザー判断として記録する。'
  '全期間(8/28-31) 実売28,884円 / 6点 / 広告26,408円 → 実CPA 4,401円 vs 分岐3,325円 = **倍率0.76**・損益 **−7,454円**。'
  '★8/31を含めた4日で見ると倍率は0.97→0.76へ悪化しており、**停止のほうが正しかった**',
  '+約8,000円/日','完了'],
 ['🚨 要確認','偏光・調光サングラスが 8/31 に 151円しか消化していない',
  '日予算5,000円に対し **3.0%**。推移は 8/28 6,498 → 8/29 6,164 → 8/30 2,816 → 8/31 **151**円と3日連続で急減。'
  'キャンペーンは ENABLED のまま。**広告の否認・素材の停止・支払いエラーのいずれかを疑う。'
  'Meta広告マネージャで広告セットと広告のステータス、エラー表示を確認してください**',
  '—','本日'],
 ['⏸ 触らない','快適マジックインソール（P2の1日目）',
  'P2(8/31-9/2)が昨日から開始。1日目は 実売42,984円 / 11点 / 広告26,154円 → **利益 +9,383円**（合格ライン 1日23,280円）。'
  '★1日目としては基準の40%。ただし月曜は最弱曜日（指数92.0）で、P1(8/28-29)は金土。**残り2日で判断する**',
  '—','9/3に執行'],
 ['⏸ 触らない','バランスケアスリッパ（増額のロック中）',
  '8/30に 10,000→12,000 へ増額。3窓とも余裕+1.07〜1.22で健全。**9/3に「1日あたり利益 ≥ 増額前の7日平均」で判定**',
  '—','9/3に執行'],
 ['⚠️ 未確定','本日の予算スナップショットが未取得',
  'このレポートの「現日予算」列と週消化率は **8/30 18時のMeta直読値（合計255,000円）** を使っている。'
  '耳かきの停止は反映されていないので、実際の日予算は 247,000〜250,000円のはず。'
  '★予算を変更していたら教えてください',
  '—','本日'],
 ['🔴 判定日','カタログ 3日判定（8/31-9/2）','CPA基準の倍率 ≥1.00 で据え置き。前回(8/27-29)は1.12。8/31の消化は17,032円','—','9/2'],
 ['🔴 判定日','優先配送 790円',
  '宣言済み基準「14日で件数 ≥ 28件」。7日実測 **4件**（アタッチ率 4/732 = 0.55%）。前週6件からさらに減少。不合格の見込み','—','9/4'],
 ['🟡 推奨','アップセル手動ペア5組の設定',
  'data/lp/upsell-pairs-2026-08-26.txt の5組をアプリに手入力する。異商品ミックス率が実質ゼロなので'
  '自動推薦は学習データがなく全店売れ筋にフォールバックする。判定は実施+14日で異商品ミックス率3.0%以上','週+約4万円','任意'],
 ['🟡 推奨','カテゴリタグ40件の付与',
  'data/tags/category-mutation-2026-08-23.graphql を GraphiQL で1回実行 ＋ ナノバブルの季節タグ是正','—','任意'],
 ['🟡 推奨','コレクション冒頭文7本の貼り付け',
  'data/lp/collection-intro-2026-08-23.txt。あわせて「暖房・防寒グッズ」コレクションのAmazon由来HTML説明を削除','—','任意'],
 ['🟡 推奨','購入通知アプリ',
  'Judge.me の Pop-up Reviews Widget は導入済みアプリの機能なので追加0円。Sales Popを足すならQikify Basic $6.99（訪問者無制限）','—','任意'],
 ['🟡 推奨','CJで原価確認',
  'あったかインソール 1,000円以下 ／ 外反母趾サポーター 800円以下 なら秋冬のテスト枠に入れる','—','任意'],
]
sh('ネクストアクション', ['今日やること・判定カレンダー・持ち越しタスク（実施が確認できるまで毎日残す）'],
   ['区分','商品/対象','内容','効果','期限'], NA, [12, 26, 78, 16, 10])

# ---------- 3 全体サマリー ----------
SUMR = []
for lbl, N, K, ACC in [('前日 8/31(月)', R['N1'], R['K1'], R['ACCT1']), ('3日 8/29-31', R['N3'], R['K3'], R['ACCT3']),
        ('7日 8/25-31', R['N7'], R['K7'], R['ACCT7']), ('30日 8/02-8/31', R['N30'], R['K30'], R['ACCT30']),
        ('★当月累積 8/01-31（8月確定）', R['NC'], R['KC'], R['ACCTC'])]:
    s = sum(N.values()); c = sum(K.values()); p = s - c - ACC
    SUMR.append([lbl, s, c, round(ACC), round(c + ACC), round((c + ACC) / s, 4), round(p), round(p / s, 4),
                 round(s / ACC, 2), round(1 / (1 - c / s), 2), round(s * FEE), round(p - s * FEE), round((p - s * FEE) / s, 4)])
sh('全体サマリー', ['全体サマリー（売上=Shopify gross−値引・広告費=Meta実測。返品は売上からも利益からも控除しない=A案）',
 '総合原価＝原価＋広告費。恒等式 総合原価率＋利益率＝100% が全窓で成立していることを確認済み',
 '縦照合3本すべて通過: Σ商品gross=全店400,180 ／ Σ値引=5,971 ／ Σキャンペーン広告費=223,482円（8/31）',
 '★8/31の広告費は data_query がブロックされたため health_check(8/02-8/31実測) − CSV(8/02-8/30実測) で導出。'
 '窓の検証: ビジュアル耳かき(8/28開始)は CSV 20,554 / HC 26,408 → 8/31 = 5,854円。推計ではなく2つの実測の差',
 '決済ブレンド率 3.423%（2026-09-01再計測: SP79.7%×3.25% ＋ KOMOJUスマホ17.1%×4.1% ＋ Paidy3.2%×4.1%。前回3.452%からPaidyが半減）',
 f'7日の全店CVR（Shopifyセッション基準）= {STORE_ORD7}注文 ÷ {STORE_SESS7:,}セッション = {STORE_ORD7/STORE_SESS7:.2%}'],
 ['窓','実売','原価','広告費','総合原価','総合原価率','利益','利益率','MER','分岐MER','決済手数料','手数料控除後利益','手数料後利益率'],
 SUMR, [20,14,13,13,13,11,13,10,8,9,12,15,12],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'#,##0',6:'0.0%',7:'#,##0',8:'0.0%',9:'0.00',10:'0.00',11:'#,##0',12:'#,##0',13:'0.0%'})
ws = wb['全体サマリー']; r = ws.max_row + 3
ws.cell(r, 1, '■ 昨日(8/31 月) × 期間比較').font = Font(name='Arial', bold=True, size=11); r += 1
S1 = sum(R['N1'].values()); C1 = sum(R['K1'].values()); A1 = R['ACCT1']
per = [('実売', S1, sum(R['N3'].values())/3, sum(R['N7'].values())/7, sum(R['N30'].values())/30),
       ('原価', C1, sum(R['K3'].values())/3, sum(R['K7'].values())/7, sum(R['K30'].values())/30),
       ('広告費', A1, R['ACCT3']/3, R['ACCT7']/7, R['ACCT30']/30),
       ('利益', S1-C1-A1, (sum(R['N3'].values())-sum(R['K3'].values())-R['ACCT3'])/3,
        (sum(R['N7'].values())-sum(R['K7'].values())-R['ACCT7'])/7,
        (sum(R['N30'].values())-sum(R['K30'].values())-R['ACCT30'])/30)]
for c, v in enumerate(['指標','8/31(月)','3日平均/日','7日平均/日','30日平均/日','vs7日平均'], 1): ws.cell(r, c, v)
for c in range(1, 7):
    x = ws.cell(r, c); x.font = TH; x.fill = HEAD; x.border = thin; x.alignment = Alignment(horizontal='center')
r += 1
for lbl, d1, m3, m7, m30 in per:
    for c, v in enumerate([lbl, round(d1), round(m3), round(m7), round(m30), round(d1/m7-1, 4)], 1):
        x = ws.cell(r, c, v); x.border = thin; x.font = NEG if (isinstance(v, (int, float)) and v < 0) else TD
        if c in (2,3,4,5): x.number_format = '#,##0'
        if c == 6: x.number_format = '+0.0%;-0.0%'
    r += 1
ws.cell(r+1, 1, '⚠️ 同曜日平均比は当面読まないこと。直近30日の日曜（8/02・8/09・8/16）が夏物崩壊前・日予算590,000時代で汚染されている。'
  f'実態は前週月曜(8/24 529,383円)比 **−25.5%**。同じ月曜どうしでこの落差なので曜日では説明できない。'
  f'ただし3日移動平均利益 {p3ma:,.0f}円 は非常ブレーキ130,000円を6日連続で上回っている').font = Font(name='Arial', size=10, color='CC0000')

# ---------- 4 商品別 ----------
P = []
for n in sorted(R['N7'], key=lambda x: -R['N7'][x]):
    s7,k7,a7,p7 = blk(n,R['N7'],R['K7'],R['A7']); s1,k1,a1,p1 = blk(n,R['N1'],R['K1'],R['A1'])
    s3,k3,a3,p3 = blk(n,R['N3'],R['K3'],R['A3']); s30,k30,a30,p30 = blk(n,R['N30'],R['K30'],R['A30'])
    sc,kc,ac,pc = blk(n,R['NC'],R['KC'],R['AC'])
    gm = 1-k7/s7 if s7 else 0; mer = s7/a7 if a7 else None; be = 1/gm if gm else None
    tg = 1/(gm-0.35) if gm-0.35 > 0 else None; yo = (mer-be) if mer and be else None
    q7 = R['Q7'].get(n, 0); o = ORD.get(n); ipo = q7/o if o else None
    bcpa = (s7-k7)/q7*ipo if q7 and ipo else None; cpa = a7/o if o and a7 else None
    jud = ('📦広告なし' if not a7 else ('🚨赤字' if p7 < 0 else ('🔧テコ入れ' if yo is not None and yo < 0.30 else
           ('🚀増額候補' if tg and mer >= tg else '✅維持'))))
    P.append([n, '季節' if n in SEA else '通年', s7, k7, round(a7), round(p7), round(p7/s7,4) if s7 else '',
        round(mer,2) if mer else '—', round(be,2) if be else '—', round(tg,2) if tg else '—',
        round(yo,2) if yo is not None else '—', s1, round(p1), round(p1/s1,4) if s1 else '',
        round(p3), round(p30), round(pc), q7, o or '', round(ipo,2) if ipo else '',
        round(cpa) if cpa else '', round(bcpa) if bcpa else '', BUD.get(MAP.get(n,n), '—'), jud])
P.append(['カタログ全部（テスト）※全商品横断','—','—','—',round(R['CAT7'])]+['—']*18+[BUD.get('カタログ全部（テスト）','—'),'—'])
sh('商品別', ['商品別（7日 8/25-31 = 判定単位）。カタログ行の広告費は「7日合計」（前日ではない）',
  '分岐CPA(注文) = 1個あたり粗利 × 点数/注文。まとめ買い商品（点数/注文>1.2）は必ず注文ベースで比較する',
  'Σ商品広告費 + カタログ = Metaアカウント7日消化 1,747,353円（差0円・w0831.pyでassert済み）',
  '★3D足臭(8/29停止)・形状記憶日傘(8/25停止)・ビジュアル耳かき(8/31停止)は7日窓に残消化が入っている',
  '★形状記憶日傘・カタログ夏以外も停止済み。判定の対象外'],
 ['商品','区分','7日売上','7日原価','7日広告','7日利益','7日利益率','MER','分岐','目標','余裕','前日売上','前日利益','前日利益率',
  '3日利益','30日利益','当月累積利益','7日販売数','7日注文数','点数/注文','CPA(注文)','分岐CPA(注文)','現日予算','判定'],
 P, [26,6]+[12]*22, {3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',12:'#,##0',13:'#,##0',14:'0.0%',
                     15:'#,##0',16:'#,##0',17:'#,##0',21:'#,##0',22:'#,##0',23:'#,##0'})

# ---------- 5 レバー一覧 ----------
LV = []
for n in sorted(R['N7'], key=lambda x: -R['N7'][x]):
    cp = MAP.get(n, n); b = BUD.get(cp)
    if not b: continue
    s7, k7, a7, p7 = blk(n, R['N7'], R['K7'], R['A7'])
    if not a7: continue
    gm = 1-k7/s7; mer = s7/a7; mm = mer*0.7
    days = [d for d in D7 if d in BSNAP[cp]]
    wu = (a7/sum(BSNAP[cp][d] for d in days)) if len(days) == 7 and sum(BSNAP[cp][d] for d in days) else None
    ends = [mer*x*gm-1 for x in (0.6, 0.7, 0.8)]
    LV.append([n, '季節' if n in SEA else '通年', b, round(wu,3) if wu else '', round(mer,2), round(1/gm,2),
        round(mer-1/gm,2), round(mm*gm-1,3), round(ends[0],3), round(ends[2],3),
        '一致' if (all(e>0 for e in ends) or all(e<0 for e in ends)) else '⚠️符号が割れる',
        round((1-mm*gm)*b*0.25*7), round((mm*gm-1)*b*0.25*7), round(p7)])
LV.sort(key=lambda x: -max(x[11], x[12]))
sh('レバー一覧', ['「どれを動かすと一番効くか」は7日平均の利益額ではなく、動かしたときの変化量で並べる',
  '削減1円あたりの利益変化 = 1 − 限界MER×粗利率 ／ 増額1円あたり = 限界MER×粗利率 − 1（限界MER = 平均MER×0.7）',
  '★頑健性: ×0.6 と ×0.8 の両端で符号が変わらないものだけ動かしてよい。割れたら据え置き',
  '増額は週消化率95%以上のものだけが対象（それ未満は増やしても使われない）'],
 ['商品','区分','日予算','週消化率','MER','分岐','余裕','1円あたり利益(×0.7)','×0.6端','×0.8端','頑健性','−25%で週','+25%で週','7日利益'],
 LV, [26,6,11,10,8,8,8,15,10,10,13,12,12,12], {3:'#,##0',4:'0.0%',12:'#,##0',13:'#,##0',14:'#,##0'})

# ---------- 6 週次診断 ----------
def wblk(days, names=None):
    s = c = 0
    for n, v in Sd.items():
        if names is not None and n not in names: continue
        g = sum(v[d][0] for d in days if d in v); dc = sum(v[d][1] for d in days if d in v)
        if not g: continue
        s += g-dc
        if n in PRICEa and n in COSTa: c += g/PRICEa[n]*COSTa[n]
    a = sum(ADd[MAP.get(n, n)].get(d, 0) for n in (names if names is not None else Sd) for d in days) if names is not None \
        else sum(ADd[cp].get(d, 0) for cp in ADd for d in days)
    return s, c, a, s-c-a
YR = {n for n in Sd if n and n not in SEA}
end = datetime.date(2026, 8, 31); WK = []
for i in range(6):
    a = end - datetime.timedelta(days=7*i+6)
    WK.append((f"{a.strftime('%m/%d')}-{(a+datetime.timedelta(days=6)).strftime('%m/%d')}",
               [(a+datetime.timedelta(days=j)).isoformat() for j in range(7)]))
WK.reverse()
WD_ = []
for lbl, days in WK:
    if days[0] < '2026-06-29': continue
    s, c, a, p = wblk(days); ss, _, sa, sp = wblk(days, SEA); ys, _, ya, yp = wblk(days, YR)
    WD_.append([lbl, round(s), round(c), round(a), round(p), round(p/s,4), round(s/a,2), round(c/s,4), round(a/s,4),
                round(sp), round(ss/sa,2) if sa else '', round(yp), round(ys/ya,2) if ya else ''])
sh('週次診断', ['全店の週次分解。原価は日次のバリアント構成が取れないナノガラス/ディスペンサーのみ30日実測ミックスの加重平均単価【近似】',
  '★夏物の崩壊が全体の正体。壊れたのは季節物だけで、通年物は伸びている',
  '★最新週(8/25-31)は日予算253,000〜255,000で運転。日予算590,000だった7月とは水準が違うので、利益額の絶対比較はしない'],
 ['週','実売','原価','広告費','利益','利益率','MER','原価率','広告費率','季節物 利益','季節物MER','通年物 利益','通年物MER'],
 WD_, [14,13,12,12,12,9,8,9,10,15,11,15,11],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'#,##0',6:'0.0%',7:'0.00',8:'0.0%',9:'0.0%',10:'#,##0',11:'0.00',12:'#,##0',13:'0.00'})

# ---------- 7 ファネル（30日・RPM/CPM つき）----------
# ★★ Meta側は health_check（30日・全クリック）。Shopify側も同じ30日窓に揃えてある
for _k, _v in MF_RAW.items():
    assert abs(sum(ADd[_k].get(d, 0) for d in D30) - _v[0]) < 1, (_k, _v[0])
FN = []
for n, (c, imp, clk) in MF.items():
    s = R['N30'].get(n, 0); k = R['K30'].get(n, 0); q = R['Q30'].get(n, 0); o = ORD30.get(n, 0)
    if not s or not o or not clk: continue
    ipo = q/o; bcpa = (s-k)/q*ipo; cpm = c/imp*1000
    ctr = clk/imp; cvr = o/clk; gpo = (s-k)/o          # 粗利/注文
    need = cpm/(gpo*1000)                              # 必要 CTR×CVR
    FN.append([n, '季節' if n in SEA else '通年', imp, round(ctr,4), clk, round(cvr,4), round(s/clk),
               round(c/clk), round(c/o), round(bcpa), round(bcpa-c/o), round(cpm), round(gpo),
               round(ctr*cvr*10000, 2), round(need*10000, 2), round((ctr*cvr)/need, 2)])
mc = statistics.median([x[5] for x in FN]) if FN else 0; mr = statistics.median([x[6] for x in FN]) if FN else 0
FN.sort(key=lambda x: -x[15])
sh('ファネル30日', ['⚠️ このシートだけ窓が「30日(8/02-8/31)」。他シートの7日窓と混ぜて読まないこと',
  '⚠️ クリックは link_clicks ではなく **全クリック**。Supermetrics の data_query がハーネス側でブロックされており、'
  '通る health_check（30日・全クリック）から取ったため。過去レポートの「CVR順位」シートと数値を並べて比較しない',
  '✅ 検算: 14キャンペーンすべてで health_check の消化 = CSV(8/02-8/31) が差0円で一致（コード内assert）',
  'CVR = Shopify注文 ÷ Meta全クリック（Metaの購入数は損益に使わない）',
  f'中央値 CVR {mc*100:.2f}% ／ 売上perクリック {mr:.0f}円',
  '★RPM/CPM = CTR×CVR×粗利/注文 ÷ CPM。1.0未満が赤字圏。★CTR単独では成否を判別できない（実測で反証済み）',
  '★必要CTR×CVR(‱) = CPM ÷ (粗利/注文 × 1000)。この列を下回っている商品が赤字'],
 ['商品','区分','インプ','CTR','クリック','CVR','売上/クリック','CPC','CPA(注文)','分岐CPA(注文)','余裕','CPM','粗利/注文',
  '実測CTR×CVR(‱)','必要CTR×CVR(‱)','RPM/CPM'],
 FN, [26,6,12,9,11,9,13,9,12,15,11,10,11,15,15,10],
 {3:'#,##0',4:'0.00%',5:'#,##0',6:'0.00%',7:'#,##0',8:'#,##0',9:'#,##0',10:'#,##0',11:'#,##0',12:'#,##0',13:'#,##0'})

# ---------- 8 日次推移 ----------
WDJ = ['月','火','水','木','金','土','日']
pr = {}
for d in ALLD:
    s = c = 0
    for n, v in Sd.items():
        if n == '' or d not in v: continue
        g, dc = v[d]; s += g-dc
        if g and n in PRICEa and n in COSTa: c += g/PRICEa[n]*COSTa[n]
    a = sum(ADd[cp].get(d, 0) for cp in ADd); pr[d] = (s, c, a, s-c-a)
DD = []
for j, d in enumerate(ALLD):
    if d < '2026-07-13': continue
    s, c, a, p = pr[d]; ma = sum(pr[x][3] for x in ALLD[j-2:j+1])/3
    DD.append([d, WDJ[datetime.date.fromisoformat(d).weekday()], round(s), round(c), round(a), round(p),
               round(p/s,4), round(s/a,2), round(ma), '⚠️非常ブレーキ' if ma < 130000 else ''])
sh('日次推移', ['日次の利益は上記【近似】原価ベース。4期間サマリーはバリアント別実数量の正確値',
  '非常ブレーキ: 3日移動平均利益 < 130,000円 で発動（カタログ→赤字商品一律−25%）'],
 ['日付','曜','実売','原価','広告費','利益','利益率','MER','3日移動平均利益','警告'],
 DD, [12,5,13,12,12,12,9,8,16,16], {3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',8:'0.00',9:'#,##0'})

# ---------- 9 本日の判定 ----------
JUDGE = [
 ['9/1','ナノバブルシャワーヘッド 本判定','倍率 ≥1.3 継続 ／ 1.0〜1.3 −50% ／ <1.0 停止',
  '7日(8/25-31) 実売117,264円 / 17点 / 広告49,128円 → 実CPA 2,890円 vs 分岐CPA 4,725円 = **倍率1.64**',
  '✅ 合格 → 日8,000円で継続',
  '利益 +29,801円 / 3窓とも余裕+0.90〜1.21。★増額は日10,000円が上限'],
 ['9/1','決済ミックス 月次再計測','直近100注文の paymentGatewayNames × 金額',
  'SP 79.7% / KOMOJUスマホ 17.1% / Paidy 3.2%','✅ 完了 → 3.423%',
  '前回3.452%から −0.029pt。**Paidyが7.5%→3.2%と半減**。FEE定数とSKILL.mdを更新済み'],
 ['8/31','ビジュアル耳かき（前日の判定の結果）','3日倍率 0.70〜1.00 → 8,000→5,000円',
  '全期間(8/28-31) 実売28,884円 / 6点 / 広告26,408円 → 実CPA 4,401円 = **倍率0.76**',
  '⏹ ユーザーは**停止**を選択（Metaで PAUSED 確認）',
  '★私の提案は減額だったが、8/31を含めると倍率が0.97→0.76へ悪化しており停止のほうが正しかった。損益 −7,454円で終了'],
 ['9/2','カタログ 3日判定（8/31-9/2）','CPA基準: 倍率 ≥1.00 で据え置き（share基準は監視指標）',
  '前回(8/27-29)は 帰属19点・倍率1.12。8/31の消化17,032円','進行中',
  '注文UTM(utm_term=6966272711408)で3日ぶんを数え直して判定する'],
 ['9/3','快適マジックインソール 増額判定（30,000の実験）','P2(8/31-9/2)の1日あたり利益 ≥ 23,280円',
  'P2 1日目(8/31) = 実売42,984円 / 11点 / 広告26,154円 → **+9,383円**（基準の40%）','⚠️ 初日は基準割れ',
  '★月曜は最弱曜日（指数92.0）。P1(8/28-29)は金土なので曜日が不利。残り2日で判断する。'
  '「グリーン」全7サイズ在庫0の影響を判定時に併記する'],
 ['9/3','バランスケアスリッパ 増額判定','1日あたり利益 ≥ 増額前(8/23-29)の7日平均',
  '8/30に 10,000→12,000 へ増額。3窓とも余裕+1.07〜1.22','進行中','増額直後のロック中なので重ねない'],
 ['9/4','優先配送 790円','14日で件数 ≥ 28件','7日実測 **4件**（アタッチ率 0.55%）','🚨不合格の見込み',
  '前週6件から減少。8/22に536→790へ値上げ済み'],
]
sh('本日の判定', ['宣言済みの基準に当てはめるだけ。基準は宣言時のまま動かさない',
  '★本日の執行はゼロ。判定2件（ナノバブル本判定・決済ミックス再計測）はどちらも完了し、操作は不要',
  '★ナノバブル = 倍率1.64で合格・日8,000円で継続',
  '★要確認が1件: **偏光・調光サングラスが8/31に151円しか消化していない**（日予算5,000の3%）。'
  'ENABLEDのままなので広告の否認かエラーを疑う'],
 ['判定日','対象','宣言済み基準','実測','結果','備考'], JUDGE, [10,26,34,34,16,46])

# ---------- 9a 予算見直しチェック（8/30 昼にユーザーが6セット変更）----------
# 変更前 = data/budget-snapshots.csv の '2026-08-30'（朝）／ 変更後 = SNAP '2026-08-30b'（Meta直読）
PREVB = collections.defaultdict(int)
for r in body:
    if r[0] == '2026-08-30': PREVB[r[1]] += int(r[3])
BR, inc, dec = [], 0, 0
for cp in sorted(BUD, key=lambda c: -BUD[c]):
    p, nn = PREVB.get(cp, 0), BUD[cp]
    n = INV.get(cp, cp)
    s7 = R['N7'].get(n, 0); k7 = R['K7'].get(n, 0); a7 = R['A7'].get(n, 0)
    s3 = R['N3'].get(n, 0); k3 = R['K3'].get(n, 0); a3 = R['A3'].get(n, 0)
    q7 = R['Q7'].get(n, 0); q3 = R['Q3'].get(n, 0)
    gm = 1 - k7/s7 if s7 else None
    mer = s7/a7 if s7 and a7 else None; be = 1/gm if gm else None
    tg = 1/(gm-0.35) if gm and gm-0.35 > 0 else None
    m3 = s3/a3 if s3 and a3 else None; b3_ = 1/(1-k3/s3) if s3 else None
    r7 = ((s7-k7)/q7)/(a7/q7) if q7 and a7 else None
    r3 = ((s3-k3)/q3)/(a3/q3) if q3 and a3 else None
    dd = [d for d in D7 if BSNAP[cp].get(d)]
    w = (sum(ADd[cp].get(d, 0) for d in dd) / sum(BSNAP[cp][d] for d in dd)) if dd else None
    per = (mer*0.7*gm - 1) if mer and gm else None
    e6 = (mer*0.6*gm - 1) if mer and gm else None
    e8 = (mer*0.8*gm - 1) if mer and gm else None
    if nn > p: inc += nn - p
    elif nn < p: dec += p - nn
    BR.append([n if n != cp else cp, p, nn, nn-p, (nn/p-1) if p else '',
        round(mer, 2) if mer else '—', round(be, 2) if be else '—', round(tg, 2) if tg else '—',
        round(mer-be, 2) if mer and be else '—', round(m3, 2) if m3 else '—',
        round(m3-b3_, 2) if m3 and b3_ else '—', round(r7, 2) if r7 else '—', round(r3, 2) if r3 else '—',
        round(w, 3) if w else '—', round(per, 3) if per is not None else '—',
        ('一致' if (e6 and e8 and ((e6 > 0 and e8 > 0) or (e6 < 0 and e8 < 0))) else '⚠️符号が割れる') if per is not None else '—',
        '★変更' if p != nn else ''])
BR.sort(key=lambda x: (x[16] == '', -abs(x[3]), -x[2]))
sh('予算見直しチェック', ['2026-08-30 昼にユーザーが日予算を見直した。**変更後の値はMetaを直読して確認済み**'
  '（campaign_and_resource_get / campaign_detail_level="ad_groups"・ENABLEDの広告セットのみ合計）',
  '合計 253,000 → **255,000円**（+2,000円/日・+0.8%）。実質は「効率の良い3つへ +8,000／薄い3つから −6,000」の付け替え',
  '★増額 3件: 快適マジックインソール +3,000(+11.1%) ／ むくみ取りかっさ +3,000(+30.0%) ／ バランスケアスリッパ +2,000(+20.0%)',
  '★減額 3件: 2WAYシートボックス −2,000(−15.4%) ／ 完全遮光・接触冷感UVハット −2,000(−25.0%) ／ 偏光・調光サングラス −2,000(−28.6%)',
  '★総評: **方向は正しい**。増額した3つは7日余裕が +0.94〜+1.62 と全商品の上位、減額した3つは3日余裕が +0.04〜+0.57 と縮んでいる側',
  '⚠️ 週消化率はこのシートから計算方法を直した。同じ日に複数回スナップショットを取った日(8/23-25)を'
  '従来は合算しており、消化率が実態より低く出ていた。修正後、快適マジックインソールは 85%→**97.7%** になり、'
  '3窓判定が「据え置き」→「🚀3窓すべて目標超＋消化95%→+25%」に変わった'],
 ['商品/キャンペーン','変更前','変更後','差','変化率','7日MER','分岐','目標','7日余裕','3日MER','3日余裕',
  '倍率7日','倍率3日','消化率(7日)','1円あたり利益','頑健性','印'],
 BR, [26,10,10,9,9,8,7,7,9,8,9,9,9,12,13,14,7],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'+0.0%;-0.0%',14:'0.0%'})

# ---------- 10 CVR推移 ----------
SES = {'2026-08-04':6639,'2026-08-05':5425,'2026-08-06':5138,'2026-08-07':4827,'2026-08-08':5527,'2026-08-09':5886,
'2026-08-10':5245,'2026-08-11':6001,'2026-08-12':4900,'2026-08-13':5126,'2026-08-14':5035,'2026-08-15':4871,
'2026-08-16':5349,'2026-08-17':3438,'2026-08-18':3523,'2026-08-19':3190,'2026-08-20':3356,'2026-08-21':3120,
'2026-08-22':3220,'2026-08-23':3438,'2026-08-24':3192,'2026-08-25':3280,'2026-08-26':2733,'2026-08-27':3145,
'2026-08-28':3149,'2026-08-29':3537,'2026-08-30':3310}
ORDD = {'2026-08-04':188,'2026-08-05':152,'2026-08-06':157,'2026-08-07':173,'2026-08-08':184,'2026-08-09':187,
'2026-08-10':141,'2026-08-11':176,'2026-08-12':124,'2026-08-13':125,'2026-08-14':120,'2026-08-15':142,
'2026-08-16':143,'2026-08-17':98,'2026-08-18':89,'2026-08-19':106,'2026-08-20':113,'2026-08-21':110,
'2026-08-22':98,'2026-08-23':106,'2026-08-24':100,'2026-08-25':99,'2026-08-26':105,'2026-08-27':130,
'2026-08-28':85,'2026-08-29':115,'2026-08-30':114}
CV = [[d, WDJ[datetime.date.fromisoformat(d).weekday()], SES[d], ORDD[d], round(ORDD[d]/SES[d], 4),
       '★昨日（確定値）' if d == '2026-08-30' else ''] for d in sorted(SES)]
o1 = sum(ORDD[d] for d in sorted(SES) if d <= '2026-08-16'); s1_ = sum(SES[d] for d in sorted(SES) if d <= '2026-08-16')
o2 = sum(ORDD[d] for d in sorted(SES) if '2026-08-17' <= d <= '2026-08-30'); s2 = sum(SES[d] for d in sorted(SES) if '2026-08-17' <= d <= '2026-08-30')
sh('CVR推移', ['全店CVR = Shopify注文 ÷ Shopifyセッション。商品別レポートのCVR（注文÷Metaクリック）とは分母が違うので混ぜないこと',
  f'★8/04-16（広告費が大きかった時期）: {o1}注文 / {s1_:,}セッション = {o1/s1_:.2%}',
  f'★8/17-30（広告費を絞った後）: {o2}注文 / {s2:,}セッション = {o2/s2:.2%} → 広告費を削ったらCVRは **{o2/s2/(o1/s1_)-1:+.1%}** 改善している',
  '★8/29は セッション3,537（8/16以降で最多）に対し注文115件 = CVR 3.25%。セッションが増えた分CV率は薄まるが、'
  'カート追加221件はこの窓の最多で、需要そのものが戻っている',
  '★8/28は セッション3,149・注文85件 = CVR 2.70% でこの窓の最低。ただし翌日に戻っており単日のブレ',
  '★セッションは 5,000〜6,000/日 → 2,700〜3,500/日 へ減った。これは日予算 590,000→253,000 の直接の結果'],
 ['日付','曜','セッション','注文','全店CVR','備考'], CV, [12,5,12,10,10,20],
 {3:'#,##0',4:'#,##0',5:'0.00%'})

# ---------- 11 曜日指数 ----------
sh('曜日指数', ['直近30日(8/02-8/31)・祝日 8/11(山の日) を除外。売上=gross−値引',
  '★同曜日平均との比較は当面使わない。直近30日の土曜（8/01・8/08・8/15）が夏物崩壊前・日予算590,000時代で汚染されているため',
  '★代わりに「前週同曜日比」で読む。8/29(土)は前週土曜(8/22 469,248円)比 **+24.3%**',
  '★日曜が最強(125.4)・水曜が最弱(88.4)。この形は7月から変わっていない'],
 ['曜日','指数(全体=100)'], [[k, round(v, 1)] for k, v in sorted(R['IDX'].items(), key=lambda x: -x[1])], [10, 16])

wb.save('data/reports/report-2026-09-01.xlsx')
print('シート:', wb.sheetnames)
print(f"3日移動平均利益 {p3ma:,.0f}円 / 非常ブレーキ130,000円")
