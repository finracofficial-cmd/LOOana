# -*- coding: utf-8 -*-
import pickle, csv, datetime, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
R=pickle.load(open('/tmp/rep0811.pkl','rb'))
FEE=0.03452
# --- 8/7 予算スナップショット（8/6の実測値を引き継ぎ。W固定のみ 8/7 01:00 にMeta直読で20,000を確認）
rows=list(csv.reader(open('/home/user/LOOana/data/budget-snapshots.csv')))
hdr=rows[0]; body=[r for r in rows[1:] if r]
if not any(r[0]=='2026-08-11' for r in body):      # 既に8/07行があれば上書きしない（当日の変更を消さないため）
    for r in [x for x in body if x[0]=='2026-08-10']: body.append(['2026-08-11',r[1],r[2],r[3]])
with open('/home/user/LOOana/data/budget-snapshots.csv','w',newline='') as f:
    w=csv.writer(f); w.writerow(hdr); w.writerows(body)
BUD={r[2]:int(r[3]) for r in body if r[0]=='2026-08-11'}
TOTB=sum(BUD.values())
wb=Workbook(); H=Font(bold=True,color='FFFFFF'); FILL=PatternFill('solid',fgColor='2F5597')
def sheet(t,cols,data,widths=None):
    ws=wb.create_sheet(t); ws.append(cols)
    for c in ws[1]: c.font=H; c.fill=FILL; c.alignment=Alignment(horizontal='center')
    for r in data: ws.append(r)
    for i,w_ in enumerate(widths or [16]*len(cols),1): ws.column_dimensions[chr(64+i) if i<27 else 'A'+chr(38+i)].width=w_
    ws.freeze_panes='A2'; return ws
wb.remove(wb.active)
# 1 全体サマリー
S=[]
for k,lbl,N,K,ACC,days in [('d1','前日 8/09(日)',R['N1'],R['K1'],R['ACCT1'],R['D1']),
    ('d3','3日 8/07-09',R['N3'],R['K3'],R['ACCT3'],R['D3']),
    ('d7','7日 8/03-09',R['N7'],R['K7'],R['ACCT7'],R['D7']),
    ('d30','30日 7/11-8/09',R['N30'],R['K30'],R['ACCT30'],R['D30'])]:
    s=sum(N.values()); c=sum(K.values()); p=s-c-ACC
    S.append([lbl,s,c,ACC,p,round(p/s*100,2),round(s/ACC,2),round(1/(1-c/s),2),round(s*FEE),round(p-s*FEE)])
sheet('全体サマリー',['窓','実売(gross-値引)','原価','広告費','利益','利益率%','MER','分岐MER','決済手数料','手数料控除後利益'],S,[16,16,14,14,14,10,8,9,13,16])
# 2 商品別
P=[]
for n in sorted(R['N7'],key=lambda x:-R['N7'][x]):
    s7,c7=R['N7'][n],R['K7'][n]; a7=R['A7'].get(n,0)
    s1,c1,a1=R['N1'].get(n,0),R['K1'].get(n,0),R['A1'].get(n,0)
    s3,c3,a3=R['N3'].get(n,0),R['K3'].get(n,0),R['A3'].get(n,0)
    s30,c30,a30=R['N30'].get(n,0),R['K30'].get(n,0),R['A30'].get(n,0)
    p7=s7-c7-a7; gm=1-c7/s7 if s7 else 0
    mer=s7/a7 if a7 else None; be=1/gm if gm else None; yo=(mer-be) if mer and be else None
    bud=BUD.get(R['MAP'].get(n,n))
    jud=('📦広告なし' if not a7 else ('🚨赤字' if p7<0 else ('🔧テコ入れ' if yo is not None and yo<0.2 else ('🚀増額候補' if yo and yo>1.5 else '✅維持'))))
    P.append([n,s7,c7,round(a7),round(p7),round(mer,2) if mer else '—',round(be,2) if be else '—',
              round(yo,2) if yo is not None else '—',s1,round(s1-c1-a1),
              round((s1-c1-a1)/s1*100,1) if s1 else '—',round(s3-c3-a3),round(s30-c30-a30),
              R['Q7'].get(n,0),bud if bud else '—',jud])
P.append(['カタログ全部（テスト）※全商品横断','—','—',round(R['CAT7']),'—','—','—','—','—','—','—','—','—','—',BUD.get('カタログ全部（テスト）','—'),'—'])
sheet('商品別',['商品','7日売上','7日原価','7日広告','7日利益','MER','分岐','余裕','前日売上','前日利益','前日利益率%','3日利益','30日利益','7日販売数','現日予算','判定'],P,[26]+[12]*15)
# 3 日次推移
COST,PRICE=R['COST'],R['PRICE']
B='/home/user/LOOana/data/daily/'
Sd=collections.defaultdict(lambda: collections.defaultdict(lambda:[0,0]))
for r in csv.DictReader(open(B+'daily_sales.csv',encoding='utf-8')):
    Sd[r['name']][r['date']][0]+=int(r['gross']); Sd[r['name']][r['date']][1]+=int(r['disc'])
ADd=collections.defaultdict(float)
for r in csv.DictReader(open(B+'daily_ad.csv',encoding='utf-8')): ADd[r['date']]+=float(r['cost'])
WD=['月','火','水','木','金','土','日']
days=sorted({d for v in Sd.values() for d in v}); pr={}
for d in days:
    s=c=0
    for n,v in Sd.items():
        if n=='' or d not in v: continue
        g,dc=v[d]; s+=g-dc
        if g==0: continue
        if n=='ナノガラス脱毛パッド': c+=round(g/3980)*749
        elif n in PRICE and n in COST: c+=round(g/PRICE[n])*COST[n]
    pr[d]=(s,c,ADd[d],s-c-ADd[d])
DD=[]
for i,d in enumerate(days):
    if d<'2026-07-24': continue
    j=days.index(d); ma=sum(pr[x][3] for x in days[j-2:j+1])/3 if j>=2 else None
    s,c,a,p=pr[d]
    DD.append([d,WD[datetime.date(*map(int,d.split('-'))).weekday()],round(s),round(a),round(p),
               round(p/s*100,2),round(ma) if ma else '—'])
sheet('日次推移',['日付','曜','実売','広告費','利益(概算)','利益率%','3日移動平均利益'],DD,[12,5,14,13,13,10,16])
# 4 曜日指数
sheet('曜日指数',['曜日','指数(全体100)'],[[k,round(v,1)] for k,v in sorted(R['IDX'].items(),key=lambda x:-x[1])],[10,14])
wb.save('/home/user/LOOana/data/reports/report-2026-08-11.xlsx')
print('saved. 日予算合計(8/07)=',f'{TOTB:,}')
