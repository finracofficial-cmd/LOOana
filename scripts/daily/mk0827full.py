# -*- coding: utf-8 -*-
"""2026-08-27 定例レポート本体（昨日=8/26 水）。"""
import pickle, csv, datetime, collections, statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
R = pickle.load(open('/tmp/rep0826w.pkl', 'rb')); FEE = 0.03452
D1, D3, D7, D30, CUM = R['D1'], R['D3'], R['D7'], R['D30'], R['CUM']
SNAP = '2026-08-26'
rows = list(csv.reader(open('data/budget-snapshots.csv'))); body = [r for r in rows[1:] if r]
assert any(r[0] == SNAP for r in body), f'{SNAP} スナップショット未記録'
BUD = collections.defaultdict(int)
for r in body:
    if r[0] == SNAP: BUD[r[1]] += int(r[3])
assert sum(BUD.values()) == 250000, sum(BUD.values())
BSNAP = collections.defaultdict(lambda: collections.defaultdict(int))
for r in body: BSNAP[r[1]][r[0].rstrip('b')] += int(r[3])

MAP = R['MAP']; INV = {v: k for k, v in MAP.items()}
# --- 7日(8/19-25) Shopify注文数（ShopifyQL GROUP BY product_title 実測）---
ORD = {'ナノガラス脱毛パッド':266,'W固定スマホ車載ホルダー':124,'快適マジックインソール':78,'ムダ毛シェーバー':45,
'2WAYシートボックス':37,'形状記憶日傘':32,'偏光・調光サングラス':27,'完全遮光・接触冷感UVハット':22,
'姿勢サポートチェア':21,'バランスケアスリッパ':20,'むくみ取りかっさ':18,'3D足臭リセットブラシ':16,'優先配送':13,
'4WAY':10,'完全遮光・形状記憶':8,'接触冷感UVアームカバー':6,'携帯電動シェーバー':3,'接触冷感UVパーカー':2,
'ナノバブルシャワーヘッド':2,'健康サンダル':1,'ヘアドライタオル':1,'湯上がりガーゼワンピース':1}
STORE_ORD7, STORE_SESS7 = 731, 22339
assert sum(ORD.values()) == 753, sum(ORD.values())
# --- 7日(8/19-25) Metaファネル（キャンペーン別実測: cost, impressions, link_clicks）---
MF_RAW = {'ナノガラス脱毛パッド':(553391,263418,6441),'W固定スマホ車載ホルダー':(242824,109854,3161),
'形状記憶日傘':(142729,33763,1369),'快適マジックインソール':(133915,84077,1908),'ムダ毛シェーバー':(124774,93503,1182),
'2WAYシートボックス':(93237,26210,796),'偏光・調光サングラス':(82481,32210,926),'姿勢サポートチェア':(81820,46072,879),
'3D足臭リセットブラシ':(64719,16010,269),'完全遮光・接触冷感UVハット':(63960,13061,612),
'バランスケアスリッパ':(46208,27649,565),'むくみ取りかっさ':(35189,8755,302),
'4WAY 取り付けOK 小型瞬間冷却ハンディファン':(31471,7659,220),'接触冷感UVアームカバー':(23978,5246,184),
'ナノバブルシャワーヘッド':(9670,4297,148),'1秒折り畳みチェア':(4133,1211,36),'接触冷感UVパーカー':(2520,667,21)}
MF = {INV.get(k, k): v for k, v in MF_RAW.items()}

SEA = {'4WAY','3WAYサーキュレーター','卓上冷感クーラー','5WAY腰掛けファン','瞬間冷却ハンディファン','接触冷感UVパーカー',
'接触冷感UVアームカバー','瞬間冷感ポンチョ','完全遮光・接触冷感UVハット','形状記憶日傘','完全遮光・形状記憶',
'偏光・調光サングラス','害虫ブロッカー','湯上がりガーゼワンピース'}

wb = Workbook(); wb.remove(wb.active)
TH = Font(name='Arial', bold=True, color='FFFFFF', size=10); TD = Font(name='Arial', size=10)
NEG = Font(name='Arial', size=10, color='CC0000'); HEAD = PatternFill('solid', fgColor='305496')
thin = Border(*[Side(style='thin', color='CCCCCC')] * 4)

def sh(title, note, cols, data, widths, fmt=None):
    ws = wb.create_sheet(title); r = 1
    if note:
        for ln in note: ws.cell(r, 1, ln).font = Font(name='Arial', size=10, bold=(r == 1)); r += 1
        r += 1
    for c, v in enumerate(cols, 1): ws.cell(r, c, v)
    for c in range(1, len(cols) + 1):
        x = ws.cell(r, c); x.font = TH; x.fill = HEAD; x.border = thin
        x.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 30; hr = r; r += 1
    for row in data:
        for c, v in enumerate(row, 1):
            x = ws.cell(r, c, v); x.border = thin; x.font = TD
            if isinstance(v, (int, float)) and v < 0: x.font = NEG
            if fmt and c in fmt: x.number_format = fmt[c]
        r += 1
    for i, w_ in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i < 27 else 'A' + chr(38 + i)].width = w_
    ws.freeze_panes = ws.cell(hr + 1, 1).coordinate
    return ws

# ===== 日次CSVを読む（3窓判定・週次・日次推移で使う）=====
Sd = collections.defaultdict(lambda: collections.defaultdict(lambda: [0, 0]))
for r_ in csv.DictReader(open('data/daily/daily_sales.csv')):
    Sd[r_['name']][r_['date']][0] += int(r_['gross']); Sd[r_['name']][r_['date']][1] += int(r_['disc'])
ADd = collections.defaultdict(lambda: collections.defaultdict(float))
for r_ in csv.DictReader(open('data/daily/daily_ad.csv')): ADd[r_['campaign']][r_['date']] += float(r_['cost'])
COSTa = dict(R['COST']); PRICEa = dict(R['PRICE'])
COSTa['ナノガラス脱毛パッド'] = (855 * 757 + 373 * 740) / 1228          # 30日実測ミックスの加重平均【近似】
COSTa['壁掛けディスペンサー'] = (41 * 2528 + 13 * 1981) / 54
PRICEa['壁掛けディスペンサー'] = 363920 / 54
ALLD = sorted({d for v in Sd.values() for d in v if d <= D1[0]})
D5 = ALLD[-5:]
def win(n, days):
    """商品nのその窓の（実売, 原価【近似】, 広告費）"""
    g = sum(Sd[n][d][0] for d in days if d in Sd[n]); dc = sum(Sd[n][d][1] for d in days if d in Sd[n])
    c = g / PRICEa[n] * COSTa[n] if g and n in PRICEa and n in COSTa else 0
    a = sum(ADd[MAP.get(n, n)].get(d, 0) for d in days)
    return g - dc, c, a

# ---------- 1 シンプル判定（3窓判定つき） ----------
def blk(n, N, K, A_):
    s = N.get(n, 0); k = K.get(n, 0); a = A_.get(n, 0); return s, k, a, s - k - a
SIMPLE = []
for n in sorted(R['N7'], key=lambda x: -R['N7'][x]):
    a7 = R['A7'].get(n, 0)
    if not a7: continue
    s7, k7, _, p7 = blk(n, R['N7'], R['K7'], R['A7'])
    s3, k3, a3, p3 = blk(n, R['N3'], R['K3'], R['A3']); s1, k1, a1, p1 = blk(n, R['N1'], R['K1'], R['A1'])
    gm = 1 - k7 / s7; be = 1 / gm; tg = 1 / (gm - 0.35) if gm - 0.35 > 0 else None; mer = s7 / a7
    cp = MAP.get(n, n); days = [d for d in D7 if d in BSNAP[cp]]
    wu = (a7 / sum(BSNAP[cp][d] for d in days)) if len(days) == 7 and sum(BSNAP[cp][d] for d in days) else None
    # ★3窓判定（3日/5日/7日・すべて前日終端）
    w3 = []
    for dd in (D3, D5, D7):
        sx, cx, ax = win(n, dd)
        w3.append((sx / ax if ax else None, 1 / (1 - cx / sx) if sx and sx > cx else None))
    below = all(m is not None and b is not None and m < b for m, b in w3)
    tight = all(m is not None and b is not None and m - b < 0.30 for m, b in w3)
    rich = all(m is not None and tg and m >= tg for m, _ in w3) and wu is not None and wu >= 0.95
    tri = '🚨3窓すべて分岐割れ→停止/−50%' if below else ('⚠️3窓すべて余裕<0.30→−25%' if tight else
          ('🚀3窓すべて目標超＋消化95%→+25%' if rich else '⏸窓が割れた→据え置き'))
    if p7 < 0 or mer < be: st = '🚨縮小・停止'
    elif (a3 and s3 / a3 < be) or mer < be + 0.30: st = '🔧改善'
    elif tg and mer >= tg and wu and wu >= 0.95: st = '🚀伸ばす候補'
    else: st = '✅維持'
    mm = mer * 0.7; per = mm * gm - 1
    # 頑健性: 限界MER = 平均MER×0.6/0.7/0.8 の3端で符号が変わらないか
    ends = [mer * x * gm - 1 for x in (0.6, 0.7, 0.8)]
    rob = '一致' if (all(e > 0 for e in ends) or all(e < 0 for e in ends)) else '⚠️符号が割れる'
    SIMPLE.append([n, round(mer, 2), round(s3 / a3, 2) if a3 else '', round(s1 / a1, 2) if a1 else '',
        round(be, 2), round(tg, 2) if tg else '', round(p7), round(p3), round(p1),
        round(wu, 3) if wu else '', round(per, 3), rob, st, tri])
sh('シンプル判定', ['LOOTY 2026-08-27 定例（昨日=8/26 水）',
 '状態: 🚨縮小・停止=7日MER<分岐 or 7日利益マイナス ／ 🔧改善=3日MER<分岐 or 余裕<0.30 ／ 🚀伸ばす候補=7日MER≥目標かつ週消化率≥95% ／ ✅維持',
 '3窓判定（3日/5日/7日・すべて8/26終端）が本番の意思決定ルール。窓が1つでも割れたら据え置き＝動かさない',
 '「広告費1円あたり利益」= 限界MER×粗利率 − 1（限界MER = 平均MER×0.7）。頑健性は×0.6/0.7/0.8 の3端で符号が変わらないかを見る'],
 ['商品','7日MER','3日MER','前日MER','分岐','目標','7日利益','3日利益','前日利益','週消化率','1円あたり利益','頑健性','状態','3窓判定'],
 SIMPLE, [26,9,9,9,8,8,12,12,12,10,12,13,14,30], {7:'#,##0',8:'#,##0',9:'#,##0',10:'0.0%'})

# ---------- 2 ネクストアクション ----------
p3ma = (sum(R['N3'].values()) - sum(R['K3'].values()) - R['ACCT3']) / 3
NA = [
 ['🎉 変化点','非常ブレーキを9日ぶりに上抜けた',
  f'3日移動平均利益 {p3ma:,.0f}円 ≧ 発動ライン130,000円。8/18から9日連続でライン割れだったのが8/26で解消。'
  '前日利益 +164,546円・利益率30.3% はこの窓の最高値。要因は広告費（231,515円＝日予算の92.6%）を絞ったまま実売が戻ったこと',
  '—','毎日'],
 ['🔴 判定日','快適マジックインソール 増額判定',
  '宣言済み基準「1日あたり利益 ≥ 23,589円」に対し 7日利益+198,171円 = **1日28,310円で合格**。'
  'MER 3.03 > 目標2.13、直近3日の消化率96.2%（22,266/22,161/19,084 ÷ 22,000）で天井に接している。'
  '増額の1円あたり利益は限界MER×0.6/0.7/0.8 の3端すべて正（+0.488〜+0.984）',
  '週+約19,000円','本日'],
 ['🔴 判定日','ムダ毛シェーバー 8/24減額の効果判定',
  '減額前(8/21-23) 1日あたり利益 +8,609円 → 減額後(8/24-26) **+11,347円**。削った帯の限界MER 0.15 は分岐1.67を大きく下回る＝'
  '**削って正解**。Δ広告 −3,204円/日で測定下限3,000円をわずかに超えている（標本は小さい）',
  '—','本日'],
 ['🚨 要対応','3D足臭リセットブラシ',
  '3窓すべて分岐割れ（3日0.85 / 5日0.83 / 7日1.08 vs 分岐1.19）。削減1円あたりの利益変化は3端すべて正（+0.270〜+0.452）で符号が一致。'
  'CR「B」を止めた後のA単独 8/26 も MER 0.41 で最悪日。A単独4日（8/20-22・8/26）を合算しても MER 1.085 < 分岐',
  '週+約12,000円','本日'],
 ['🔴 判定日','ナノガラス コピー修正判定',
  '宣言済み基準「点数/注文 1.06 → 1.15」。7日実測 296個/266注文 = **1.113**（8/26朝は1.10）。判定日は明日8/28',
  '—','8/28'],
 ['🔴 判定日','ナノバブルシャワーヘッド 中間ゲート',
  '宣言済み基準「3日累計 Shopify実注文 ≥ 5件」。8/25 0件 → 8/26 **2件**。残り2日で3件必要。'
  '帰属は注文UTM（utm_campaign = 6934285209608）で測る','—','8/28'],
 ['🔴 判定日','カタログ 減額判定',
  '宣言済み基準「ラストクリック注文シェア < 消化シェアの0.8倍 → 10,000→5,000」。'
  '8/26実測は 注文シェア2.9%（68件中2件）÷ 消化シェア7.35% = **0.40倍**で基準を大きく下回っている。'
  '⚠️ 宣言済みの非常ブレーキ「カタログ 18,000 → 10,000」は依然として未実行','—','8/29'],
 ['🔴 判定日','3D足臭 CR A単独判定',
  '8/26からA単独のクリーン窓が開始（8/26-30）。RPM/CPM ≥1.3 増額 ／ <1.0 停止 ／ 1.0〜1.3 据え置き。'
  '★上の「要対応」で予算を半減しても、CRテスト自体は継続される（率で判定するため半減で無効化されない）','—','8/30'],
 ['🔴 判定日','4WAY・接触冷感UVアームカバー 停止判定',
  '宣言済み基準「全店週次利益(8/24-30) ≥ 696,901円」。8/24-26 の3日で **408,634円**（7日換算 953,479円）で大きく上回るペース。'
  '両方とも8/23が最終消化で停止済み','—','8/30'],
 ['🔴 判定日','ナノバブルシャワーヘッド 本判定',
  '倍率 ≥1.3 継続 ／ 1.0〜1.3 は−50% ／ <1.0 停止。★増額は日10,000円が上限（3〜5月にリーチを2倍に広げて壊した実測がある）','—','9/1'],
 ['🔴 判定日','優先配送 790円',
  '宣言済み基準「14日で件数 ≥ 28件」。7日実測 **13件**（アタッチ率 13/731 = 1.78%）。'
  '8/22に536→790へ値上げ済みで、8/26も790円で計上を確認','—','9/4'],
 ['🟡 推奨','アップセル手動ペア5組の設定',
  'data/lp/upsell-pairs-2026-08-26.txt の5組をアプリに手入力する。異商品ミックス率が実質ゼロ（1.78%）なので'
  '自動推薦は学習データがなく全店売れ筋にフォールバックする。判定は実施+14日で異商品ミックス率3.0%以上','週+約4万円','任意'],
 ['🟡 推奨','カテゴリタグ40件の付与',
  'data/tags/category-mutation-2026-08-23.graphql を GraphiQL で1回実行 ＋ ナノバブルの季節タグ是正','—','任意'],
 ['🟡 推奨','コレクション冒頭文7本の貼り付け',
  'data/lp/collection-intro-2026-08-23.txt。あわせて「暖房・防寒グッズ」コレクションのAmazon由来HTML説明を削除','—','任意'],
 ['🟡 推奨','購入通知アプリ',
  'Judge.me の Pop-up Reviews Widget は導入済みアプリの機能なので追加0円。Sales Popを足すならQikify Basic $6.99（訪問者無制限）','—','任意'],
 ['🟡 推奨','CJで原価確認',
  'あったかインソール 1,000円以下 ／ 外反母趾サポーター 800円以下 なら秋冬のテスト枠に入れる','—','任意'],
]
sh('ネクストアクション', ['今日やること・判定カレンダー・持ち越しタスク（実施が確認できるまで毎日残す）'],
   ['区分','商品/対象','内容','効果','期限'], NA, [12, 26, 78, 16, 10])

# ---------- 3 全体サマリー ----------
SUMR = []
for lbl, N, K, ACC in [('前日 8/26(水)', R['N1'], R['K1'], R['ACCT1']), ('3日 8/24-26', R['N3'], R['K3'], R['ACCT3']),
        ('7日 8/20-26', R['N7'], R['K7'], R['ACCT7']), ('30日 7/28-8/26', R['N30'], R['K30'], R['ACCT30']),
        ('★当月累積 8/01-26', R['NC'], R['KC'], R['ACCTC'])]:
    s = sum(N.values()); c = sum(K.values()); p = s - c - ACC
    SUMR.append([lbl, s, c, round(ACC), round(c + ACC), round((c + ACC) / s, 4), round(p), round(p / s, 4),
                 round(s / ACC, 2), round(1 / (1 - c / s), 2), round(s * FEE), round(p - s * FEE), round((p - s * FEE) / s, 4)])
sh('全体サマリー', ['全体サマリー（売上=Shopify gross−値引・広告費=Meta実測。返品は売上からも利益からも控除しない=A案）',
 '総合原価＝原価＋広告費。恒等式 総合原価率＋利益率＝100% が全窓で成立していることを確認済み',
 '縦照合3本すべて通過: Σ商品gross=全店513,470 ／ Σ値引=14,931 ／ Σキャンペーン広告費=アカウント実消化',
 '決済ブレンド率 3.452%（2026-08-01再計測: SP76.2%×3.25% ＋ KOMOJUスマホ16.3%×4.1% ＋ Paidy7.5%×4.1%）',
 f'7日の全店CVR（Shopifyセッション基準）= {STORE_ORD7}注文 ÷ {STORE_SESS7:,}セッション = {STORE_ORD7/STORE_SESS7:.2%}'],
 ['窓','実売','原価','広告費','総合原価','総合原価率','利益','利益率','MER','分岐MER','決済手数料','手数料控除後利益','手数料後利益率'],
 SUMR, [20,14,13,13,13,11,13,10,8,9,12,15,12],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'#,##0',6:'0.0%',7:'#,##0',8:'0.0%',9:'0.00',10:'0.00',11:'#,##0',12:'#,##0',13:'0.0%'})
ws = wb['全体サマリー']; r = ws.max_row + 3
ws.cell(r, 1, '■ 昨日(8/26 水) × 期間比較').font = Font(name='Arial', bold=True, size=11); r += 1
S1 = sum(R['N1'].values()); C1 = sum(R['K1'].values()); A1 = R['ACCT1']
per = [('実売', S1, sum(R['N3'].values())/3, sum(R['N7'].values())/7, sum(R['N30'].values())/30),
       ('原価', C1, sum(R['K3'].values())/3, sum(R['K7'].values())/7, sum(R['K30'].values())/30),
       ('広告費', A1, R['ACCT3']/3, R['ACCT7']/7, R['ACCT30']/30),
       ('利益', S1-C1-A1, (sum(R['N3'].values())-sum(R['K3'].values())-R['ACCT3'])/3,
        (sum(R['N7'].values())-sum(R['K7'].values())-R['ACCT7'])/7,
        (sum(R['N30'].values())-sum(R['K30'].values())-R['ACCT30'])/30)]
for c, v in enumerate(['指標','8/26(水)','3日平均/日','7日平均/日','30日平均/日','vs7日平均'], 1): ws.cell(r, c, v)
for c in range(1, 7):
    x = ws.cell(r, c); x.font = TH; x.fill = HEAD; x.border = thin; x.alignment = Alignment(horizontal='center')
r += 1
for lbl, d1, m3, m7, m30 in per:
    for c, v in enumerate([lbl, round(d1), round(m3), round(m7), round(m30), round(d1/m7-1, 4)], 1):
        x = ws.cell(r, c, v); x.border = thin; x.font = NEG if (isinstance(v, (int, float)) and v < 0) else TD
        if c in (2,3,4,5): x.number_format = '#,##0'
        if c == 6: x.number_format = '+0.0%;-0.0%'
    r += 1
ws.cell(r+1, 1, '⚠️ 「同曜日平均比 −37.9%」は読まないこと。直近30日の月曜（7/27・8/3・8/10・8/17）が夏物崩壊前で汚染されている。'
  f'実態は7日平均比 +2.4%。3日移動平均利益 {p3ma:,.0f}円 は非常ブレーキ130,000円を下回っているが、'
  '前日利益134,495円は床を上回った').font = Font(name='Arial', size=10, color='CC0000')

# ---------- 4 商品別 ----------
P = []
for n in sorted(R['N7'], key=lambda x: -R['N7'][x]):
    s7,k7,a7,p7 = blk(n,R['N7'],R['K7'],R['A7']); s1,k1,a1,p1 = blk(n,R['N1'],R['K1'],R['A1'])
    s3,k3,a3,p3 = blk(n,R['N3'],R['K3'],R['A3']); s30,k30,a30,p30 = blk(n,R['N30'],R['K30'],R['A30'])
    sc,kc,ac,pc = blk(n,R['NC'],R['KC'],R['AC'])
    gm = 1-k7/s7 if s7 else 0; mer = s7/a7 if a7 else None; be = 1/gm if gm else None
    tg = 1/(gm-0.35) if gm-0.35 > 0 else None; yo = (mer-be) if mer and be else None
    q7 = R['Q7'].get(n, 0); o = ORD.get(n); ipo = q7/o if o else None
    bcpa = (s7-k7)/q7*ipo if q7 and ipo else None; cpa = a7/o if o and a7 else None
    jud = ('📦広告なし' if not a7 else ('🚨赤字' if p7 < 0 else ('🔧テコ入れ' if yo is not None and yo < 0.30 else
           ('🚀増額候補' if tg and mer >= tg else '✅維持'))))
    P.append([n, '季節' if n in SEA else '通年', s7, k7, round(a7), round(p7), round(p7/s7,4) if s7 else '',
        round(mer,2) if mer else '—', round(be,2) if be else '—', round(tg,2) if tg else '—',
        round(yo,2) if yo is not None else '—', s1, round(p1), round(p1/s1,4) if s1 else '',
        round(p3), round(p30), round(pc), q7, o or '', round(ipo,2) if ipo else '',
        round(cpa) if cpa else '', round(bcpa) if bcpa else '', BUD.get(MAP.get(n,n), '—'), jud])
P.append(['カタログ全部（テスト）※全商品横断','—','—','—',round(R['CAT7'])]+['—']*18+[BUD.get('カタログ全部（テスト）','—'),'—'])
sh('商品別', ['商品別（7日 8/20-26 = 判定単位）。カタログ行の広告費は「7日合計」（前日ではない）',
  '分岐CPA(注文) = 1個あたり粗利 × 点数/注文。まとめ買い商品（点数/注文>1.2）は必ず注文ベースで比較する',
  'Σ商品広告費 + カタログ = Metaアカウント7日消化（差0円・w0824.pyでassert済み）',
  '★3D足臭の点数/注文は 15/13 = 1.15（n=13）。以前 n=1 で 2.00 と報告したのは誤り'],
 ['商品','区分','7日売上','7日原価','7日広告','7日利益','7日利益率','MER','分岐','目標','余裕','前日売上','前日利益','前日利益率',
  '3日利益','30日利益','当月累積利益','7日販売数','7日注文数','点数/注文','CPA(注文)','分岐CPA(注文)','現日予算','判定'],
 P, [26,6]+[12]*22, {3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',12:'#,##0',13:'#,##0',14:'0.0%',
                     15:'#,##0',16:'#,##0',17:'#,##0',21:'#,##0',22:'#,##0',23:'#,##0'})

# ---------- 5 レバー一覧 ----------
LV = []
for n in sorted(R['N7'], key=lambda x: -R['N7'][x]):
    cp = MAP.get(n, n); b = BUD.get(cp)
    if not b: continue
    s7, k7, a7, p7 = blk(n, R['N7'], R['K7'], R['A7'])
    if not a7: continue
    gm = 1-k7/s7; mer = s7/a7; mm = mer*0.7
    days = [d for d in D7 if d in BSNAP[cp]]
    wu = (a7/sum(BSNAP[cp][d] for d in days)) if len(days) == 7 and sum(BSNAP[cp][d] for d in days) else None
    ends = [mer*x*gm-1 for x in (0.6, 0.7, 0.8)]
    LV.append([n, '季節' if n in SEA else '通年', b, round(wu,3) if wu else '', round(mer,2), round(1/gm,2),
        round(mer-1/gm,2), round(mm*gm-1,3), round(ends[0],3), round(ends[2],3),
        '一致' if (all(e>0 for e in ends) or all(e<0 for e in ends)) else '⚠️符号が割れる',
        round((1-mm*gm)*b*0.25*7), round((mm*gm-1)*b*0.25*7), round(p7)])
LV.sort(key=lambda x: -max(x[11], x[12]))
sh('レバー一覧', ['「どれを動かすと一番効くか」は7日平均の利益額ではなく、動かしたときの変化量で並べる',
  '削減1円あたりの利益変化 = 1 − 限界MER×粗利率 ／ 増額1円あたり = 限界MER×粗利率 − 1（限界MER = 平均MER×0.7）',
  '★頑健性: ×0.6 と ×0.8 の両端で符号が変わらないものだけ動かしてよい。割れたら据え置き',
  '増額は週消化率95%以上のものだけが対象（それ未満は増やしても使われない）'],
 ['商品','区分','日予算','週消化率','MER','分岐','余裕','1円あたり利益(×0.7)','×0.6端','×0.8端','頑健性','−25%で週','+25%で週','7日利益'],
 LV, [26,6,11,10,8,8,8,15,10,10,13,12,12,12], {3:'#,##0',4:'0.0%',12:'#,##0',13:'#,##0',14:'#,##0'})

# ---------- 6 週次診断 ----------
def wblk(days, names=None):
    s = c = 0
    for n, v in Sd.items():
        if names is not None and n not in names: continue
        g = sum(v[d][0] for d in days if d in v); dc = sum(v[d][1] for d in days if d in v)
        if not g: continue
        s += g-dc
        if n in PRICEa and n in COSTa: c += g/PRICEa[n]*COSTa[n]
    a = sum(ADd[MAP.get(n, n)].get(d, 0) for n in (names if names is not None else Sd) for d in days) if names is not None \
        else sum(ADd[cp].get(d, 0) for cp in ADd for d in days)
    return s, c, a, s-c-a
YR = {n for n in Sd if n and n not in SEA}
end = datetime.date(2026, 8, 25); WK = []
for i in range(6):
    a = end - datetime.timedelta(days=7*i+6)
    WK.append((f"{a.strftime('%m/%d')}-{(a+datetime.timedelta(days=6)).strftime('%m/%d')}",
               [(a+datetime.timedelta(days=j)).isoformat() for j in range(7)]))
WK.reverse()
WD_ = []
for lbl, days in WK:
    if days[0] < '2026-06-29': continue
    s, c, a, p = wblk(days); ss, _, sa, sp = wblk(days, SEA); ys, _, ya, yp = wblk(days, YR)
    WD_.append([lbl, round(s), round(c), round(a), round(p), round(p/s,4), round(s/a,2), round(c/s,4), round(a/s,4),
                round(sp), round(ss/sa,2) if sa else '', round(yp), round(ys/ya,2) if ya else ''])
sh('週次診断', ['全店の週次分解。原価は日次のバリアント構成が取れないナノガラス/ディスペンサーのみ30日実測ミックスの加重平均単価【近似】',
  '★夏物の崩壊が全体の正体。壊れたのは季節物だけで、通年物は伸びている',
  '★最新週(8/18-24)は日予算256,000〜258,000で運転。日予算590,000だった7月とは水準が違うので、利益額の絶対比較はしない'],
 ['週','実売','原価','広告費','利益','利益率','MER','原価率','広告費率','季節物 利益','季節物MER','通年物 利益','通年物MER'],
 WD_, [14,13,12,12,12,9,8,9,10,15,11,15,11],
 {2:'#,##0',3:'#,##0',4:'#,##0',5:'#,##0',6:'0.0%',7:'0.00',8:'0.0%',9:'0.0%',10:'#,##0',11:'0.00',12:'#,##0',13:'0.00'})

# ---------- 7 CVR順位・ファネル（RPM/CPM つき）----------
FN = []
for n, (c, imp, clk) in MF.items():
    s = R['N7'].get(n, 0); k = R['K7'].get(n, 0); q = R['Q7'].get(n, 0); o = ORD.get(n, 0)
    if not s or not o or not clk: continue
    ipo = q/o; bcpa = (s-k)/q*ipo; cpm = c/imp*1000
    ctr = clk/imp; cvr = o/clk; gpo = (s-k)/o          # 粗利/注文
    need = cpm/(gpo*1000)                              # 必要 CTR×CVR
    FN.append([n, '季節' if n in SEA else '通年', imp, round(ctr,4), clk, round(cvr,4), round(s/clk),
               round(c/clk), round(c/o), round(bcpa), round(bcpa-c/o), round(cpm), round(gpo),
               round(ctr*cvr*10000, 2), round(need*10000, 2), round((ctr*cvr)/need, 2)])
mc = statistics.median([x[5] for x in FN]) if FN else 0; mr = statistics.median([x[6] for x in FN]) if FN else 0
FN.sort(key=lambda x: -x[15])
sh('CVR順位', ['CVR = Shopify注文 ÷ Metaリンククリック（Metaの購入数は損益に使わない）',
  f'中央値 CVR {mc*100:.2f}% ／ 売上perクリック {mr:.0f}円',
  '★RPM/CPM = CTR×CVR×粗利/注文 ÷ CPM。1.0未満が赤字圏。★CTR単独では成否を判別できない（実測で反証済み）',
  '★必要CTR×CVR(‱) = CPM ÷ (粗利/注文 × 1000)。この列を下回っている商品が赤字'],
 ['商品','区分','インプ','CTR','クリック','CVR','売上/クリック','CPC','CPA(注文)','分岐CPA(注文)','余裕','CPM','粗利/注文',
  '実測CTR×CVR(‱)','必要CTR×CVR(‱)','RPM/CPM'],
 FN, [26,6,12,9,11,9,13,9,12,15,11,10,11,15,15,10],
 {3:'#,##0',4:'0.00%',5:'#,##0',6:'0.00%',7:'#,##0',8:'#,##0',9:'#,##0',10:'#,##0',11:'#,##0',12:'#,##0',13:'#,##0'})

# ---------- 8 日次推移 ----------
WDJ = ['月','火','水','木','金','土','日']
pr = {}
for d in ALLD:
    s = c = 0
    for n, v in Sd.items():
        if n == '' or d not in v: continue
        g, dc = v[d]; s += g-dc
        if g and n in PRICEa and n in COSTa: c += g/PRICEa[n]*COSTa[n]
    a = sum(ADd[cp].get(d, 0) for cp in ADd); pr[d] = (s, c, a, s-c-a)
DD = []
for j, d in enumerate(ALLD):
    if d < '2026-07-13': continue
    s, c, a, p = pr[d]; ma = sum(pr[x][3] for x in ALLD[j-2:j+1])/3
    DD.append([d, WDJ[datetime.date.fromisoformat(d).weekday()], round(s), round(c), round(a), round(p),
               round(p/s,4), round(s/a,2), round(ma), '⚠️非常ブレーキ' if ma < 130000 else ''])
sh('日次推移', ['日次の利益は上記【近似】原価ベース。4期間サマリーはバリアント別実数量の正確値',
  '非常ブレーキ: 3日移動平均利益 < 130,000円 で発動（カタログ→赤字商品一律−25%）'],
 ['日付','曜','実売','原価','広告費','利益','利益率','MER','3日移動平均利益','警告'],
 DD, [12,5,13,12,12,12,9,8,16,16], {3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.0%',8:'0.00',9:'#,##0'})

# ---------- 9 本日の判定 ----------
JUDGE = [
 ['8/27','快適マジックインソール 増額判定','1日あたり利益 ≥ 23,589円',
  '7日利益 198,171円 = **1日28,310円**','✅合格 → +25%（22,000→27,500）',
  'MER 3日3.05 / 5日3.10 / 7日3.03、いずれも目標2.13超。直近3日の消化率96.2%。増額の1円あたり利益は3端すべて正'],
 ['8/27','ムダ毛シェーバー','8/24の 20,000→17,000 の効果判定',
  '1日あたり利益 8,609 → **11,347円**／削った帯の限界MER 0.15（分岐1.67）','✅減額は正解 → 17,000で維持',
  'Δ広告 −3,204円/日で測定下限をわずかに超えるだけ。標本は小さいので「さらに削る」根拠には使わない'],
 ['8/27','3D足臭リセットブラシ','3窓すべて分岐割れ → 停止/−50%',
  '3日0.85 / 5日0.83 / 7日1.08 vs 分岐1.19。A単独4日合算でも1.085','🚨発動 → −50%（10,000→5,000）',
  '削減1円あたりの利益変化は×0.6/0.7/0.8 の3端すべて正（+0.270〜+0.452）。8/30のCR判定は率で見るので半減でも継続できる'],
 ['8/28','ナノガラス コピー修正判定','点数/注文 1.06 → 1.15',
  '7日実測 296個/266注文 = **1.113**','進行中','8/26朝1.10 → 1.113。動いてはいるが基準には届いていない'],
 ['8/28','ナノバブルシャワーヘッド 中間ゲート','3日累計 Shopify実注文 ≥ 5件。下回れば停止',
  '8/25 0件 → 8/26 **2件**（広告7,627円・MER 1.83）','進行中',
  '帰属は注文UTM（utm_campaign = 6934285209608）。Metaの計上は使わない'],
 ['8/29','カタログ 減額判定','ラストクリック注文シェア < 消化シェアの0.8倍 → 10,000→5,000',
  '8/26実測 2.9% ÷ 7.35% = **0.40倍**','進行中',
  '⚠️ カタログは18,000のまま（旧14,000＋夏以外4,000）。宣言した10,000への減額は未実行'],
 ['8/30','3D足臭 CR A単独判定','RPM/CPM ≥1.3 増額 ／ <1.0 停止 ／ 1.0〜1.3 据え置き',
  '8/26からA単独。初日は MER 0.41 の最悪日','進行中','予算を半減しても率での判定は成立する'],
 ['8/30','4WAY・接触冷感UVアームカバー 停止判定','全店週次利益(8/24-30) ≥ 696,901円',
  '8/24-26 の3日で **408,634円**（7日換算 953,479円）','✅ペース上は大きく上回る','両方とも8/23が最終消化'],
 ['9/1','ナノバブルシャワーヘッド 本判定','倍率 ≥1.3 継続 ／ 1.0〜1.3 は−50% ／ <1.0 停止','進行中','—',
  '★増額は日10,000円が上限（3〜5月にリーチを71,205→145,211人へ広げて壊した実測がある）'],
 ['9/4','優先配送 790円','14日で件数 ≥ 28件','7日実測 **13件**（アタッチ率1.78%）','進行中','8/22に536→790へ値上げ済み'],
]
sh('本日の判定', ['宣言済みの基準に当てはめるだけ。基準は宣言時のまま動かさない',
  '★快適マジックインソールは増額判定に合格（1日28,310円 ≥ 23,589円）',
  '★3D足臭は3窓すべて分岐割れ。削減1円あたりの符号も3端で一致したので発動する'],
 ['判定日','対象','宣言済み基準','実測','結果','備考'], JUDGE, [10,26,34,34,16,46])

# ---------- 10 CVR推移 ----------
SES = {'2026-08-04':6639,'2026-08-05':5425,'2026-08-06':5138,'2026-08-07':4827,'2026-08-08':5527,'2026-08-09':5886,
'2026-08-10':5245,'2026-08-11':6001,'2026-08-12':4900,'2026-08-13':5126,'2026-08-14':5035,'2026-08-15':4871,
'2026-08-16':5349,'2026-08-17':3438,'2026-08-18':3523,'2026-08-19':3190,'2026-08-20':3356,'2026-08-21':3120,
'2026-08-22':3220,'2026-08-23':3438,'2026-08-24':3192,'2026-08-25':3280,'2026-08-26':2733}
ORDD = {'2026-08-04':188,'2026-08-05':152,'2026-08-06':157,'2026-08-07':173,'2026-08-08':184,'2026-08-09':187,
'2026-08-10':141,'2026-08-11':176,'2026-08-12':124,'2026-08-13':125,'2026-08-14':120,'2026-08-15':142,
'2026-08-16':143,'2026-08-17':98,'2026-08-18':89,'2026-08-19':106,'2026-08-20':113,'2026-08-21':110,
'2026-08-22':98,'2026-08-23':106,'2026-08-24':100,'2026-08-25':99,'2026-08-26':105}
CV = [[d, WDJ[datetime.date.fromisoformat(d).weekday()], SES[d], ORDD[d], round(ORDD[d]/SES[d], 4),
       '★昨日（確定値）' if d == '2026-08-26' else ''] for d in sorted(SES)]
o1 = sum(ORDD[d] for d in sorted(SES) if d <= '2026-08-16'); s1_ = sum(SES[d] for d in sorted(SES) if d <= '2026-08-16')
o2 = sum(ORDD[d] for d in sorted(SES) if '2026-08-17' <= d <= '2026-08-26'); s2 = sum(SES[d] for d in sorted(SES) if '2026-08-17' <= d <= '2026-08-26')
sh('CVR推移', ['全店CVR = Shopify注文 ÷ Shopifyセッション。商品別レポートのCVR（注文÷Metaクリック）とは分母が違うので混ぜないこと',
  f'★8/04-16（広告費が大きかった時期）: {o1}注文 / {s1_:,}セッション = {o1/s1_:.2%}',
  f'★8/17-25（広告費を絞った後）: {o2}注文 / {s2:,}セッション = {o2/s2:.2%} → 広告費を削ったらCVRは **{o2/s2/(o1/s1_)-1:+.1%}** 改善している',
  '★8/26は セッション2,733（この窓の最少）に対し注文105件 = **CVR 3.84%** でこの窓の最高。セッションが減ってCVRが上がった',
  '★セッションは 5,000〜6,000/日 → 2,700〜3,400/日 へ減った。これは日予算 590,000→250,000 の直接の結果'],
 ['日付','曜','セッション','注文','全店CVR','備考'], CV, [12,5,12,10,10,20],
 {3:'#,##0',4:'#,##0',5:'0.00%'})

# ---------- 11 曜日指数 ----------
sh('曜日指数', ['直近30日(7/28-8/26)・祝日 8/11(山の日) を除外。売上=gross−値引',
  '★同曜日平均との比較は当面使わない。直近30日の水曜（7/29・8/5・8/12・8/19）が夏物崩壊前で汚染されているため',
  '★代わりに「7日平均比」で読む。8/26は **+4.0%**（同曜日平均比では−33.3%だが、これは分母が汚染されている）'],
 ['曜日','指数(全体=100)'], [[k, round(v, 1)] for k, v in sorted(R['IDX'].items(), key=lambda x: -x[1])], [10, 16])

wb.save('data/reports/report-2026-08-27.xlsx')
print('シート:', wb.sheetnames)
print(f"3日移動平均利益 {p3ma:,.0f}円 / 非常ブレーキ130,000円")
