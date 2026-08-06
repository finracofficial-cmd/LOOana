#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""2026-08-06 朝レポートを実測予算で更新＋予算履歴シート追加＋アクション最終版に差し替え"""
import csv, collections, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
thin=Border(*[Side(style='thin',color='D9D9D9')]*4)
H=PatternFill('solid',fgColor='1F4E79'); ALT=PatternFill('solid',fgColor='F2F7FB')
RED=PatternFill('solid',fgColor='FFC7CE'); GRN=PatternFill('solid',fgColor='C6EFCE'); YEL=PatternFill('solid',fgColor='FFF2CC')
# 実測予算（8/6夕 Meta直読み・キャンペーン合算）
VER={'ナノガラス脱毛パッド':80000,'4WAY':50000,'害虫ブロッカー':45000,'形状記憶日傘':35000,
'接触冷感UVパーカー':30000,'完全遮光・形状記憶':28000,'偏光・調光サングラス':22000,'ムダ毛シェーバー':20000,
'接触冷感UVアームカバー':17000,'5WAY腰掛けファン':16250,'W固定スマホ車載ホルダー':15000,'健康サンダル':13000,
'携帯電動シェーバー':11000,'卓上冷感クーラー':10000,'姿勢サポートチェア':10000,'瞬間冷感ポンチョ':9000,
'完全遮光・接触冷感UVハット':8000,'バランスケアスリッパ':5000,'3WAYサーキュレーター':0,
'UV歯ブラシ除菌器':0,'壁掛けディスペンサー':0}
wb=openpyxl.load_workbook('data/reports/report-2026-08-06.xlsx')
ws=wb['商品別']
fixed=0
for r in range(2,ws.max_row+1):
    n=ws.cell(r,1).value
    if n in VER and ws.cell(r,30).value!=VER[n]:
        ws.cell(r,30,VER[n]); fixed+=1
print(f'商品別シート: 現日予算を実測へ更新 {fixed}件')
# ---- 予算履歴シート ----
hist=collections.defaultdict(dict)   # (campaign) -> date -> total
for r in csv.DictReader(open('data/budget-snapshots.csv',encoding='utf-8')):
    d=r['snapshot_date']
    if d>='2026-08-01':
        hist[r['campaign']][d]=hist[r['campaign']].get(d,0)+int(r['daily_budget'])
DATES=['2026-08-01','2026-08-02','2026-08-03','2026-08-04','2026-08-05','2026-08-06']
# 8/6朝の状態（朝スナップショットを置換してしまったため、当日の変更履歴を明示行で持つ）
AM={'4WAY 取り付けOK 小型瞬間冷却ハンディファン':60000,'害虫ブロッカー':55000,'卓上冷感クーラー':24000,
'3WAYサーキュレーター扇風機':16000,'接触冷感UVパーカー':34000,'完全遮光・形状記憶・晴雨兼用・UV日傘':34000}
if '予算履歴' in wb.sheetnames: del wb['予算履歴']
ws2=wb.create_sheet('予算履歴',2)
ws2['A1']='キャンペーン別 日予算の履歴（8/01〜8/06）。8/06は「朝→夕(Meta直読み実測)」を分けて表示'
ws2['A1'].font=Font(bold=True,size=13); ws2.merge_cells('A1:I1')
hdr=['キャンペーン','8/01','8/02','8/03','8/04','8/05','8/06朝','8/06夕(実測)','8/06中の変更']
for c,v in enumerate(hdr,1):
    cell=ws2.cell(3,c,v); cell.font=Font(bold=True,color='FFFFFF'); cell.fill=H; cell.border=thin
    cell.alignment=Alignment(horizontal='center',wrap_text=True)
r=3
order=sorted(hist,key=lambda k:-hist[k].get('2026-08-06',0))
for camp in order:
    r+=1
    ev=hist[camp].get('2026-08-06',0)
    am=AM.get(camp, ev)
    vals=[camp]+[hist[camp].get(d,'—') for d in DATES[:-1]]+[am,ev]
    chg=f'{am:,} → {ev:,}（ユーザー実施）' if am!=ev else ''
    vals.append(chg)
    for c,v in enumerate(vals,1):
        cell=ws2.cell(r,c,v); cell.border=thin
        if isinstance(v,int): cell.number_format='#,##0'
        if (r-4)%2==1: cell.fill=ALT
    if am!=ev:
        ws2.cell(r,8).fill=RED if ev<am else GRN
tot=[sum(hist[c].get(d,0) for c in hist) for d in DATES[:-1]]
am_t=sum(AM.get(c,hist[c].get('2026-08-06',0)) for c in hist)
ev_t=sum(hist[c].get('2026-08-06',0) for c in hist)
r+=1
row=['合計']+tot+[am_t,ev_t,f'−{am_t-ev_t:,}円/日']
for c,v in enumerate(row,1):
    cell=ws2.cell(r,c,v); cell.font=Font(bold=True); cell.border=thin
    if isinstance(v,int): cell.number_format='#,##0'
r+=2
for nt in ['8/05以前はdata/budget-snapshots.csvの日次記録（Meta実測ベース）','8/06夕はcampaign_and_resource_getで全19キャンペーンを直読みした確定値',
 '8/06中の変更（ユーザー実施6件）: 4WAY 60,000→50,000／害虫 40,000→30,000(+０７１９15,000は不変・B-コピー広告停止)／卓上 24,000→10,000／3WAY 16,000→停止／パーカー 34,000→30,000／完全遮光 34,000→28,000',
 '8/05の変更: カタログ30,000→23,000／4WAY 70,000→60,000／5WAY 20,000→16,250／ポンチョ13,000→9,000／座椅子停止／W固定10,000→15,000']:
    ws2.cell(r,1,'※ '+nt).font=Font(size=9,color='555555'); r+=1
for c,w in zip('ABCDEFGHI',[34,10,10,10,10,10,10,12,30]): ws2.column_dimensions[c].width=w
ws2.freeze_panes='B4'
# ---- カテゴリ別アクション最終版に差し替え ----
if 'カテゴリ別アクション' in wb.sheetnames: del wb['カテゴリ別アクション']
ws3=wb.create_sheet('カテゴリ別アクション（最終）',1)
ACT=[
('🌀 送風機器（今季回復しない・段階縮小）',[
 ('4WAY',50000,50000,'8/9判定','【追い打ち削減は撤回】今日60,000→50,000を実施済み。同日の再削減は3クリーン日ルール違反かつ限界MER1.89>分岐1.58で額を失う。8/9判定=3日(8/6-8/8)の1日あたり利益<20,000円なら37,500へ−25%'),
 ('5WAY腰掛けファン',16250,16250,'8/8判定','1日利益≥16,779で維持／不合格→12,000'),
 ('卓上冷感クーラー',10000,10000,'8/10判定','昨日減額でロック中。判定→7,500'),
 ('3WAYサーキュレーター',0,0,'停止済✅','在庫売り切りへ')]),
('☀️ 日差し・UV（8/11まで維持→縮小設計）',[
 ('形状記憶日傘',35000,35000,'8/11判定','合算判定後に26,000へ縮小開始'),
 ('完全遮光・形状記憶',28000,28000,'8/11判定','判定後、成績の良い方へ寄せる'),
 ('偏光・調光サングラス',22000,16500,'今日','削って利益額+192円/日＝タダで率↑（8/12 LP判定は中止）'),
 ('完全遮光・接触冷感UVハット',8000,8000,'8/7判定','CPA2,540<分岐3,875 ✅見込み→維持')]),
('👕 冷感アパレル（最健全・骨格を守る）',[
 ('接触冷感UVパーカー',30000,30000,'維持','CVR4.78%店内3位。削らない・戻さない'),
 ('接触冷感UVアームカバー',17000,12750,'今日⚠️未実施','余裕+0.09で実質ゼロの予算'),
 ('瞬間冷感ポンチョ',9000,6750,'今日','利益額±0で率+0.07pt')]),
('🦟 虫（触らない）',[
 ('害虫ブロッカー',45000,45000,'8/7判定','削ると額も率も下がる唯一の商品。CPA2,881<3,113✅。8/14基準4,274から段階縮小')]),
('✨ ムダ毛・脱毛（利益の44%・維持）',[
 ('ナノガラス脱毛パッド',80000,80000,'維持','7日利益537,414円で店内1位'),
 ('ムダ毛シェーバー',20000,20000,'維持','増額は削減基調に合わせ見送り'),
 ('携帯電動シェーバー',11000,8250,'今日','削って利益額+230円/日＝タダ')]),
('🦶 足（唯一の赤字）',[
 ('健康サンダル',13000,0,'今日 停止','1円あたり−0.151。停止で+1,933円/日'),
 ('バランスケアスリッパ',5000,5000,'8/13判定','今朝合格→判定で2,500へ')]),
('🚗車・💺姿勢・📦横断（維持）',[
 ('W固定スマホ車載ホルダー',15000,15000,'8/8判定','維持確認のみ（増額なし）'),
 ('姿勢サポートチェア',10000,10000,'8/11本判定','CVR3.47%は基準超え。CPAが未達'),
 ('カタログ全部（テスト）',23000,23000,'8/13判定','増分測定中につき固定。黒字（余裕+780円/注文）')]),
('🆕 新商品（9月の崖への備え）',[
 ('テスト3枠',0,27000,'8/10開始','3WAY+サンダル停止分29,000で賄う。車用シートクッション/防災3in1/スカルプブラシ。CJ原価・納期の確認待ち')]),
]
ws3['A1']='カテゴリ別・商品別アクション 最終版（2026-08-06夕。制約: 床10万死守/基本削減/入れ替えのみ）'
ws3['A1'].font=Font(bold=True,size=13); ws3.merge_cells('A1:F1')
hdr=['商品','現予算(実測)','提案後','実施','根拠・条件']
r=3
for c,v in enumerate(hdr,1):
    cell=ws3.cell(r,c,v); cell.font=Font(bold=True,color='FFFFFF'); cell.fill=H; cell.border=thin
for cat,items in ACT:
    r+=1; ws3.cell(r,1,cat).font=Font(bold=True)
    for c in range(1,6): ws3.cell(r,c).fill=YEL; ws3.cell(r,c).border=thin
    for n,cur,new,when,why in items:
        r+=1
        for c,v in enumerate([n,cur,new,when,why],1):
            cell=ws3.cell(r,c,v); cell.border=thin
            if isinstance(v,int): cell.number_format='#,##0'
            cell.alignment=Alignment(wrap_text=True,vertical='top')
        if new<cur: ws3.cell(r,3).fill=RED
        elif new>cur: ws3.cell(r,3).fill=GRN
        ws3.row_dimensions[r].height=34
for c,w in zip('ABCDE',[28,12,12,13,80]): ws3.column_dimensions[c].width=w
ws3.freeze_panes='A4'
wb.save('data/reports/report-2026-08-06.xlsx')
print('✅ 更新完了:', wb.sheetnames[:5])
