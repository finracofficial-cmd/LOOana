# -*- coding: utf-8 -*-
"""LOOTY日次レポートのExcel生成。使い方: python3 report_xlsx.py <data.json> <out.xlsx>
data.json: {"date":"YYYY-MM-DD(曜)", "headline":[[label,昨日,同曜日,3日,7日,30日,vs7日],...],
 "summary":[[期間,売上,原価,広告費,利益,利益率,手数料,控除後利益,控除後率],...],
 "products":[[商品,7日売上,7日原価,7日広告,7日利益,前日売上,前日利益,前日率,3日利益,30日利益,MER,分岐,余裕,判定,予算],...],
 "notes":[str,...]}"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

data = json.load(open(sys.argv[1]))
wb = Workbook()
TH = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TD = Font(name="Arial", size=10)
NEG = Font(name="Arial", size=10, color="CC0000")
HEAD = PatternFill("solid", fgColor="305496")
ALT = PatternFill("solid", fgColor="F2F6FC")
YELLOW = PatternFill("solid", fgColor="FFF2A8")
GREEN = PatternFill("solid", fgColor="D6EFD8")
RED = PatternFill("solid", fgColor="FADBD8")
thin = Border(*[Side(style="thin", color="CCCCCC")]*4)

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row, c); cell.font = TH; cell.fill = HEAD; cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ---- Sheet0: ネクストアクション ----
na = wb.active; na.title = "ネクストアクション"
na["A1"] = f"LOOTY ネクストアクション {data['date']}"
na["A1"].font = Font(name="Arial", bold=True, size=14)
r = 3
for sec, rows_ in data.get("next_actions", []):
    c = na.cell(r, 1, sec); c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="305496")
    for cc in range(1, 7): na.cell(r, cc).fill = PatternFill("solid", fgColor="305496")
    r += 1
    if rows_ and isinstance(rows_[0], list) and len(rows_[0]) > 1:
        for c2, v in enumerate(rows_[0], 1): na.cell(r, c2, v)
        style_header(na, r, len(rows_[0])); r += 1
        for row in rows_[1:]:
            for c2, v in enumerate(row, 1):
                cell = na.cell(r, c2, v); cell.font = TD; cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
    else:
        for row in rows_:
            na.cell(r, 1, row[0] if isinstance(row, list) else row).font = TD
            na.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
    r += 1
for col, w in zip("ABCDEF", [30, 42, 12, 12, 30, 14]): na.column_dimensions[col].width = w

# ---- Sheet1: 全体サマリー ----
ws = wb.create_sheet("全体サマリー")
ws["A1"] = f"LOOTY 日次損益レポート {data['date']}"
ws["A1"].font = Font(name="Arial", bold=True, size=13)
r = 3
ws.cell(r, 1, "■ 昨日の全体実績 × 期間比較").font = Font(name="Arial", bold=True, size=11)
r += 1
hh = ["指標", "昨日", "同曜日平均", "3日平均/日", "7日平均/日", "30日平均/日", "昨日vs7日平均"]
for c, v in enumerate(hh, 1): ws.cell(r, c, v)
style_header(ws, r, len(hh))
for row in data["headline"]:
    r += 1
    for c, v in enumerate(row, 1):
        cell = ws.cell(r, c, v); cell.font = TD; cell.border = thin
        # 実績期間比較の数値はカンマ表記（ユーザー指定 2026-07-26）
        if isinstance(v, (int, float)) and not isinstance(v, bool) and c in (2,3,4,5,6):
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
r += 2
ws.cell(r, 1, "■ 期間別サマリー（カタログ広告費込み・全商品）").font = Font(name="Arial", bold=True, size=11)
r += 1
sh = ["期間", "売上(gross−値引)", "原価", "広告費", "利益", "利益率", "手数料(3.49%)", "控除後利益", "控除後利益率"]
for c, v in enumerate(sh, 1): ws.cell(r, c, v)
style_header(ws, r, len(sh))
for row in data["summary"]:
    r += 1
    for c, v in enumerate(row, 1):
        cell = ws.cell(r, c, v); cell.font = TD; cell.border = thin
        if isinstance(v, (int, float)) and c in (2,3,4,5,7,8): cell.number_format = "#,##0"
        if isinstance(v, (int, float)) and c in (6,9): cell.number_format = "0.0%"
r += 2
for note in data.get("notes", []):
    ws.cell(r, 1, "・" + note).font = Font(name="Arial", size=9, color="606060")
    r += 1
for col, w in zip("ABCDEFGHI", [22,14,12,12,12,9,12,12,12]): ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ---- Sheet2: 商品別 ----
ws2 = wb.create_sheet("商品別")
ph = ["商品",
      "7日売上", "7日原価", "7日広告費", "7日利益", "7日利益率", "原価率", "広告費率",
      "3日売上", "3日原価", "3日広告費", "3日利益", "3日利益率",
      "前日売上", "前日原価", "前日広告費", "前日利益", "前日利益率",
      "30日利益", "MER(7日)", "分岐", "目標MER", "余裕", "増分MER",
      "CVR(7日)", "CVR順位", "残シーズン(週)",
      "値引(7日)", "返品(7日)",
      "現日予算(円)", "判定", "推奨アクション"]
for c, v in enumerate(ph, 1): ws2.cell(1, c, v)
style_header(ws2, 1, len(ph))
ws2.row_dimensions[1].height = 30
MONEY  = (2,3,4,5, 9,10,11,12, 14,15,16,17, 19, 28,29, 30)
PCT    = (6,7,8, 13, 18, 25)
RATIO  = (20,21,22,23,24)
PROFIT = (5,12,17,19)        # 赤字を赤字表示する利益列
MARGIN = (6,13,18)           # 利益率列（<30%黄・赤字は赤）
for i, row in enumerate(data["products"]):
    r = i + 2
    for c, v in enumerate(row, 1):
        cell = ws2.cell(r, c, v); cell.border = thin
        cell.font = NEG if (isinstance(v,(int,float)) and v < 0 and c in PROFIT+MARGIN) else TD
        if isinstance(v, (int, float)):
            if c in MONEY: cell.number_format = "#,##0"
            if c in PCT:   cell.number_format = "0.0%"
            if c in RATIO: cell.number_format = "0.00"
        if i % 2 == 1: cell.fill = ALT
        if c == 32: cell.alignment = Alignment(wrap_text=True, vertical="top")
        if not isinstance(v, (int, float)): continue
        if c in MARGIN and v < 0.30: cell.fill = RED if v < 0 else YELLOW
        if c == 7 and v > 0.33: cell.fill = RED        # 原価率>33%
        if c == 8 and v > 0.40: cell.fill = RED        # 広告費率>40%
        if c == 24: cell.fill = GREEN if v >= 2.91 else (RED if v < 1.44 else YELLOW)
        if c == 25: cell.fill = GREEN if v >= 0.0318 else RED      # CVR 自店中央値3.18%
        if c == 27 and v <= 5: cell.fill = YELLOW                  # 残シーズン5週以下
        if c == 29 and v > 0: cell.fill = YELLOW                   # 返品あり（原価には影響させない・別窓管理）
# MER vs 目標MER: 目標を超えていれば緑、下回れば赤（商品ごとの目標で判定）
for i, row in enumerate(data["products"]):
    r = i + 2
    mer, tgt = row[19], row[21]
    if isinstance(mer,(int,float)) and isinstance(tgt,(int,float)):
        ws2.cell(r, 20).fill = GREEN if mer >= tgt else RED
WIDTHS = [34, 11,10,11,11,9, 8,9, 10,10,10,10,9, 10,10,10,10,9, 11, 9,7,8,7,8, 9,9,11, 10,10, 11, 22, 62]
for c, w in enumerate(WIDTHS, 1): ws2.column_dimensions[ws2.cell(1,c).column_letter].width = w
ws2.freeze_panes = "B2"  # ヘッダー行＋商品列を固定（ユーザー指定）

# ---- Sheet3: 判定フロー（フロー図＋当日の全商品トレース） ----
if data.get("flow"):
    ws3 = wb.create_sheet("判定フロー")
    ws3["A1"] = "予算増減・CR投入の判定フロー　※商品別シートの判定はこのフローから機械的に生成している"
    ws3["A1"].font = Font(name="Arial", bold=True, size=12)
    r = 3
    for step, q, branch in data["flow"]:
        if not q:
            r += 1; continue
        if not step:                                  # セクション見出し
            c = ws3.cell(r, 1, q); c.font = Font(name="Arial", bold=True, size=11, color="1F4E78")
            ws3.cell(r, 1).fill = PatternFill("solid", fgColor="EAF1FB")
            r += 1; continue
        ws3.cell(r, 1, step).font = Font(name="Arial", bold=True, size=10, color="C00000")
        ws3.cell(r, 2, q).font = Font(name="Arial", bold=True, size=10)
        ws3.cell(r, 3, branch).font = Font(name="Arial", size=9, color="404040")
        for c in (1, 2, 3):
            ws3.cell(r, c).border = thin
            ws3.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 2
    ws3.cell(r, 1, "■ 本日の全商品トレース（どの分岐を通ってその結論になったか）").font = Font(name="Arial", bold=True, size=11)
    r += 1
    th = ["商品", "日販", "7日利益", "MER", "目標MER", "分岐", "余裕", "消化率", "増分MER",
          "判定日", "現予算", "提案予算", "結論", "通った分岐"]
    for c, v in enumerate(th, 1): ws3.cell(r, c, v)
    style_header(ws3, r, len(th))
    hdr = r
    for i, row in enumerate(data["trace"]):
        r += 1
        for c, v in enumerate(row, 1):
            cell = ws3.cell(r, c, v); cell.border = thin; cell.font = TD
            if isinstance(v, (int, float)):
                if c in (3, 11, 12): cell.number_format = "#,##0"
                if c == 2: cell.number_format = "0.0"
                if c in (4, 5, 6, 7, 9): cell.number_format = "0.00"
                if c == 8: cell.number_format = "0%"
            if i % 2 == 1: cell.fill = ALT
            if c == 14: cell.alignment = Alignment(wrap_text=True, vertical="top")
        if row[10] != row[11]:                        # 予算が動く行を強調
            for c in range(1, 15): ws3.cell(r, c).fill = YELLOW
    for c, w in enumerate([30, 7, 11, 7, 8, 7, 7, 8, 9, 10, 10, 10, 26, 78], 1):
        ws3.column_dimensions[ws3.cell(hdr, c).column_letter].width = w
    ws3.column_dimensions["B"].width = 46
    ws3.column_dimensions["C"].width = 92
    ws3.freeze_panes = ws3.cell(hdr + 1, 2)

# ---- 追加シート（data["extra"]: [{"name","title","header","rows","widths","money","pct","notes"}]） ----
for ex in data.get("extra", []):
    wse = wb.create_sheet(ex["name"])
    wse["A1"] = ex["title"]
    wse["A1"].font = Font(name="Arial", bold=True, size=12)
    r = 3
    for c, v in enumerate(ex["header"], 1): wse.cell(r, c, v)
    style_header(wse, r, len(ex["header"]))
    hdr = r
    money = set(ex.get("money", [])); pct = set(ex.get("pct", [])); dec2 = set(ex.get("dec2", []))
    for i, row in enumerate(ex["rows"]):
        r += 1
        if len(row) == 1 and isinstance(row[0], str):          # 区切り見出し
            c = wse.cell(r, 1, row[0]); c.font = Font(name="Arial", bold=True, size=10, color="1F4E78")
            for cc in range(1, len(ex["header"])+1): wse.cell(r, cc).fill = PatternFill("solid", fgColor="EAF1FB")
            continue
        for c, v in enumerate(row, 1):
            cell = wse.cell(r, c, v); cell.border = thin; cell.font = TD
            if isinstance(v, (int, float)):
                if c in money: cell.number_format = "#,##0"
                if c in pct:   cell.number_format = "0.0%"
                if c in dec2:  cell.number_format = "0.00"
                if v < 0: cell.font = NEG
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if i % 2 == 1 and not cell.fill.fgColor.rgb == "00FFF2A8": cell.fill = ALT
    for c, w in enumerate(ex.get("widths", []), 1):
        wse.column_dimensions[wse.cell(hdr, c).column_letter].width = w
    wse.freeze_panes = wse.cell(hdr + 1, 2)
    if ex.get("notes"):
        r += 2
        for nt in ex["notes"]:
            cell = wse.cell(r, 1, nt); cell.font = Font(name="Arial", size=9, color="606060")
            cell.alignment = Alignment(wrap_text=True, vertical="top"); r += 1

wb.save(sys.argv[2])
print("saved", sys.argv[2])
