#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""2026-08-06 夕方版レポート: 日次利益（床防衛）×予算現在地（Meta直読み）×決定アクション"""
import csv, collections, datetime, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
B='data/daily/'
COST={'4WAY':1730,'害虫ブロッカー':867,'形状記憶日傘':1309,'完全遮光・形状記憶':1309,'卓上冷感クーラー':1996,
'接触冷感UVパーカー':1434,'5WAY腰掛けファン':1848,'偏光・調光サングラス':893,'3WAYサーキュレーター':2485,
'接触冷感UVアームカバー':784,'健康サンダル':1162,'瞬間冷感ポンチョ':987,'ムダ毛シェーバー':2798,
'携帯電動シェーバー':1743,'バランスケアスリッパ':1304,'湯上がりガーゼワンピース':1968,'UV歯ブラシ除菌器':3240,
'優先配送':0,'瞬間冷却ハンディファン':2053,'完全遮光・接触冷感UVハット':1411,'ネックマッサージャー':3034,
'リカバリーサンダル':1309,'ナノバブルシャワーヘッド':2255,'姿勢サポートベルト':1429,'癒しの指圧マット':1648,
'スマートノーズEMS美顔器':2453,'W固定スマホ車載ホルダー':1182,'ジェットウォッシャー':1259,'姿勢サポートチェア':1811,
'ナノガラス脱毛パッド':751,'壁掛けディスペンサー':2450}
PRICE={'4WAY':4980,'害虫ブロッカー':3980,'形状記憶日傘':4980,'完全遮光・形状記憶':4980,'卓上冷感クーラー':5980,
'接触冷感UVパーカー':3980,'5WAY腰掛けファン':4980,'偏光・調光サングラス':3980,'3WAYサーキュレーター':6980,
'接触冷感UVアームカバー':3980,'健康サンダル':4980,'瞬間冷感ポンチョ':3980,'携帯電動シェーバー':4980,
'バランスケアスリッパ':4980,'湯上がりガーゼワンピース':5980,'UV歯ブラシ除菌器':8980,'瞬間冷却ハンディファン':5980,
'完全遮光・接触冷感UVハット':4980,'ナノバブルシャワーヘッド':6980,'ネックマッサージャー':9980,'リカバリーサンダル':3980,
'姿勢サポートベルト':5980,'癒しの指圧マット':5980,'優先配送':536,'ナノガラス脱毛パッド':3980,'ムダ毛シェーバー':6980,
'壁掛けディスペンサー':6980,'スマートノーズEMS美顔器':5980,'W固定スマホ車載ホルダー':3980,'ジェットウォッシャー':4980,
'姿勢サポートチェア':5980}
S=collections.defaultdict(lambda: collections.defaultdict(lambda:[0,0]))
for r in csv.DictReader(open(B+'daily_sales.csv',encoding='utf-8')):
    S[r['name']][r['date']][0]+=int(r['gross']); S[r['name']][r['date']][1]+=int(r['disc'])
AD=collections.defaultdict(float)
for r in csv.DictReader(open(B+'daily_ad.csv',encoding='utf-8')):
    AD[r['date']]+=float(r['cost'])
WD=['月','火','水','木','金','土','日']
def daily(d):
    s=c=0
    for n,v in S.items():
        if n=='' or d not in v: continue
        g,dc=v[d]; s+=g-dc; c+=round(g/PRICE.get(n,10**9))*COST.get(n,0)
    return s,c,AD[d],s-c-AD[d]
days=[d for d in sorted(AD) if d>='2026-07-16']
wb=openpyxl.Workbook(); thin=Border(*[Side(style='thin',color='D9D9D9')]*4)
H=PatternFill('solid',fgColor='1F4E79'); ALT=PatternFill('solid',fgColor='F2F7FB')
RED=PatternFill('solid',fgColor='FFC7CE'); YEL=PatternFill('solid',fgColor='FFF2CC'); GRN=PatternFill('solid',fgColor='C6EFCE')
def sheet(name,title,hdr,rows_,widths,money=(),pct=(),notes=()):
    ws=wb.create_sheet(name)
    ws['A1']=title; ws['A1'].font=Font(bold=True,size=13); ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(hdr))
    r=3
    for c,v in enumerate(hdr,1):
        cell=ws.cell(r,c,v); cell.font=Font(bold=True,color='FFFFFF'); cell.fill=H; cell.border=thin
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.row_dimensions[r].height=26
    for i,row in enumerate(rows_):
        r+=1
        for c,v in enumerate(row,1):
            cell=ws.cell(r,c,v); cell.border=thin
            if isinstance(v,(int,float)):
                if c in money: cell.number_format='#,##0'
                if c in pct: cell.number_format='0.0%'
            if i%2==1: cell.fill=ALT
            cell.alignment=Alignment(wrap_text=True,vertical='top') if c==len(hdr) else Alignment(vertical='top')
    r+=2
    for nt in notes:
        ws.cell(r,1,'※ '+nt).font=Font(size=9,color='555555'); r+=1
    for c,w in enumerate(widths,1): ws.column_dimensions[ws.cell(3,c).column_letter].width=w
    ws.freeze_panes='A4'
    return ws
# ---- ①日次利益 ----
rows=[]; hist=[]
for d in days:
    s,c,a,p=daily(d); hist.append(p)
    ma=sum(hist[-3:])/min(3,len(hist))
    w=WD[datetime.date(*map(int,d.split('-'))).weekday()]
    dist=min(p,ma)-100000
    rows.append([d,w,round(s),round(c),round(a),round(p),p/s,round(ma),round(dist),'🚨' if p<130000 else ('⚠️' if p<180000 else '')])
ws=sheet('日次利益（床防衛）','日次利益と床（最低10万円）との距離 — 2026-08-06夕方版。売上=Shopify実測／広告費=Meta実測',
 ['日','曜','売上(g−値引)','原価','広告費','利益','利益率','3日移動平均','床との距離','警告'],
 rows,[11,4,12,10,10,10,8,11,10,6],money=(3,4,5,6,8,9),pct=(7,),
 notes=['同曜日の週次比較: 水曜 643,045→277,105→112,474 と毎週ほぼ半減（季節減衰が主因）',
 '非常ブレーキ: 3日移動平均<130,000円 → カタログ→10,000／4WAYさらに−25%／日傘−25% を即発動',
 '原価はバリアント加重平均単価を含む概算（正値との差±1%未満）'])
for i,row in enumerate(rows):
    if row[5]<130000:
        for c in range(1,11): ws.cell(4+i,c).fill=RED
    elif row[5]<180000:
        for c in range(1,11): ws.cell(4+i,c).fill=YEL
# ---- ②予算現在地 ----
BUD=[['🌀 送風','4WAY','4WAY',50000,'→35,000 今日','昨日+2,530円/消化58,237円＝弱日タダ働き。以後週−25%'],
['🌀','4WAY',' ','','',''],
]
BUD=[['🌀 送風機器','4WAY',50000,35000,'今日','弱い日はタダ働き（昨日+2,530円/消化58,237円）。以後毎週−25%'],
['🌀','5WAY腰掛けファン',16250,16250,'8/8判定','1日利益≥16,779で維持／不合格12,000'],
['🌀','卓上冷感クーラー',10000,10000,'8/10判定','昨日減額でロック中。判定→7,500'],
['🌀','3WAYサーキュレーター',0,0,'停止済✅','在庫売り切りへ'],
['☀️ 日差し・UV','形状記憶日傘',35000,35000,'8/11判定','合算判定後に26,000へ縮小開始'],
['☀️','完全遮光・形状記憶',28000,28000,'8/11判定','判定後、成績の良い方へ寄せる'],
['☀️','偏光・調光サングラス',22000,16500,'今日','削って利益額+192円/日＝タダで率↑。8/12 LP判定は中止'],
['☀️','完全遮光・接触冷感UVハット',8000,8000,'8/7判定','CPA2,540<分岐3,875で合格見込み→維持'],
['👕 冷感アパレル','接触冷感UVパーカー',30000,30000,'維持','CVR4.78%店内3位。削らない・戻さない'],
['👕','接触冷感UVアームカバー',17000,12750,'今日⚠️','未実施の指示。余裕+0.09で実質ゼロ'],
['👕','瞬間冷感ポンチョ',9000,6750,'今日','利益額±0で率+0.07pt'],
['🦟 虫','害虫ブロッカー（本体+０７１９）',45000,45000,'触らない','削ると額も率も下がる唯一の商品。8/7判定CPA<3,113（現2,881✅）',],
['✨ ムダ毛・脱毛','ナノガラス（セット2+3）',80000,80000,'維持','7日利益 店内1位'],
['✨','ムダ毛シェーバー',20000,20000,'維持','増額は削減基調に合わせ見送り'],
['✨','携帯電動シェーバー',11000,8250,'今日','削って利益額+230円/日＝タダ'],
['🦶 足','健康サンダル',13000,0,'今日 停止','唯一の1円あたりマイナス（−0.151）。+1,933円/日'],
['🦶','バランスケアスリッパ',5000,5000,'8/13判定','今朝合格→判定で2,500へ'],
['🚗 車','W固定スマホ車載ホルダー',15000,15000,'維持','8/8判定は維持確認のみ（増額なし）'],
['💺 姿勢','姿勢サポートチェア',10000,10000,'8/11本判定','値下げテスト中。CVR3.47%は基準超え'],
['📦 横断','カタログ全部（テスト）',23000,23000,'8/13判定','増分測定中につき固定。黒字（余裕+780円/注文）'],
['🆕 新商品','テスト3枠（8/10開始）',0,27000,'8/10','3WAY+サンダル停止分29,000で賄う。候補: 車用シートクッション/防災3in1/スカルプブラシ'],
]
tot_c=sum(r[2] for r in BUD); tot_n=sum(r[3] for r in BUD)
BUD.append(['合計','',tot_c,tot_n,'',f'全店 {tot_c:,} → {tot_n:,}円/日（純減 {tot_n-tot_c:+,}）'])
ws2=sheet('予算現在地とアクション','予算の現在地（Meta全19キャンペーン直読み 8/6夕）と決定アクション',
 ['カテゴリ','商品','現予算(実測)','提案後','実施','根拠・条件'],
 BUD,[14,30,12,12,11,64],money=(3,4))
for i,row in enumerate(BUD):
    d=row[3]-row[2] if isinstance(row[3],int) and isinstance(row[2],int) else 0
    if d<0:
        ws2.cell(4+i,4).fill=RED
    elif d>0:
        ws2.cell(4+i,4).fill=GRN
# ---- ③判定カレンダー ----
CAL=[['8/07','害虫ブロッカー','7日CPA(注文)<3,113円','2,881円 ✅','合格→45,000維持'],
['8/07','完全遮光・接触冷感UVハット','7日CPA(注文)<4,569円','2,540円 ✅','合格→8,000維持'],
['8/07','姿勢サポートチェア','中間CTR≥2.0%','1.95% ⚠️僅差','未達なら停止検討'],
['8/08','5WAY腰掛けファン','1日利益≥16,779円','判定日に測定','不合格→12,000'],
['8/08','W固定スマホ車載ホルダー','7日CPA<2,798円','1,253円 ✅','維持の確認のみ'],
['8/08','関連商品セクション','ミックス率≥2%','8/3実測2.7%','続投確定へ'],
['8/10','卓上冷感クーラー','7日MER≥1.83','2.03','不合格→7,500'],
['8/10','新商品テスト3枠','出稿開始','CJ原価・納期待ち','—'],
['8/11','日傘2本合算','合算注文数維持（基準241注文/7日）','判定日に測定','縮小の設計を決定'],
['8/11','姿勢サポートチェア','CPA<2,375かつCVR≥3.15%','2,760円/3.47%','価格7,980へ戻すか判断'],
['8/13','カタログ','全店売上減<69,384円/週','測定中','不合格→30,000へ戻す'],
['8/13','バランスケアスリッパ','7日利益>0','+3,295円','合格でも2,500へ縮小'],
['8/14','害虫ブロッカー','7日CPA<4,274円（分岐×0.8）','2,881円 ✅','超過で−25%・分岐超えで半減/停止'],
['毎朝','全店（床防衛）','3日移動平均≥130,000円','173,817円','割れたら非常ブレーキ3点を即発動'],]
sheet('判定カレンダー','判定カレンダー（宣言済み基準で執行。前倒し・後ろ倒しをしない）',
 ['日付','対象','合格基準','現在値','処置'],CAL,[8,26,34,20,32])
del wb['Sheet']
out='data/reports/report-2026-08-06-evening.xlsx'
wb.save(out); print('saved',out)
