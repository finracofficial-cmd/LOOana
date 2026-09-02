# -*- coding: utf-8 -*-
"""商品別の日次タブを持つExcelを生成する。
使い方: python3 product_tabs.py <daily_sales.csv> <daily_ad.csv> <out.xlsx>
  daily_sales.csv: date,name,gross,disc   （Shopify実測・商品別日次）
  daily_ad.csv:    date,campaign,cost     （Meta実測・キャンペーン別日次）
タブ構成: 一覧（全商品サマリー）＋ 商品ごとに1タブ ＋ カタログ全部 ＋ 非広告商品
各商品タブ: 日付/曜日/曜日指数/売上(gross-値引)/値引/販売数/原価/広告費/利益/利益率/MER/曜日補正MER/日予算/消化率/メモ
"""
import csv, sys, collections, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SALES, AD, OUT = sys.argv[1], sys.argv[2], sys.argv[3]
INTO = sys.argv[4] if len(sys.argv) > 4 else None   # 既存ブックに商品タブを追記する場合はそのパス
SNAP = '/home/user/LOOana/data/budget-snapshots.csv'

COST = {'温感EMSフェイシャルワンド':2079,'ビジュアル耳かき':1655,'3D足臭リセットブラシ':609,'1秒折り畳みチェア':1584,'4WAY':1730,'害虫ブロッカー':867,'形状記憶日傘':1309,'完全遮光・形状記憶':1309,'卓上冷感クーラー':1996,
 '接触冷感UVパーカー':1434,'5WAY腰掛けファン':1848,'偏光・調光サングラス':893,'3WAYサーキュレーター':2485,
 '接触冷感UVアームカバー':784,'健康サンダル':1162,'瞬間冷感ポンチョ':987,'ムダ毛シェーバー':2798,
 '携帯電動シェーバー':1743,'バランスケアスリッパ':1304,'湯上がりガーゼワンピース':1968,'UV歯ブラシ除菌器':3240,
 '優先配送':0,'瞬間冷却ハンディファン':2053,'完全遮光・接触冷感UVハット':1411,'W固定スマホ車載ホルダー':1069,'ネックマッサージャー':3034,
 'リカバリーサンダル':1309,'ナノバブルシャワーヘッド':2255,'姿勢サポートベルト':1429,'癒しの指圧マット':1648,'姿勢サポートチェア':1811,
 # 加重平均（バリエーション構成が日次で取れないため30日実測ミックスで加重）
 'ナノガラス脱毛パッド':752, '壁掛けディスペンサー':2405, '2WAYシートボックス':1943,'オートソープディスペンサー':3191,'スタイルアップインナー':1202, '快適マジックインソール':677,'むくみ取りかっさ':1120}
PRICE = {'温感EMSフェイシャルワンド':5980,'ビジュアル耳かき':4980,'3D足臭リセットブラシ':3980,'1秒折り畳みチェア':4980,'4WAY':4980,'害虫ブロッカー':3980,'形状記憶日傘':4980,'完全遮光・形状記憶':4980,'卓上冷感クーラー':5980,
 '接触冷感UVパーカー':4980,'5WAY腰掛けファン':4980,'偏光・調光サングラス':3980,'3WAYサーキュレーター':5980,
 '接触冷感UVアームカバー':3980,'健康サンダル':4980,'瞬間冷感ポンチョ':3980,'携帯電動シェーバー':4980,
 'バランスケアスリッパ':4980,'湯上がりガーゼワンピース':5980,'UV歯ブラシ除菌器':8980,'瞬間冷却ハンディファン':5980,
 '完全遮光・接触冷感UVハット':5980,'ナノバブルシャワーヘッド':6980,'むくみ取りかっさ':3980,'ネックマッサージャー':9980,
 'リカバリーサンダル':3980,'姿勢サポートベルト':5980,'癒しの指圧マット':5980,'優先配送':536,
 'ナノガラス脱毛パッド':3980, '壁掛けディスペンサー':6756, '2WAYシートボックス':4980,'オートソープディスペンサー':7980,'スタイルアップインナー':3980, '快適マジックインソール':3980, 'ムダ毛シェーバー':None,
 'W固定スマホ車載ホルダー':3980, '完全遮光・接触冷感UVハット':None, '3WAYサーキュレーター':None, '姿勢サポートチェア':None}   # None=売価が窓内で変動
BLENDED = {'3WAYサーキュレーター':'2026-08-02に売価5,980→6,980円へ値上げ（原価2,485円据置）。原価率41.6%→35.6%・分岐MER1.71→1.55・目標MER4.27→3.40・分岐CPA3,495→4,495円。販売数は日付に応じた売価で算出',
           '完全遮光・接触冷感UVハット':'2026-07-31に新規出稿開始。同日に売価5,980→4,980円へ値下げしたため、販売数は日付に応じた売価で算出',
           'W固定スマホ車載ホルダー':'2026-08-01に新規出稿開始（通年商品）。売価3,980円・原価1,182円・原価率29.7%・分岐CPA2,798円・分岐MER1.42・目標MER2.83',
           '姿勢サポートチェア':'2026-08-04に売価7,980→5,980円で広告再開（原価1,811円据置）。原価率30.3%・分岐CPA4,169円・分岐MER1.43・目標MER2.88。⚠️8/4時点で広告セットが2本ACTIVE（姿勢サポートチェア10,000＋座椅子20260516 9,000）',
           'ナノガラス脱毛パッド':'原価は白/緑757円・黒/桃740円の30日実測ミックス(835:364)で加重平均（752円）。4期間サマリーはバリエーション別実数量の正確値',
           '快適マジックインソール':'2026-08-11 17:03 JSTに新規出稿（通年・足カテゴリ・日予算8,000円）。売価3,980円・原価652〜708円（サイズ依存）・単純平均677円・原価率17.0%（店内最低）・分岐MER1.205・目標MER2.084・分岐CPA3,303円。⚠️原価はサイズで8.6%ばらつくため、実売が出たらバリアント別の実数量で加重し直す。中間判定8/15・本判定8/19',
           '2WAYシートボックス':'2026-08-06に新規出稿開始（通年・車カテゴリ）。売価4,980円・原価1,943円・原価率39.0%・分岐MER1.64・目標MER3.84・分岐CPA3,037円。本判定8/13',
           'ムダ毛シェーバー':'売価は7/27に5,980→6,980円へ値上げ。販売数は日付に応じた売価で算出（値上げ当日7/27のみ新旧混在を分解: 旧1個+新5個=6個）',
           '壁掛けディスペンサー':'日次の売価・原価は3本セット(6,980/2,528円)と2本セット(5,980/1,981円)の30日実測ミックスで加重平均。'
                            '販売数・原価は日次では概算になる（4期間サマリーはバリエーション別実数量の正確値）'}
# Shopify商品名 → Metaキャンペーン名
CAMP = {'4WAY':'4WAY 取り付けOK 小型瞬間冷却ハンディファン','完全遮光・形状記憶':'完全遮光・形状記憶・晴雨兼用・UV日傘',
        '3WAYサーキュレーター':'3WAYサーキュレーター扇風機','壁掛けディスペンサー':'ディスペンサー'}
TAB = {'温感EMSフェイシャルワンド':'温感EMSワンド','ビジュアル耳かき':'ビジュアル耳かき','3D足臭リセットブラシ':'3D足臭ブラシ','1秒折り畳みチェア':'1秒折り畳みチェア','4WAY':'4WAY','害虫ブロッカー':'害虫ブロッカー','形状記憶日傘':'形状記憶日傘',
 'ナノガラス脱毛パッド':'ナノガラス脱毛パッド','完全遮光・形状記憶':'完全遮光UV日傘','接触冷感UVパーカー':'UVパーカー',
 '卓上冷感クーラー':'卓上冷感クーラー','5WAY腰掛けファン':'5WAY腰掛けファン','偏光・調光サングラス':'サングラス',
 '接触冷感UVアームカバー':'UVアームカバー','3WAYサーキュレーター':'3WAYサーキュレーター','瞬間冷感ポンチョ':'瞬間冷感ポンチョ',
 'ムダ毛シェーバー':'ムダ毛シェーバー','健康サンダル':'健康サンダル','UV歯ブラシ除菌器':'UV歯ブラシ除菌器',
 '携帯電動シェーバー':'携帯電動シェーバー','バランスケアスリッパ':'バランスケアスリッパ','壁掛けディスペンサー':'壁掛けディスペンサー',
 '湯上がりガーゼワンピース':'湯上がりワンピース',
 '完全遮光・接触冷感UVハット':'UVハット','W固定スマホ車載ホルダー':'車載ホルダー','姿勢サポートチェア':'姿勢サポートチェア','2WAYシートボックス':'2WAYシートボックス','快適マジックインソール':'快適インソール',
 'ナノバブルシャワーヘッド':'ナノバブル','むくみ取りかっさ':'むくみ取りかっさ'}
EVENTS = {('3D足臭リセットブラシ','2026-08-20'):'出稿開始 日予算8,000円→10,000円',
 ('温感EMSフェイシャルワンド','2026-09-02'):'出稿開始 日予算10,000円（夕方開始・原価はピンク2,079の保守側）',
 ('ビジュアル耳かき','2026-08-28'):'出稿開始 日予算8,000円',
 ('3D足臭リセットブラシ','2026-08-29'):'停止（最終消化日・11日で85,838円/11注文＝CPA 7,803円）',
 ('快適マジックインソール','2026-08-30'):'増額 27,000→30,000',('むくみ取りかっさ','2026-08-30'):'増額 10,000→13,000',
 ('バランスケアスリッパ','2026-08-30'):'増額 10,000→12,000',('2WAYシートボックス','2026-08-30'):'減額 13,000→11,000',
 ('偏光・調光サングラス','2026-08-30'):'減額 7,000→5,000',('完全遮光・接触冷感UVハット','2026-08-30'):'減額 8,000→6,000',
 ('1秒折り畳みチェア','2026-08-19'):'出稿開始',('1秒折り畳みチェア','2026-08-20'):'停止（0/49クリック）',
 ('むくみ取りかっさ','2026-08-23'):'出稿開始 日予算10,000円',
 ('4WAY','2026-08-23'):'停止（最終消化日）',('接触冷感UVアームカバー','2026-08-23'):'停止（最終消化日）',
 ('ムダ毛シェーバー','2026-08-24'):'減額 20,000→17,000',
 ('形状記憶日傘','2026-08-25'):'停止（最終消化日・7日MER 1.17で分岐1.36割れ）',
 ('ナノバブルシャワーヘッド','2026-08-25'):'新セット「- コピー」で再開 日予算8,000円。中間ゲート8/28・本判定9/1',
 ('快適マジックインソール','2026-08-27'):'増額 22,000→27,000。P1=8/24-26・P2=8/28-30で8/31判定（基準 1日28,310円）',
 ('バランスケアスリッパ','2026-08-26'):'増額 6,000→10,000（+67%・ユーザー判断で確定）',
 ('害虫ブロッカー','2026-07-27'):'CRを1本→3本に追加',
          ('4WAY','2026-07-29'):'予備CR(広告B)を1本追加',
          ('ナノガラス脱毛パッド','2026-07-29'):'増額 68,000→76,000（メインセット55,000→63,000）',
          ('接触冷感UVアームカバー','2026-07-28'):'CRを1本追加',
          ('携帯電動シェーバー','2026-07-28'):'CRを1本追加',
          ('UV歯ブラシ除菌器','2026-07-23'):'複製リセット(A-コピー)投入',
          ('ムダ毛シェーバー','2026-07-27'):'値上げ 5,980→6,980円',
          ('卓上冷感クーラー','2026-07-27'):'LP改修（不安解消ブロック追加）',
          ('壁掛けディスペンサー','2026-07-28'):'LP全面改修＋価格メリット刷新',
          ('姿勢サポートチェア','2026-08-04'):'5,980円で広告再開(10,000円/日)。旧セット座椅子20260516も並走',
          ('UV歯ブラシ除菌器','2026-08-04'):'広告停止（ユーザー判断）',
          ('形状記憶日傘','2026-08-04'):'同時並行テストセット(25,000円/日)を停止',
          ('3WAYサーキュレーター','2026-08-02'):'値上げ 5,980→6,980円',
          ('湯上がりガーゼワンピース','2026-07-26'):'キャンペーン停止',
          ('完全遮光・接触冷感UVハット','2026-07-31'):'新規出稿開始（日予算8,000）＋売価 5,980→4,980へ値下げ',
          ('W固定スマホ車載ホルダー','2026-08-01'):'新規出稿開始（日予算8,000・分岐CPA2,798円・通年商品）',
          ('害虫ブロッカー','2026-08-01'):'本体セット減額 60,000→43,000',
          ('接触冷感UVアームカバー','2026-08-02'):'増額を巻き戻し 22,000→17,000（増額後2日で分岐割れ）',
          ('UV歯ブラシ除菌器','2026-08-02'):'広告Aを複製して3本体制に（複製リセット）',
          ('ナノガラス脱毛パッド','2026-08-02'):'★セット３のみ増額 13,000→17,000（セット２は消化率86.6%のため63,000据置）。判定8/6',
          ('ムダ毛シェーバー','2026-08-02'):'★増額 16,000→20,000。値上げ判定は6日46個で合格確定。判定8/6',
          ('3WAYサーキュレーター','2026-08-02'):'★値上げ 5,980→6,980円（原価率41.6%→35.6%）。判定8/9=7日販売数≥37個',
          ('害虫ブロッカー','2026-08-03'):'★配分変更 本体43,000→40,000／０７１９12,000→15,000（CR判定不合格・CTR2.67%）。判定8/7',
          ('W固定スマホ車載ホルダー','2026-08-03'):'★増額 8,000→10,000（CPA1,191円=分岐の43%・消化率119%）。判定8/8',
          ('卓上冷感クーラー','2026-08-03'):'LP判定不合格（CVR1.77%<2.5%）→LP施策は打ち切り。予算24,000は据置',
          ('壁掛けディスペンサー','2026-08-02'):'★広告停止（需要プール枯渇。返品率12.7%の解消と新素材を条件に10月以降で再開検討）',
          ('健康サンダル','2026-08-06'):'★広告停止（7日−13,531円・CVR−30%）',
          ('3WAYサーキュレーター','2026-08-06'):'★広告停止',
          ('卓上冷感クーラー','2026-08-06'):'★減額 24,000→10,000（夕方さらに10,000→7,500）',
          ('4WAY','2026-08-06'):'★減額 60,000→50,000（夕方さらに50,000→37,500）',
          ('害虫ブロッカー','2026-08-06'):'★本体減額 40,000→30,000＋広告「B - コピー」停止。夜に秋CR「A - コピー」を本体へ投入',
          ('接触冷感UVパーカー','2026-08-06'):'★減額 34,000→30,000',
          ('完全遮光・形状記憶','2026-08-06'):'★減額 34,000→28,000',
          ('接触冷感UVアームカバー','2026-08-06'):'★減額 17,000→12,500（8/09にさらに8,000）',
          ('2WAYシートボックス','2026-08-06'):'★新規出稿開始（日予算10,000・分岐CPA3,037円）。中間8/09・本判定8/13',
          ('携帯電動シェーバー','2026-08-06'):'★減額 11,000→8,250',
          ('瞬間冷感ポンチョ','2026-08-05'):'★減額 13,000→9,000。判定8/12',
          ('5WAY腰掛けファン','2026-08-05'):'★増額の巻き戻し 20,000→16,250',
          ('害虫ブロッカー','2026-08-07'):'LP全季節化（虫の季節ブロック追加）。判定は売上/クリック≥200円の維持',
          ('偏光・調光サングラス','2026-08-06'):'LP秋冬化（西日・運転訴求）。判定8/13=7日CVR≥3.0%',
          ('形状記憶日傘','2026-08-06'):'LP全季節化（晴雨兼用・秋雨/台風）。判定8/18',
          ('完全遮光・形状記憶','2026-08-06'):'LP全季節化。判定8/18',
          ('バランスケアスリッパ','2026-08-07'):'新素材D投入（第1関門8/10・第2関門8/14=外部CTR≥2.00%かつ売上/クリック≥163円）',
          ('2WAYシートボックス','2026-08-08'):'★増額 10,000→12,500',
          ('W固定スマホ車載ホルダー','2026-08-06'):'★増額 15,000→20,000（1日で+100%）',
          ('W固定スマホ車載ホルダー','2026-08-08'):'★増額 20,000→23,000',
          ('W固定スマホ車載ホルダー','2026-08-09'):'★増額 23,000→26,000',
          ('W固定スマホ車載ホルダー','2026-08-10'):'★増額 26,000→32,500。週次ランプで判定8/16（増分MER≥2.88）',
          ('2WAYシートボックス','2026-08-09'):'★増額 12,500→15,000',
          ('4WAY','2026-08-10'):'★減額 37,500→28,000',
          ('完全遮光・形状記憶','2026-08-10'):'★減額 28,000→21,000',
          ('完全遮光・接触冷感UVハット','2026-08-10'):'★増額 8,000→12,000',
          ('姿勢サポートチェア','2026-08-05'):'旧セット「座椅子20260516」(9,000)を停止 → 10,000のみ',
          ('快適マジックインソール','2026-08-11'):'★新規出稿（日予算8,000・分岐CPA3,303円・通年）。中間判定8/15・本判定8/19',
          ('瞬間冷感ポンチョ','2026-08-11'):'★広告停止（3窓すべて分岐割れ・7日−27,856円）',
          ('携帯電動シェーバー','2026-08-11'):'★減額 8,250→4,000（3窓すべて分岐1.54割れ）',
          ('完全遮光・形状記憶','2026-08-11'):'★減額 21,000→15,750（3窓すべて余裕<0.30）',
          ('偏光・調光サングラス','2026-08-11'):'★減額 22,000→16,500（3窓すべて余裕<0.30）',
          ('4WAY','2026-08-11'):'★減額 28,000→21,000（季節の前倒し縮小トリガーに該当）'}
HOLIDAYS = {'2026-07-20':'海の日'}
WD = ['月','火','水','木','金','土','日']

# ---- 読み込み ----
S = collections.defaultdict(lambda: collections.defaultdict(lambda: [0,0]))   # [name][date] = [gross, disc]
for r in csv.DictReader(open(SALES, encoding='utf-8')):
    S[r['name']][r['date']][0] += int(r['gross']); S[r['name']][r['date']][1] += int(r['disc'])
A = collections.defaultdict(lambda: collections.defaultdict(float))            # [campaign][date] = cost
for r in csv.DictReader(open(AD, encoding='utf-8')):
    A[r['campaign']][r['date']] += float(r['cost'])
B = collections.defaultdict(lambda: collections.defaultdict(int))              # [campaign][date] = daily budget
for r in csv.DictReader(open(SNAP, encoding='utf-8')):
    B[r['campaign']][r['snapshot_date']] += int(r['daily_budget'])

DATES = sorted({d for v in S.values() for d in v})
def wd(d): return WD[datetime.date(*map(int, d.split('-'))).weekday()]
# 曜日指数（祝日を除外して算出）
store = {d: sum(v[d][0]-v[d][1] for v in S.values() if d in v) for d in DATES}
base = [d for d in DATES if d not in HOLIDAYS]
byw = collections.defaultdict(list)
for d in base: byw[wd(d)].append(store[d])
avg = sum(store[d] for d in base)/len(base)
IDX = {k: sum(v)/len(v)/avg*100 for k, v in byw.items()}

# ---- スタイル ----
wb = load_workbook(INTO) if INTO else Workbook()
TH = Font(name='Arial', bold=True, color='FFFFFF', size=10)
TD = Font(name='Arial', size=10)
NEG = Font(name='Arial', size=10, color='CC0000')
HEAD = PatternFill('solid', fgColor='305496')
ALT = PatternFill('solid', fgColor='F2F6FC')
GREEN = PatternFill('solid', fgColor='D6EFD8')
RED = PatternFill('solid', fgColor='FADBD8')
YEL = PatternFill('solid', fgColor='FFF2A8')
GRAY = PatternFill('solid', fgColor='EDEDED')
thin = Border(*[Side(style='thin', color='CCCCCC')]*4)

def header(ws, row, n):
    for c in range(1, n+1):
        x = ws.cell(row, c); x.font = TH; x.fill = HEAD; x.border = thin
        x.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def _mixed_qty(g, p_old, p_new):
    """新旧売価が混在する日の販売数を解く（a×旧 + b×新 = gross）"""
    for b in range(0, g // p_new + 1):
        rem = g - b * p_new
        if rem % p_old == 0:
            return rem // p_old + b
    return None

def series(name):
    """その商品の日次系列を返す"""
    camp = CAMP.get(name, name)
    out = []
    for d in DATES:
        g, dc = S[name].get(d, [0, 0])
        sales = g - dc
        cost_ad = A[camp].get(d, 0.0)
        p = PRICE.get(name)
        if name == 'ムダ毛シェーバー' and g:      # 7/27に5,980→6,980へ値上げ（日付対応の売価で算出）
            if d < '2026-07-27': q = round(g / 5980)
            elif g % 6980 == 0:  q = g // 6980
            else:                q = _mixed_qty(g, 5980, 6980)   # 値上げ当日の新旧混在
        elif name == '完全遮光・接触冷感UVハット' and g:   # 7/31に5,980→4,980へ値下げ
            if d < '2026-07-31': q = round(g / 5980)
            elif g % 4980 == 0:  q = g // 4980
            else:                q = _mixed_qty(g, 5980, 4980)
        elif name == '3WAYサーキュレーター' and g:        # 8/2に5,980→6,980へ値上げ
            if d < '2026-08-02': q = round(g / 5980)
            elif g % 6980 == 0:  q = g // 6980
            else:                q = _mixed_qty(g, 5980, 6980)   # 値上げ当日の新旧混在
        elif name == '接触冷感UVパーカー' and g:          # 8/13 09:58 JST に3,980→4,980へ値上げ
            if d < '2026-08-13': q = round(g / 3980)
            elif g % 4980 == 0:  q = g // 4980
            elif g % 3980 == 0:  q = g // 3980
            else:                q = _mixed_qty(g, 3980, 4980)   # 値上げ当日の新旧混在
        elif name == '姿勢サポートチェア' and g:          # 8/4に7,980→5,980へ値下げして広告再開
            if d < '2026-08-04': q = round(g / 7980)
            elif g % 5980 == 0:  q = g // 5980
            else:                q = _mixed_qty(g, 7980, 5980)
        elif p:
            q = round(g / p)
        else:                       # 窓内で売価が変動し構成が解けない商品は数量を出さない
            q = None
        cost = q * COST.get(name, 0) if q is not None else None
        profit = sales - (cost or 0) - cost_ad
        bud = B.get(camp, {}).get(d)
        out.append(dict(date=d, wd=wd(d), idx=IDX[wd(d)], sales=sales, disc=dc, qty=q,
                        cost=cost, ad=cost_ad, profit=profit, bud=bud))
    return out

def block(rows, days):
    s = sum(r['sales'] for r in rows[-days:]); c = sum((r['cost'] or 0) for r in rows[-days:])
    a = sum(r['ad'] for r in rows[-days:])
    return s, c, round(a), round(s-c-a)

ADV = [n for n in TAB if any(A[CAMP.get(n, n)].values())]
ADV.sort(key=lambda n: -sum(v[0]-v[1] for v in S[n].values()))

# ---- Sheet: 一覧 ----
if INTO:
    ws = wb.create_sheet('一覧')
else:
    ws = wb.active; ws.title = '一覧'
ws['A1'] = f'LOOTY 商品別 日次パフォーマンス（{DATES[0]}〜{DATES[-1]}・全てShopify/Metaの実測）'
ws['A1'].font = Font(name='Arial', bold=True, size=13)
ws['A2'] = '各商品タブに日次の 売上／値引／販売数／原価／広告費／利益／利益率／MER／曜日補正MER／日予算／消化率 を載せています。'
ws['A3'] = '曜日指数（祝日7/20を除外した実測）: ' + '・'.join(f'{k}{v:.1f}' for k, v in sorted(IDX.items(), key=lambda x: -x[1]))
r = 5
hh = ['商品', '売価', '単価原価', '原価率(30日)', '分岐MER', '目標MER',
      '30日売上', '30日原価', '30日広告費', '30日利益', '30日利益率', '30日MER',
      '7日売上', '7日広告費', '7日利益', '7日MER', '前日利益', '現日予算']
for c, v in enumerate(hh, 1): ws.cell(r, c, v)
header(ws, r, len(hh)); r += 1
for n in ADV:
    rows = series(n)
    s30, c30, a30, p30 = block(rows, 30); s7, c7, a7, p7 = block(rows, 7)
    p1 = rows[-1]['profit']
    cr = c30/s30 if s30 else 0
    br = round(1/(1-cr), 2) if cr < 1 else ''
    tg = round(1/(1-cr-0.35), 2) if 1-cr-0.35 > 0 else ''
    bud = B.get(CAMP.get(n, n), {}).get(DATES[-1])
    vals = [n, PRICE.get(n) or '変動', COST.get(n, 0), round(cr, 4), br, tg,
            s30, c30, a30, p30, round(p30/s30, 4) if s30 else '', round(s30/a30, 2) if a30 else '',
            s7, a7, p7, round(s7/a7, 2) if a7 else '', round(p1), bud or '']
    for c, v in enumerate(vals, 1):
        x = ws.cell(r, c, v); x.border = thin; x.font = TD
        if isinstance(v, (int, float)) and v < 0 and c in (10, 11, 15, 17): x.font = NEG
        if c in (7, 8, 9, 10, 13, 14, 15, 17, 2, 3, 18): x.number_format = '#,##0'
        if c in (4, 11): x.number_format = '0.0%'
        if c in (5, 6, 12, 16): x.number_format = '0.00'
    if isinstance(br, float) and a30:
        ws.cell(r, 12).fill = GREEN if s30/a30 >= (tg or 99) else (RED if s30/a30 < br else YEL)
    r += 1
for col, w in zip('ABCDEFGHIJKLMNOPQR', [26,9,9,11,9,9, 12,11,12,12,11,9, 11,11,11,9, 11,11]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'B6'

# ---- 商品ごとのタブ ----
PH = ['日付','曜日','曜日指数','売上(gross−値引)','値引','販売数','売価(実測)','単価原価','原価','原価率',
      '広告費','利益','利益率','MER','曜日補正MER','日予算','消化率','メモ']
for n in ADV:
    rows = series(n)
    w = wb.create_sheet(TAB[n][:31])
    w['A1'] = n; w['A1'].font = Font(name='Arial', bold=True, size=13)
    s30, c30, a30, p30 = block(rows, 30)
    cr = c30/s30 if s30 else 0
    br = 1/(1-cr) if cr < 1 else 0
    tg = 1/(1-cr-0.35) if 1-cr-0.35 > 0 else 0
    if s30 > 0:
        w['A2'] = (f'売価 {PRICE.get(n) or "窓内で変動"}円 ／ 単価原価 {COST.get(n,0):,}円 ／ '
                   f'原価率(30日) {cr*100:.1f}% ／ 分岐MER {br:.2f} ／ 目標MER {tg:.2f}')
    else:   # 30日窓に売上がない新商品は実測から比率を出せない（0%と誤表示しない）
        _cr = COST.get(n,0)/PRICE[n] if PRICE.get(n) else None
        w['A2'] = (f'売価 {PRICE.get(n) or "窓内で変動"}円 ／ 単価原価 {COST.get(n,0):,}円 ／ '
                   + (f'原価率(定価ベース) {_cr*100:.1f}% ／ 分岐MER {1/(1-_cr):.2f} ／ 目標MER {1/(1-_cr-0.35):.2f}'
                      if _cr else '原価率・分岐・目標 = 30日窓に売上なしのため算出不可')
                   + ' ※30日窓に売上がないため実測比率ではなく定価から算出')
    if n in BLENDED:
        w['A3'] = '※ ' + BLENDED[n]; w['A3'].font = Font(name='Arial', size=9, color='CC0000')
    rr = 5
    w.cell(rr, 1, '■ 期間サマリー').font = Font(name='Arial', bold=True, size=11); rr += 1
    sh = ['期間', '売上', '原価', '広告費', '利益', '利益率', 'MER', '余裕(MER−分岐)']
    for c, v in enumerate(sh, 1): w.cell(rr, c, v)
    header(w, rr, len(sh)); rr += 1
    for lbl, dd in [(f'前日({DATES[-1][5:]})', 1), (f'3日({DATES[-3][5:]}-{DATES[-1][5:]})', 3), (f'7日({DATES[-7][5:]}-{DATES[-1][5:]})', 7), (f'30日({DATES[-30][5:]}-{DATES[-1][5:]})', 30)]:
        s, c_, a, p = block(rows, dd)
        vals = [lbl, s, c_, a, p, round(p/s, 4) if s else '', round(s/a, 2) if a else '',
                round(s/a - br, 2) if a and br else '']
        for c, v in enumerate(vals, 1):
            x = w.cell(rr, c, v); x.border = thin; x.font = NEG if (isinstance(v,(int,float)) and v < 0 and c in (5,6,8)) else TD
            if c in (2,3,4,5): x.number_format = '#,##0'
            if c == 6: x.number_format = '0.0%'
            if c in (7,8): x.number_format = '0.00'
        rr += 1
    rr += 1
    w.cell(rr, 1, '■ 日次').font = Font(name='Arial', bold=True, size=11); rr += 1
    for c, v in enumerate(PH, 1): w.cell(rr, c, v)
    header(w, rr, len(PH)); w.row_dimensions[rr].height = 28; rr += 1
    top = rr
    for i, x in enumerate(rows):
        memo = []
        if x['date'] in HOLIDAYS: memo.append(f'祝日({HOLIDAYS[x["date"]]})')
        prev = rows[i-1]['bud'] if i else None
        if x['bud'] is not None and prev is not None and x['bud'] != prev:
            memo.append(f'日予算 {prev:,}→{x["bud"]:,}円')
        if x['bud'] and (x['ad']/x['bud'] > 1.8 or x['ad']/x['bud'] < 0.4):
            memo.append('⚠️消化率が異常（当日中の予算変更 or 開始初日 or スナップショット欠落を確認）')
        if (n, x['date']) in EVENTS: memo.append(EVENTS[(n, x['date'])])
        mer = x['sales']/x['ad'] if x['ad'] else None
        gross = x['sales'] + x['disc']
        unit_price = round(gross / x['qty']) if x['qty'] else ''          # その日の実測売価（値引前）
        unit_cost  = round(x['cost'] / x['qty']) if x['qty'] else ''      # その日の実測単価原価（バリアント加重後）
        cost_rate  = round(x['cost'] / x['sales'], 4) if x['sales'] and x['cost'] else ''
        vals = [x['date'], x['wd'], round(x['idx'], 1), x['sales'], x['disc'], x['qty'],
                unit_price, unit_cost, x['cost'], cost_rate,
                round(x['ad']), round(x['profit']),
                round(x['profit']/x['sales'], 4) if x['sales'] else '',
                round(mer, 2) if mer else '', round(mer/(x['idx']/100), 2) if mer else '',
                x['bud'] or '', round(x['ad']/x['bud'], 3) if x['bud'] else '', ' ／ '.join(memo)]
        for c, v in enumerate(vals, 1):
            cell = w.cell(rr, c, v); cell.border = thin
            cell.font = NEG if (isinstance(v, (int, float)) and v < 0 and c in (12, 13)) else TD
            if i % 2 == 1: cell.fill = ALT
            if c in (4,5,7,8,9,11,12,16): cell.number_format = '#,##0'
            if c in (10,13,17): cell.number_format = '0.0%'
            if c in (3,14,15): cell.number_format = '0.00'
            if c == 18: cell.alignment = Alignment(wrap_text=True, vertical='top')
        if mer:
            w.cell(rr, 14).fill = GREEN if mer >= tg else (RED if mer < br else YEL)
        if memo: w.cell(rr, 18).fill = YEL
        if x['date'] in HOLIDAYS:
            for c in range(1, 4): w.cell(rr, c).fill = GRAY
        rr += 1
    for col, wid in zip('ABCDEFGHIJKLMNOPQR', [12,6,9,16,10,8,11,10,11,9,11,11,9,8,11,10,9,40]):
        w.column_dimensions[col].width = wid
    w.freeze_panes = f'A{top}'
    w.auto_filter.ref = f'A{top-1}:R{rr-1}'

# ---- カタログ全部（テスト） ----
w = wb.create_sheet('カタログ全部')
w['A1'] = 'カタログ全部（テスト）※全商品横断のカタログ広告'
w['A1'].font = Font(name='Arial', bold=True, size=13)
w['A2'] = 'Shopify側でどの商品にも紐づかないため売上・利益は分解できない。広告費のみを日次で管理する。'
w['A3'] = 'この行を含めると、商品タブの広告費の合計＝Metaアカウントの実消化額と一致する。'
rr = 5
for c, v in enumerate(['日付', '曜日', '広告費', '日予算', '消化率'], 1): w.cell(rr, c, v)
header(w, rr, 5); rr += 1
for d in DATES:
    a = A['カタログ全部（テスト）'].get(d, 0); b = B.get('カタログ全部（テスト）', {}).get(d)
    for c, v in enumerate([d, wd(d), round(a), b or '', round(a/b, 3) if b else ''], 1):
        x = w.cell(rr, c, v); x.border = thin; x.font = TD
        if c in (3, 4): x.number_format = '#,##0'
        if c == 5: x.number_format = '0.0%'
    rr += 1
for col, wid in zip('ABCDE', [12, 6, 12, 11, 9]): w.column_dimensions[col].width = wid

# ---- 非広告商品 ----
w = wb.create_sheet('非広告商品')
w['A1'] = '広告を出していない商品（広告費0円。利益＝売上−原価）'
w['A1'].font = Font(name='Arial', bold=True, size=13)
rr = 3
noad = [n for n in S if n not in ADV and n != '']
for c, v in enumerate(['商品', '売価', '単価原価', '30日売上', '30日販売数', '30日原価', '30日利益', '30日利益率'], 1):
    w.cell(rr, c, v)
header(w, rr, 8); rr += 1
for n in sorted(noad, key=lambda k: -sum(v[0]-v[1] for v in S[k].values())):
    s = sum(v[0]-v[1] for v in S[n].values()); g = sum(v[0] for v in S[n].values())
    p = PRICE.get(n); q = round(g/p) if p else None
    c_ = q*COST.get(n, 0) if q is not None else 0
    for c, v in enumerate([n, p or '変動', COST.get(n, 0), s, q, c_, s-c_, round((s-c_)/s, 4) if s else ''], 1):
        x = w.cell(rr, c, v); x.border = thin; x.font = TD
        if c in (2, 3, 4, 6, 7): x.number_format = '#,##0'
        if c == 8: x.number_format = '0.0%'
    rr += 1
for col, wid in zip('ABCDEFGH', [26, 9, 10, 12, 11, 11, 12, 11]): w.column_dimensions[col].width = wid

wb.save(OUT)
print('saved', OUT, '/ タブ数', len(wb.sheetnames))
